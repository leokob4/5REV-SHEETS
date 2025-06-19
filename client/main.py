from fastapi import FastAPI, Request, Form, HTTPException, Depends
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.security import OAuth2PasswordBearer
from jose import jwt, JWTError
from pydantic import BaseModel
import openpyxl
import bcrypt
from datetime import datetime, timedelta

import os
import sys

# Define o caminho para a raiz do projeto (assumindo main.py está em client/)
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)

# === Config ===
SECRET_KEY = "plm_secret" # Mantenha esta chave segura e idealmente em variáveis de ambiente
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

# === Init App ===
app = FastAPI()
# A pasta 'templates' deve estar na raiz do projeto
templates = Jinja2Templates(directory=os.path.join(project_root, "templates"))
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/login")

# === In-memory sheet-driven DB ===
users_db = {}
modules_db = {} # REVERTIDO: De tools_db para modules_db
permissions_db = {} # REVERTIDO: De access_db para permissions_db

def load_sheets():
    """
    Carrega os dados das planilhas para o banco de dados em memória do backend.
    Agora usa modules.xlsx e permissions.xlsx, conforme main.xlsx.
    """
    global users_db, modules_db, permissions_db

    # Caminhos relativos à raiz do projeto para consistência
    APP_SHEETS_DIR = os.path.join(project_root, "app_sheets")
    MAIN_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "main.xlsx")

    try:
        # Carrega o main.xlsx para obter as referências de arquivos
        if not os.path.exists(MAIN_EXCEL_PATH):
            print(f"Erro: Arquivo main.xlsx não encontrado em {MAIN_EXCEL_PATH}")
            return # Ou levante uma exceção, dependendo da criticidade

        wb_main = openpyxl.load_workbook(MAIN_EXCEL_PATH)
        if "refs" not in wb_main.sheetnames:
            print(f"Erro: Planilha 'refs' não encontrada em {MAIN_EXCEL_PATH}")
            return

        refs_sheet = wb_main["refs"]
        refs = {}
        for row_idx in range(2, refs_sheet.max_row + 1):
            row_values = [cell.value for cell in refs_sheet[row_idx]]
            if len(row_values) >= 2 and row_values[0] is not None and row_values[1] is not None:
                refs[str(row_values[1])] = str(row_values[0])
            else:
                if not all(v is None for v in row_values): # Only warn for non-empty malformed rows
                    print(f"Aviso: Ignorando linha malformada na planilha 'refs' (linha {row_idx}): {row_values}")


        # Carrega users.xlsx
        users_excel_name = refs.get("users")
        if users_excel_name:
            wb_users = openpyxl.load_workbook(os.path.join(APP_SHEETS_DIR, users_excel_name))
            sheet = wb_users["users"]
            users_db = {}
            # Assume cabeçalhos na primeira linha
            headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            header_map = {h: idx for idx, h in enumerate(headers)}
            
            required_user_headers = ["username", "password_hash", "role"]
            if not all(h in header_map for h in required_user_headers):
                print(f"Aviso: Cabeçalhos esperados ausentes na planilha 'users' de {users_excel_name}. Esperado: {required_user_headers}")

            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]
                if all(v is None for v in row_values): # Ignora linhas completamente vazias
                    continue

                username = row_values[header_map.get("username")] if "username" in header_map and header_map["username"] < len(row_values) else None
                password_hash = row_values[header_map.get("password_hash")] if "password_hash" in header_map and header_map["password_hash"] < len(row_values) else None
                role = row_values[header_map.get("role")] if "role" in header_map and header_map["role"] < len(row_values) else "user" # Define 'user' como padrão se a role for None

                if username and password_hash:
                    users_db[str(username)] = {
                        "username": str(username),
                        "password_hash": str(password_hash),
                        "role": str(role),
                    }
                else:
                    print(f"Aviso: Ignorando linha vazia ou malformada em {users_excel_name} (planilha 'users', linha {row_idx}): {row_values}")
            print(f"Carregados {len(users_db)} usuários de {users_excel_name}.")
        else:
            print("Aviso: 'users' não referenciado em main.xlsx.")

        # Carrega modules.xlsx (REVERTIDO)
        modules_excel_name = refs.get("modules") # Referência a 'modules' em main.xlsx
        if modules_excel_name:
            wb_modules = openpyxl.load_workbook(os.path.join(APP_SHEETS_DIR, modules_excel_name))
            sheet = wb_modules["modules"] # Assume a planilha 'modules'
            modules_db = {}
            # Mapeia cabeçalhos
            headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            header_map = {h: idx for idx, h in enumerate(headers)}
            
            required_module_headers = ["id", "name", "description"] # Assumindo estes cabeçalhos em modules.xlsx
            if not all(h in header_map for h in required_module_headers):
                print(f"Aviso: Cabeçalhos esperados ausentes na planilha 'modules' de {modules_excel_name}. Esperado: {required_module_headers}")
            
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]

                if all(v is None for v in row_values): # Ignora linhas completamente vazias
                    continue

                mod_id = row_values[header_map.get("id")] if "id" in header_map and header_map["id"] < len(row_values) else None
                mod_name = row_values[header_map.get("name")] if "name" in header_map and header_map["name"] < len(row_values) else None
                mod_description = row_values[header_map.get("description")] if "description" in header_map and header_map["description"] < len(row_values) else ""

                if mod_id and mod_name:
                    modules_db[str(mod_id)] = {
                        "id": str(mod_id),
                        "name": str(mod_name),
                        "description": str(mod_description),
                    }
                else:
                    print(f"Aviso: Ignorando linha vazia ou malformada em {modules_excel_name} (planilha 'modules', linha {row_idx}): {row_values}")
            print(f"Carregadas {len(modules_db)} módulos de {modules_excel_name}.")
        else:
            print("Aviso: 'modules' não referenciado em main.xlsx.")

        # Carrega permissions.xlsx (REVERTIDO)
        permissions_excel_name = refs.get("permissions") # Referência a 'permissions' em main.xlsx
        if permissions_excel_name:
            wb_perms = openpyxl.load_workbook(os.path.join(APP_SHEETS_DIR, permissions_excel_name))
            sheet = wb_perms["permissions"] # Assume a planilha 'permissions'
            permissions_db = {}
            # Mapeia cabeçalhos
            headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            header_map = {h: idx for idx, h in enumerate(headers)}

            required_perm_headers = ["role", "allowed_modules"] # Assumindo estes cabeçalhos em permissions.xlsx
            if not all(h in header_map for h in required_perm_headers):
                print(f"Aviso: Cabeçalhos esperados ausentes na planilha 'permissions' de {permissions_excel_name}. Esperado: {required_perm_headers}")

            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]
                
                if all(v is None for v in row_values): # Ignora linhas completamente vazias
                    continue

                role_name = row_values[header_map.get("role")] if "role" in header_map and header_map["role"] < len(row_values) else None
                allowed_modules_str = row_values[header_map.get("allowed_modules")] if "allowed_modules" in header_map and header_map["allowed_modules"] < len(row_values) else ""

                if role_name:
                    # Trata "all" ou lista de IDs separados por vírgula
                    permissions_db[str(role_name)] = [s.strip() for s in str(allowed_modules_str).split(',')] if str(allowed_modules_str).strip().lower() != "all" else "all"
                else:
                    print(f"Aviso: Ignorando linha vazia ou malformada em {permissions_excel_name} (planilha 'permissions', linha {row_idx}): {row_values}")
            print(f"Carregadas {len(permissions_db)} permissões de {permissions_excel_name}.")
        else:
            print("Aviso: 'permissions' não referenciado em main.xlsx.")

    except FileNotFoundError as e:
        print(f"Erro: Um dos arquivos Excel não foi encontrado. Verifique se {MAIN_EXCEL_PATH}, {os.path.join(APP_SHEETS_DIR, users_excel_name or 'users.xlsx')}, {os.path.join(APP_SHEETS_DIR, modules_excel_name or 'modules.xlsx')} e {os.path.join(APP_SHEETS_DIR, permissions_excel_name or 'permissions.xlsx')} existem e estão acessíveis: {e}")
    except KeyError as e:
        print(f"Erro: Planilha ou cabeçalho esperado não encontrado ao carregar. Verifique os nomes das planilhas e cabeçalhos. Detalhes: {e}")
    except Exception as e:
        print(f"Erro inesperado ao carregar planilhas: {e}")

# === Auth ===
def authenticate_user(username, password):
    user = users_db.get(username)
    if not user:
        return None
    # Use bcrypt.checkpw para verificar a senha
    if not bcrypt.checkpw(password.encode('utf-8'), user["password_hash"].encode('utf-8')):
        return None
    return user

def create_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Credenciais inválidas")
        user = users_db.get(username)
        if user is None:
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")

# === Routes ===
@app.on_event("startup")
def startup_event():
    load_sheets()

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = None):
    return templates.TemplateResponse("login.html", {"request": request, "error": error})

@app.post("/login")
async def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    user = authenticate_user(username, password)
    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Nome de usuário ou senha inválidos."})
    token = create_token({"sub": user["username"]})
    response = RedirectResponse("/dashboard", status_code=302)
    response.set_cookie(key="token", value=token, httponly=True)
    return response

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, current_user: dict = Depends(get_current_user)):
    # Usa modules_db e permissions_db
    allowed_modules = permissions_db.get(current_user["role"], [])
    
    if allowed_modules == "all":
        visible_modules = list(modules_db.values())
    else:
        visible_modules = [modules_db[mid] for mid in allowed_modules if mid in modules_db]

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": current_user,
        "modules": visible_modules # Volta para 'modules'
    })

@app.post("/admin/reload")
async def admin_reload(current_user: dict = Depends(get_current_user)):
    """
    Recarrega as planilhas do backend.
    Apenas para usuários com o papel 'admin'.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores podem recarregar os dados.")
    
    load_sheets()
    return RedirectResponse("/dashboard", status_code=302)

