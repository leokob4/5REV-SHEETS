import sys
import os
import bcrypt
import openpyxl
import json
import subprocess
import threading 
import importlib 

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QToolBar, QAction, QTabWidget, QMenu, QToolButton,
    QWidget, QVBoxLayout, QSplitter, QTreeWidget, QTreeWidgetItem,
    QLabel, QLineEdit, QPushButton, QHBoxLayout, QMessageBox, QGraphicsView,
    QGraphicsScene, QGraphicsRectItem, QGraphicsLineItem, QDialog, QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem, QHeaderView, QInputDialog, QComboBox, QGraphicsTextItem,
    QTextEdit 
)
from PyQt5.QtCore import Qt, QPointF, QFileInfo
from PyQt5.QtGui import QBrush, QPen, QColor, QFont 

# --- Correção para ModuleNotFoundError: No module named 'ui' ---
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# --- Importar Módulos das Ferramentas ---
# Agora importamos SearchBarWidget e MiniConsoleWidget
from ui.tools.product_data import ProductDataTool
from ui.tools.bom_manager import BomManagerTool
from ui.tools.configurador import ConfiguradorTool
from ui.tools.colaboradores import ColaboradoresTool
from ui.tools.items import ItemsTool
from ui.tools.manufacturing import ManufacturingTool
from ui.tools.pcp import PcpTool
from ui.tools.estoque import EstoqueTool
from ui.tools.financeiro import FinanceiroTool 
from ui.tools.pedidos import PedidosTool
from ui.tools.manutencao import ManutencaoTool
from ui.tools.engenharia_data import EngenhariaDataTool 
from ui.tools.excel_viewer_tool import ExcelViewerTool 
from ui.tools.structure_view_tool import StructureViewTool
from ui.tools.rpi_tool import RpiTool 
from ui.tools.engenharia_workflow_tool import EngenhariaWorkflowTool 
from ui.tools.user_settings_tool import UserSettingsTool
from app_sheets.tools.tools_line_generator import ToolsLineGeneratorTool 

# NOVAS IMPORTAÇÕES DE WIDGETS MODULARIZADOS
from ui.tools.search_bar import SearchBarWidget 
from ui.tools.mini_console import MiniConsoleWidget

# --- Configuração dos Caminhos dos Arquivos ---
USER_SHEETS_DIR = os.path.join(project_root, "user_sheets")
APP_SHEETS_DIR = os.path.join(project_root, "app_sheets")

COLABORADORES_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "colaboradores.xlsx")
CONFIGURADOR_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "configurador.xlsx")
FINANCEIRO_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "financeiro.xlsx")
MANUTENCAO_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "manutencao.xlsx")
OUTPUT_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "output.xlsx") 
PEDIDOS_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "pedidos.xlsx")
PROGRAMACAO_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "programacao.xlsx") 
RPI_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "RPI.xlsx")
ESTOQUE_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "estoque.xlsx") 
DB_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "db.xlsx") 
ENGENHARIA_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "engenharia.xlsx") 

BOM_DATA_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "bom_data.xlsx") 
ITEMS_DATA_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "items_data.xlsx") 
MANUFACTURING_DATA_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "manufacturing_data.xlsx")

USERS_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "users.xlsx")
ACCESS_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "access.xlsx") 
TOOLS_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "tools.xlsx") 
MAIN_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "main.xlsx") 
MODULES_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "modules.xlsx") 
PERMISSIONS_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "permissions.xlsx") 

UPDATE_METADATA_SCRIPT_PATH = os.path.join(APP_SHEETS_DIR, "tools", "update_user_sheets_metadata.py")
SHEET_VALIDATOR_SCRIPT_PATH = os.path.join(APP_SHEETS_DIR, "tools", "sheet_validator_simple.py") 
CREATE_ENGENHARIA_SCRIPT_PATH = os.path.join(APP_SHEETS_DIR, "tools", "create_engenharia_xlsx.py")
TOOLS_LINE_GENERATOR_SCRIPT_PATH = os.path.join(APP_SHEETS_DIR, "tools", "tools_line_generator.py") 

# Lista de arquivos protegidos (atualizada com os novos módulos)
PROTECTED_FILES = [
    os.path.basename(COLABORADORES_EXCEL_PATH),
    os.path.basename(CONFIGURADOR_EXCEL_PATH),
    os.path.basename(FINANCEIRO_EXCEL_PATH),
    os.path.basename(MANUTENCAO_EXCEL_PATH),
    os.path.basename(OUTPUT_EXCEL_PATH),
    os.path.basename(PEDIDOS_EXCEL_PATH),
    os.path.basename(PROGRAMACAO_EXCEL_PATH),
    os.path.basename(RPI_EXCEL_PATH),
    os.path.basename(ESTOQUE_EXCEL_PATH),
    os.path.basename(DB_EXCEL_PATH), 
    os.path.basename(ENGENHARIA_EXCEL_PATH), 
    os.path.basename(BOM_DATA_EXCEL_PATH), 
    os.path.basename(ITEMS_DATA_EXCEL_PATH), 
    os.path.basename(MANUFACTURING_DATA_EXCEL_PATH),
    
    os.path.basename(USERS_EXCEL_PATH), 
    os.path.basename(ACCESS_EXCEL_PATH), 
    os.path.basename(TOOLS_EXCEL_PATH), 
    os.path.basename(MAIN_EXCEL_PATH), 
    os.path.basename(MODULES_EXCEL_PATH), 
    os.path.basename(PERMISSIONS_EXCEL_PATH), 
    
    os.path.basename(UPDATE_METADATA_SCRIPT_PATH),
    os.path.basename(SHEET_VALIDATOR_SCRIPT_PATH), 
    os.path.basename(CREATE_ENGENHARIA_SCRIPT_PATH),
    os.path.basename(TOOLS_LINE_GENERATOR_SCRIPT_PATH),
    os.path.basename(os.path.join(project_root, 'ui', 'tools', 'search_bar.py')), # NOVO
    os.path.basename(os.path.join(project_root, 'ui', 'tools', 'mini_console.py'))  # NOVO
]

# Garante que os diretórios existam
os.makedirs(USER_SHEETS_DIR, exist_ok=True)
os.makedirs(APP_SHEETS_DIR, exist_ok=True)
os.makedirs(os.path.dirname(UPDATE_METADATA_SCRIPT_PATH), exist_ok=True) 
os.makedirs(os.path.dirname(CREATE_ENGENHARIA_SCRIPT_PATH), exist_ok=True) 
os.makedirs(os.path.dirname(SHEET_VALIDATOR_SCRIPT_PATH), exist_ok=True) 
os.makedirs(os.path.dirname(TOOLS_LINE_GENERATOR_SCRIPT_PATH), exist_ok=True) 
os.makedirs(os.path.join(project_root, 'ui', 'tools'), exist_ok=True) # Garante que ui/tools existe


# === FUNÇÕES AUXILIARES DE PLANILHA ===
def load_users_from_excel_util():
    """Carrega dados de usuário do arquivo users.xlsx (agora em app_sheets)."""
    users = {}
    try:
        if not os.path.exists(USERS_EXCEL_PATH):
            QMessageBox.critical(None, "Arquivo Não Encontrado", f"O arquivo de usuários não foi encontrado: {USERS_EXCEL_PATH}")
            return {}

        wb = openpyxl.load_workbook(USERS_EXCEL_PATH)
        users_sheet = wb["users"] 
        
        headers = [cell.value for cell in users_sheet[1]] if users_sheet.max_row >= 1 else []
        header_map = {h: idx for idx, h in enumerate(headers)}

        required_headers = ["id", "username", "password_hash", "role"]
        if not all(h in header_map for h in required_headers):
            QMessageBox.warning(None, "Cabeçalhos Ausentes", 
                                f"A planilha 'users' em {USERS_EXCEL_PATH} não possui todos os cabeçalhos esperados. "
                                f"Esperado: {', '.join(required_headers)}")
            return {}

        for row_idx in range(2, users_sheet.max_row + 1):
            row_values = [cell.value for cell in users_sheet[row_idx]]
            
            if all(v is None for v in row_values):
                continue

            user_id = row_values[header_map["id"]] if "id" in header_map and header_map["id"] < len(row_values) else None
            username = row_values[header_map["username"]] if "username" in header_map and header_map["username"] < len(row_values) else None
            password_hash = row_values[header_map["password_hash"]] if "password_hash" in header_map and header_map["password_hash"] < len(row_values) else None
            role = row_values[header_map["role"]] if "role" in header_map and header_map["role"] < len(row_values) else None

            if username is not None and password_hash is not None:
                users[str(username)] = { 
                    "id": user_id,
                    "username": str(username),
                    "password_hash": str(password_hash),
                    "role": str(role) if role is not None else "user"
                }
            else:
                print(f"Aviso: Ignorando linha malformada na planilha 'users' (linha {row_idx}): {row_values}")

    except KeyError:
        QMessageBox.critical(None, "Erro de Planilha", f"A planilha 'users' não foi encontrada em {USERS_EXCEL_PATH}. Por favor, certifique-se de que o nome da planilha seja 'users'.")
        return {}
    except Exception as e:
        QMessageBox.critical(None, "Erro de Carregamento", f"Erro ao carregar usuários: {e}")
        return {}
    return users


def register_user(username, password, role="user"):
    """Registra um novo usuário no arquivo users.xlsx (agora em app_sheets)."""
    try:
        wb = openpyxl.load_workbook(USERS_EXCEL_PATH)
        sheet = wb["users"]
        
        headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
        header_map = {h: idx for idx, h in enumerate(headers)}

        required_headers = ["id", "username", "password_hash", "role"]
        if not all(h in header_map for h in required_headers):
            for h in required_headers:
                if h not in header_map:
                    headers.append(h)
            sheet.insert_rows(1) 
            for col_idx, h in enumerate(headers):
                sheet.cell(row=1, column=col_idx + 1).value = h
            header_map = {h: idx for idx, h in enumerate(headers)} 

        next_id = 1
        existing_ids = set()
        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            current_id = row_values[header_map["id"]] if "id" in header_map and header_map["id"] < len(row_values) else None
            current_username = row_values[header_map["username"]] if "username" in header_map and header_map["username"] < len(row_values) else None

            if all(v is None for v in row_values):
                continue

            if current_id is not None:
                existing_ids.add(current_id)
            if current_username == username:
                raise ValueError("Nome de usuário já existe.")
        while next_id in existing_ids:
            next_id += 1

        password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        
        new_row_data = [""] * len(headers)
        new_row_data[header_map["id"]] = next_id
        new_row_data[header_map["username"]] = username
        new_row_data[header_map["password_hash"]] = password_hash
        new_row_data[header_map["role"]] = role
        
        sheet.append(new_row_data)
        wb.save(USERS_EXCEL_PATH)
    except FileNotFoundError:
        QMessageBox.critical(None, "Arquivo Não Encontrado", f"O arquivo de usuários não foi encontrado em: {USERS_EXCEL_PATH}. Não é possível registrar o usuário.")
    except KeyError:
        QMessageBox.critical(None, "Erro de Planilha", f"A planilha 'users' não foi encontrada em {USERS_EXCEL_PATH}. Não é possível registrar o usuário.")
    except Exception as e:
        QMessageBox.critical(None, "Erro", f"Ocorreu um erro durante o registro: {e}")


def load_tools_from_excel_util():
    """
    Carrega dados da ferramenta do arquivo Excel dedicado às ferramentas (tools.xlsx).
    Adaptado para os novos cabeçalhos: mod_id, mod_name, module_path, class_name etc.
    """
    tools = {}
    try:
        if not os.path.exists(TOOLS_EXCEL_PATH):
            QMessageBox.critical(None, "Arquivo Não Encontrado", f"O arquivo de ferramentas não foi encontrado: {TOOLS_EXCEL_PATH}. Por favor, certifique-se de que ele exista.")
            return {}

        wb = openpyxl.load_workbook(TOOLS_EXCEL_PATH)
        sheet = wb["tools"] 
        
        headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
        header_map = {h: idx for idx, h in enumerate(headers)}

        required_headers = ["mod_id", "mod_name", "module_path", "class_name"] 
        if not all(h in header_map for h in required_headers):
            QMessageBox.warning(None, "Cabeçalhos Ausentes", 
                                f"A planilha 'tools' em {TOOLS_EXCEL_PATH} não possui todos os cabeçalhos esperados. "
                                f"Esperado: {', '.join(required_headers)}")
            return {}

        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            if all(v is None for v in row_values):
                continue

            mod_id = row_values[header_map["mod_id"]] if "mod_id" in header_map and header_map["mod_id"] < len(row_values) and row_values[header_map["mod_id"]] is not None else None
            mod_name = row_values[header_map["mod_name"]] if "mod_name" in header_map and header_map["mod_name"] < len(row_values) and row_values[header_map["mod_name"]] is not None else None
            mod_description = row_values[header_map.get("mod_description")] if "mod_description" in header_map and header_map["mod_description"] < len(row_values) else None
            module_path = row_values[header_map["module_path"]] if "module_path" in header_map and header_map["module_path"] < len(row_values) and row_values[header_map["module_path"]] is not None else None
            class_name = row_values[header_map["class_name"]] if "class_name" in header_map and header_map["class_name"] < len(row_values) and row_values[header_map["class_name"]] is not None else None 
            mod_work_table = row_values[header_map.get("MOD_WORK_TABLE")] if "MOD_WORK_TABLE" in header_map and header_map["MOD_WORK_TABLE"] < len(row_values) else None
            mod_work_table_path = row_values[header_map.get("MOD_WORK_TABLE_PATH")] if "MOD_WORK_TABLE_PATH" in header_map and header_map["MOD_WORK_TABLE_PATH"] < len(row_values) else None

            if mod_id is not None and mod_name is not None and (module_path is not None or class_name is not None): 
                tools[str(mod_id)] = {
                    "id": str(mod_id),
                    "name": str(mod_name),
                    "description": str(mod_description) if mod_description is not None else "",
                    "path": str(module_path) if module_path is not None else "", 
                    "class_name": str(class_name) if class_name is not None else "", 
                    "mod_work_table": str(mod_work_table) if mod_work_table is not None else "",
                    "mod_work_table_path": str(mod_work_table_path) if mod_work_table_path is not None else ""
                }
            else:
                print(f"Aviso: Ignorando linha malformada ou incompleta na planilha 'tools' (linha {row_idx}): {row_values}")

    except KeyError as ke:
        QMessageBox.critical(None, "Erro de Planilha", f"A planilha 'tools' não foi encontrada em {TOOLS_EXCEL_PATH} ou cabeçalho ausente ({ke}). Por favor, certifique-se de que o nome da planilha seja 'tools' e os cabeçalhos estejam corretos.")
        return {}
    except Exception as e:
        QMessageBox.critical(None, "Erro de Carregamento", f"Erro ao carregar ferramentas: {e}")
        return {}
    return tools


def load_role_permissions_util():
    """Carrega permissões de papel do arquivo access.xlsx (agora em app_sheets)."""
    perms = {}
    try:
        if not os.path.exists(ACCESS_EXCEL_PATH):
            QMessageBox.critical(None, "Arquivo Não Encontrado", f"O arquivo de permissões não foi encontrado: {ACCESS_EXCEL_PATH}")
            return {}

        wb = openpyxl.load_workbook(ACCESS_EXCEL_PATH)
        sheet = wb["access"] 
        
        headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
        header_map = {h: idx for idx, h in enumerate(headers)}

        required_headers = ["role", "allowed_tools"]
        if not all(h in header_map for h in required_headers):
            QMessageBox.warning(None, "Cabeçalhos Ausentes", 
                                f"A planilha 'access' em {ACCESS_EXCEL_PATH} não possui todos os cabeçalhos esperados. "
                                f"Esperado: {', '.join(required_headers)}")
            return {}

        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            if all(v is None for v in row_values):
                continue

            role_name = row_values[header_map["role"]] if "role" in header_map and header_map["role"] < len(row_values) and row_values[header_map["role"]] is not None else None
            allowed_tools_str = row_values[header_map["allowed_tools"]] if "allowed_tools" in header_map and header_map["allowed_tools"] < len(row_values) and row_values[header_map["allowed_tools"]] is not None else ""

            if role_name:
                perms[str(role_name)] = [s.strip() for s in allowed_tools_str.split(',')] if allowed_tools_str.strip().lower() != "all" else "all"
            else:
                print(f"Aviso: Ignorando linha malformada na planilha 'access' (linha {row_idx}): {row_values}")
        return perms
    except KeyError:
        QMessageBox.critical(None, "Erro de Planilha", f"A planilha 'access' não foi encontrada em {ACCESS_EXCEL_PATH}. Por favor, certifique-se de que o nome da planilha seja 'access'.")
        return {}
    except Exception as e:
        QMessageBox.critical(None, "Erro de Carregamento", f"Erro ao carregar permissões: {e}")
        return {}
    return perms


def load_workspace_items_from_excel_util():
    """
    Carrega os itens do espaço de trabalho da planilha 'Estrutura' em engenharia.xlsx.
    Mapeia 'part_number' para 'name' e 'part_type' para 'type'.
    """
    workspace_items = []
    try:
        if not os.path.exists(ENGENHARIA_EXCEL_PATH): 
            QMessageBox.warning(None, "Arquivo Não Encontrado", 
                                f"O arquivo 'engenharia.xlsx' não foi encontrado: {ENGENHARIA_EXCEL_PATH}\n"
                                "Por favor, crie-o com uma planilha 'Estrutura' e as colunas 'part_number' e 'part_type'.")
            return []

        wb = openpyxl.load_workbook(ENGENHARIA_EXCEL_PATH) 
        sheet_name = "Estrutura" 
        if sheet_name not in wb.sheetnames:
            QMessageBox.warning(None, "Planilha Ausente", 
                                f"A planilha '{sheet_name}' não foi encontrada em {ENGENHARIA_EXCEL_PATH}. "
                                f"Por favor, certifique-se de que a planilha exista e tenha as colunas 'part_number' e 'part_type'.")
            return []

        sheet = wb[sheet_name]
        
        headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
        header_map = {h: idx for idx, h in enumerate(headers)}

        required_headers = ["part_number", "part_type"] 
        if not all(h in header_map for h in required_headers):
            QMessageBox.warning(None, "Cabeçalhos Ausentes", 
                                f"A planilha '{sheet_name}' em {ENGENHARIA_EXCEL_PATH} não possui todos os cabeçalhos esperados para o workspace. "
                                f"Esperado: {', '.join(required_headers)}")
            return []

        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            if all(v is None for v in row_values):
                continue 

            part_number = row_values[header_map["part_number"]] if "part_number" in header_map and header_map["part_number"] < len(row_values) else None
            part_type = row_values[header_map["part_type"]] if "part_type" in header_map and header_map["part_type"] < len(row_values) else None

            if part_number is not None and part_type is not None:
                workspace_items.append({"name": str(part_number), "type": str(part_type)})
            else:
                print(f"Aviso: Ignorando linha malformada na planilha '{sheet_name}' (linha {row_idx}): {row_values}")

    except Exception as e:
        QMessageBox.critical(None, "Erro de Carregamento", f"Erro ao carregar itens do espaço de trabalho: {e}")
        return []
    return workspace_items


# === JANELA DE LOGIN ===
class LoginWindow(QWidget):
    """
    A janela de login para o aplicativo.
    Lida com a autenticação e registro de usuários.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("5revolution Login")
        self.setGeometry(400, 200, 300, 180) 
        self.users = load_users_from_excel_util() 

        self._init_ui()

    def _init_ui(self):
        """Inicializa os elementos da interface do usuário para a janela de login."""
        layout = QVBoxLayout()

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Nome de Usuário")
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Senha")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.returnPressed.connect(self.authenticate) 

        login_btn = QPushButton("Entrar")
        login_btn.clicked.connect(self.authenticate)

        register_btn = QPushButton("Registrar")
        register_btn.clicked.connect(self.handle_register)

        layout.addWidget(QLabel("Bem-vindo ao 5revolution"))
        layout.addWidget(self.username_input)
        layout.addWidget(self.password_input)

        btns_layout = QHBoxLayout()
        btns_layout.addWidget(login_btn)
        btns_layout.addWidget(register_btn)

        layout.addLayout(btns_layout)
        self.setLayout(layout)

    def authenticate(self):
        """Autentica o usuário com base nas credenciais fornecidas."""
        uname = self.username_input.text().strip()
        pwd = self.password_input.text().strip()

        if not uname or not pwd:
            QMessageBox.warning(self, "Falha no Login", "Nome de usuário e senha não podem estar vazios.")
            return

        user = self.users.get(uname)

        if not user or not bcrypt.checkpw(pwd.encode(), user["password_hash"].encode()):
            QMessageBox.warning(self, "Falha no Login", "Nome de usuário ou senha inválidos.")
            return

        self.main = TeamcenterStyleGUI(user)
        self.main.showMaximized() 
        self.close() 

    def handle_register(self):
        """Lida com o registro de usuário."""
        uname = self.username_input.text().strip()
        pwd = self.password_input.text().strip()

        if not uname or not pwd:
            QMessageBox.warning(self, "Erro de Validação", "Nome de usuário e senha são obrigatórios para o registro.")
            return

        try:
            register_user(uname, pwd)
            QMessageBox.information(self, "Registrado", f"Usuário '{uname}' registrado com sucesso com o papel 'user'.")
            self.users = load_users_from_excel_util() 
            self.username_input.clear()
            self.password_input.clear()
        except ValueError as ve:
            QMessageBox.warning(self, "Falha no Registro", str(ve))
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro durante o registro: {e}")


# === FERRAMENTA GUI PARA VALIDADOR DE CABEÇALHOS DO DB ===
# Esta classe é definida AQUI dentro de gui.py e não é importada de um arquivo externo
class DbHeadersUpdaterTool(QWidget):
    """
    Ferramenta GUI para executar e exibir o resultado do validador de planilhas.
    Permite que o usuário visualize a saída do script 'sheet_validator_simple.py'
    diretamente na interface.
    """
    def __init__(self, refresh_callback=None): 
        super().__init__()
        self.setWindowTitle("Validador de Consistência de Planilhas")
        self.script_path = SHEET_VALIDATOR_SCRIPT_PATH # Caminho para o script externo do validador
        self.refresh_callback = refresh_callback 
        self._init_ui()

    def _init_ui(self):
        """Inicializa os elementos da interface do usuário."""
        layout = QVBoxLayout()

        description_label = QLabel(
            "Esta ferramenta executa o script 'sheet_validator_simple.py', que compara a estrutura real dos "
            "cabeçalhos de todas as planilhas do sistema com o esquema registrado na 'db_db' em 'db.xlsx'.\n\n"
            "Resultados esperados:\n"
            "- Sucesso: Mensagem indicando que não foram encontradas inconsistências.\n"
            "- Erro: Detalhes de quais planilhas e cabeçalhos estão inconsistentes com a 'db_db'."
        )
        description_label.setWordWrap(True) 
        layout.addWidget(description_label)

        self.run_button = QPushButton("Executar Validação")
        self.run_button.clicked.connect(self._run_validation_script)
        layout.addWidget(self.run_button)

        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setPlaceholderText("Aguardando execução da validação... Clique em 'Executar Validação' para iniciar. A saída detalhada do script aparecerá aqui.")
        layout.addWidget(self.output_text)

        self.setLayout(layout)

    def _run_validation_script(self):
        """
        Executa o script de validação externa e exibe a saída no QTextEdit.
        """
        self.output_text.clear()
        self.output_text.append("Executando validação... Por favor, aguarde.")
        self.run_button.setEnabled(False) 

        cmd = [sys.executable, self.script_path, "run"] 

        try:
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, bufsize=1)
            
            def read_stream(stream, is_stderr=False):
                for line in stream:
                    self.output_text.append(f"ERRO: {line.strip()}" if is_stderr else line.strip())
                    QApplication.processEvents() 

            stdout_thread = threading.Thread(target=read_stream, args=(process.stdout,))
            stderr_thread = threading.Thread(target=read_stream, args=(process.stderr, True))

            stdout_thread.start()
            stderr_thread.start()

            stdout_thread.join() 
            stderr_thread.join()

            process.wait() 

            if process.returncode == 0:
                self.output_text.append("\n--- Validação Concluída com Sucesso ---")
                if self.refresh_callback: 
                    self.refresh_callback()
            else:
                self.output_text.append(f"\n--- Validação Concluída com Erros (Código: {process.returncode}) ---")
                self.output_text.append("\nPor favor, revise a saída acima para identificar as inconsistências.")
                
        except FileNotFoundError:
            self.output_text.append(f"ERRO: O executável Python ou o script '{os.path.basename(self.script_path)}' não foi encontrado. Verifique o PATH ou o caminho do script.")
        except Exception as e:
            self.output_text.append(f"ERRO Inesperado ao executar script: {e}")
        finally:
            self.run_button.setEnabled(True) 


# === CLASSE PRINCIPAL DA GUI ===
class TeamcenterStyleGUI(QMainWindow):
    def __init__(self, user_data):
        super().__init__()
        self.user_data = user_data
        self.current_user_role = user_data["role"]

        self.users = {}
        self.access_permissions = {}
        self.available_tools_metadata = {}
        self.workspace_items = []

        self._load_all_configuration_data() 

        self.setWindowTitle("5revolution ERP")
        self.setGeometry(100, 100, 1200, 800) 
        self._init_ui()

    def _load_all_configuration_data(self):
        """Carrega todos os dados de configuração dos arquivos Excel."""
        self.users = load_users_from_excel_util()
        self.access_permissions = load_role_permissions_util()
        self.available_tools_metadata = load_tools_from_excel_util()
        self.workspace_items = load_workspace_items_from_excel_util()
        print("Dados de configuração carregados/recarregados.")

    def _refresh_gui_data(self):
        """Recarrega os dados de configuração e atualiza os componentes da GUI."""
        self._load_all_configuration_data()
        tools_menu_btn = self.findChild(QToolButton, "tools_menu_btn")
        if tools_menu_btn:
            self._populate_tools_menu(tools_menu_btn.menu()) 
        self._populate_workspace_tree() 
        self._populate_file_system_tree() 
        # Garante que a barra de pesquisa e o console também sejam reinicializados se necessário
        self.search_bar_widget.clear_search() # Limpa a busca
        self.mini_console_widget.clear_output() # Limpa o console
        self.mini_console_widget.append_output("GUI recarregada. Console limpo.")
        print("GUI atualizada com novos dados de configuração.")


    def _init_ui(self):
        """Inicializa os componentes principais da interface do usuário."""
        self._create_toolbar_menu()
        self._create_main_layout()
        self._setup_initial_content()

    def _create_toolbar_menu(self):
        """Cria a barra de ferramentas superior e seus menus."""
        toolbar = self.addToolBar("Main Toolbar")
        toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)

        # Menu Arquivo
        file_menu_btn = QToolButton(self)
        file_menu_btn.setText("Arquivo")
        file_menu_btn.setPopupMode(QToolButton.InstantPopup)
        file_menu = QMenu(self)
        exit_action = QAction("Sair", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        file_menu_btn.setMenu(file_menu)
        toolbar.addWidget(file_menu_btn)

        # Menu Ferramentas (dinâmico com base nas permissões)
        tools_menu_btn = QToolButton(self)
        tools_menu_btn.setObjectName("tools_menu_btn") 
        tools_menu_btn.setText("Ferramentas")
        tools_menu_btn.setPopupMode(QToolButton.InstantPopup)
        tools_menu = QMenu(self)
        self._populate_tools_menu(tools_menu)
        tools_menu_btn.setMenu(tools_menu)
        toolbar.addWidget(tools_menu_btn)

        # Menu Ajuda
        help_menu_btn = QToolButton(self)
        help_menu_btn.setText("Ajuda")
        help_menu_btn.setPopupMode(QToolButton.InstantPopup)
        help_menu = QMenu(self)
        about_action = QAction("Sobre", self)
        about_action.triggered.connect(lambda: QMessageBox.information(self, "Sobre 5revolution", "Sistema ERP 5revolution v1.0"))
        help_menu.addAction(about_action)
        help_menu_btn.setMenu(help_menu)
        toolbar.addWidget(help_menu_btn)

        # Menu Admin (apenas para administradores)
        if self.current_user_role == "admin":
            admin_menu_btn = QToolButton(self)
            admin_menu_btn.setText("👑 Ferramentas Admin")
            admin_menu_btn.setPopupMode(QToolButton.InstantPopup)
            admin_menu = QMenu(self)
            
            create_update_sheets_action = QAction("Criar/Reinicializar/atualizar planilhas", self)
            create_update_sheets_action.setToolTip(
                "Executa o script 'update_user_sheets_metadata.py' com a ação 'create_or_update_sheets'.\n"
                "Função: Cria novas planilhas ou atualiza as existentes com os cabeçalhos definidos na 'db_db' em 'db.xlsx'.\n"
                "Importante: Preserva os dados existentes a partir da segunda linha (dados do app), apenas ajustando a estrutura de cabeçalhos e excluindo linhas totalmente vazias abaixo da 2ª linha."
            )
            create_update_sheets_action.triggered.connect(self._run_create_or_update_all_sheets) 
            admin_menu.addAction(create_update_sheets_action)

            sync_db_schema_action = QAction("Sincronizar 'db_db' com Estrutura Real das Planilhas", self)
            sync_db_schema_action.setToolTip(
                "Executa o script 'update_user_sheets_metadata.py' com a ação 'update_db_schema'.\n"
                "Função: Coleta os cabeçalhos de *todas* as planilhas Excel do projeto (em 'user_sheets' e 'app_sheets', excluindo o próprio 'db.xlsx').\n"
                "Resultado: Registra essa estrutura real na planilha 'db_db' em 'db.xlsx', servindo como a 'fonte da verdade' para validações futuras."
            )
            sync_db_schema_action.triggered.connect(self._run_sync_db_db_schema) 
            admin_menu.addAction(sync_db_schema_action)

            validate_db_consistency_action = QAction("Abrir Validador de Consistência de Planilhas (GUI)", self)
            validate_db_consistency_action.setToolTip(
                "Abre uma interface gráfica dedicada para executar o 'sheet_validator_simple.py'.\n"
                "Função: Compara os cabeçalhos de todas as planilhas do projeto com o esquema registrado na 'db_db' em 'db.xlsx'.\n"
                "Saída: Mostra detalhadamente na interface quais planilhas e cabeçalhos apresentam inconsistências, ou uma mensagem de sucesso."
            )
            validate_db_consistency_action.triggered.connect(lambda: self._open_tool("MOD000018", refresh_callback=self._refresh_gui_data)) 
            admin_menu.addAction(validate_db_consistency_action)

            generate_tool_line_action = QAction("Adicionar Nova Ferramenta ao Sistema", self)
            generate_tool_line_action.setToolTip(
                "Abre uma interface gráfica para adicionar novas entradas de módulo/ferramenta na planilha 'tools.xlsx'.\n"
                "Função: Facilita o registro de novas ferramentas no sistema, gerando automaticamente um 'mod_id' sequencial e coletando informações essenciais."
            )
            generate_tool_line_action.triggered.connect(lambda: self._open_tool("MOD000019", refresh_callback=self._refresh_gui_data)) 
            admin_menu.addAction(generate_tool_line_action)


            admin_menu_btn.setMenu(admin_menu)
            toolbar.addWidget(admin_menu_btn)


    def _populate_tools_menu(self, menu):
        """
        Popula o menu de ferramentas com base nas permissões do usuário e nos metadados das ferramentas.
        Utiliza 'allowed_tools' do access.xlsx diretamente.
        """
        menu.clear() 
        user_allowed_tools_list = self.access_permissions.get(self.current_user_role, [])
        
        allowed_tool_ids_for_user = set()
        if user_allowed_tools_list == "all":
            allowed_tool_ids_for_user = set(self.available_tools_metadata.keys())
        else:
            allowed_tool_ids_for_user = set(user_allowed_tools_list)
            
        for tool_id, tool_info in self.available_tools_metadata.items():
            if tool_id in ["MOD000019", "MOD000018"]: 
                continue 
            
            if tool_id in allowed_tool_ids_for_user:
                action = QAction(tool_info["name"], self)
                action.triggered.connect(lambda checked, t_id=tool_id: self._open_tool(t_id, refresh_callback=self._refresh_gui_data))
                menu.addAction(action)

    def _open_tool(self, tool_id, refresh_callback=None):
        """
        Abre a ferramenta selecionada em uma nova aba, carregando a classe dinamicamente de tools.xlsx.
        """
        tool_info = self.available_tools_metadata.get(tool_id)
        
        if not tool_info:
            QMessageBox.warning(self, "Ferramenta Não Encontrada", f"Metadados para a ferramenta com ID '{tool_id}' não encontrados em tools.xlsx.")
            return

        tool_name = tool_info["name"]
        module_path = tool_info["path"] 
        class_name = tool_info["class_name"] 

        ToolClass = None # Inicializa ToolClass
        # Casos especiais que não dependem diretamente de module_path + class_name de tools.xlsx
        if tool_id == "MOD000019": # ToolsLineGeneratorTool (app_sheets.tools.tools_line_generator)
            ToolClass = ToolsLineGeneratorTool
        elif tool_id == "MOD000018": # DbHeadersUpdaterTool (classe interna do gui.py)
            ToolClass = DbHeadersUpdaterTool
        else:
            # Importação dinâmica da classe da ferramenta
            try:
                # Adiciona o diretório base ao sys.path temporariamente se não for um módulo de nível superior
                if module_path.startswith("ui.tools"):
                    base_module_dir = os.path.join(project_root, "ui")
                elif module_path.startswith("app_sheets.tools"):
                    base_module_dir = os.path.join(project_root, "app_sheets")
                else:
                    base_module_dir = project_root # Default

                if base_module_dir not in sys.path:
                    sys.path.insert(0, base_module_dir) # Garante que o diretório base do módulo esteja no sys.path
                    
                module = importlib.import_module(module_path)
                ToolClass = getattr(module, class_name)
            except ImportError as e:
                QMessageBox.critical(self, "Erro de Importação", 
                                     f"Não foi possível importar o módulo '{module_path}' para a ferramenta '{tool_name}'. Verifique 'module_path' em tools.xlsx e a existência do arquivo. Erro: {e}")
                print(f"Erro: Não foi possível importar o módulo '{module_path}'. Verifique o module_path em tools.xlsx e a existência do arquivo. Erro: {e}")
                return
            except AttributeError as e:
                QMessageBox.critical(self, "Erro de Classe", 
                                     f"A classe '{class_name}' não foi encontrada no módulo '{module_path}' para a ferramenta '{tool_name}'. Verifique 'class_name' em tools.xlsx. Erro: {e}")
                print(f"Erro: A classe '{class_name}' não foi encontrada no módulo '{module_path}'. Verifique o class_name em tools.xlsx. Erro: {e}")
                return
            except Exception as e:
                QMessageBox.critical(self, "Erro Inesperado", f"Ocorreu um erro inesperado ao carregar a ferramenta '{tool_name}': {e}")
                print(f"Erro inesperado ao carregar ferramenta: {e}")
                import traceback
                traceback.print_exc()
                return

        # Tenta instanciar a ferramenta com os parâmetros corretos
        tool_instance = None
        try:
            if tool_id in ["MOD000019", "MOD000018"]: 
                tool_instance = ToolClass(refresh_callback=refresh_callback)
            elif tool_id == "MOD000014" or tool_id == "mod_user_settings": 
                tool_instance = ToolClass(self.user_data) 
            elif tool_id == "MOD000012" or tool_id == "mod4": 
                tool_instance = ToolClass(file_path=ENGENHARIA_EXCEL_PATH, sheet_name="Estrutura")
            elif tool_id == "MOD000013" or tool_id == "mod_workflow": 
                tool_instance = ToolClass(file_path=ENGENHARIA_EXCEL_PATH, sheet_name="Workflows")
            elif tool_id == "MOD000015" or tool_id == "mod_excel_viewer": 
                 tool_instance = ToolClass(file_path=None) 
            else:
                work_table_path = tool_info.get("mod_work_table_path")
                if work_table_path:
                    full_work_table_path = os.path.normpath(os.path.join(project_root, work_table_path.strip('/\\')))
                    if os.path.exists(full_work_table_path):
                        tool_instance = ToolClass(file_path=full_work_table_path)
                    else:
                        QMessageBox.warning(self, "Caminho Inválido", f"O arquivo de trabalho para '{tool_name}' não foi encontrado: {full_work_table_path}")
                        return
                else:
                    tool_instance = ToolClass()

            if tool_instance:
                for i in range(self.central_widget.count()):
                    if self.central_widget.tabText(i) == tool_name:
                        self.central_widget.setCurrentIndex(i)
                        return

                self.central_widget.addTab(tool_instance, tool_name)
                self.central_widget.setCurrentWidget(tool_instance)
            else:
                QMessageBox.warning(self, "Erro de Instanciação", f"Não foi possível criar uma instância para a ferramenta '{tool_name}'. Verifique o construtor da classe ou os parâmetros necessários.")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Abrir Ferramenta", f"Não foi possível abrir a ferramenta '{tool_name}': {e}")
            print(f"Detalhes do erro ao abrir a ferramenta '{tool_name}': {e}")
            import traceback
            traceback.print_exc()


    def _create_main_layout(self):
        """Cria o layout principal com a árvore de navegação e as abas de trabalho."""
        main_splitter = QSplitter(Qt.Horizontal)
        self.setCentralWidget(main_splitter)

        # Painel Esquerdo (Árvore de Navegação e Console)
        left_panel_container = QWidget()
        left_panel_layout = QVBoxLayout(left_panel_container)
        left_panel_layout.setContentsMargins(0,0,0,0) # Remove margens do layout

        # Splitter para a Árvore de Navegação e o Console
        left_splitter = QSplitter(Qt.Vertical)

        # Widget para a Árvore de Navegação
        tree_view_widget = QWidget()
        tree_view_layout = QVBoxLayout(tree_view_widget)
        tree_view_layout.setContentsMargins(0,0,0,0)
        tree_view_layout.addWidget(QLabel("<h2>Espaço de Trabalho</h2>"))

        # Instancia e adiciona a barra de pesquisa
        self.tree_widget = QTreeWidget() # Instancia aqui para passar para o SearchBarWidget
        self.tree_widget.setHeaderLabels(["Nome", "Tipo"]) 
        self.tree_widget.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        
        self.search_bar_widget = SearchBarWidget(self.tree_widget)
        tree_view_layout.addWidget(self.search_bar_widget)
        tree_view_layout.addWidget(self.tree_widget)
        
        left_splitter.addWidget(tree_view_widget)

        # Instancia e adiciona o mini-console
        self.mini_console_widget = MiniConsoleWidget()
        # Opcional: conectar o sinal de comando para uma função no GUI principal
        self.mini_console_widget.command_entered.connect(self._handle_console_command)
        left_splitter.addWidget(self.mini_console_widget)

        # Define tamanhos iniciais para o splitter esquerdo (árvore e console)
        # Ex: 70% para a árvore, 30% para o console
        left_splitter.setSizes([self.height() * 0.7, self.height() * 0.3])

        left_panel_layout.addWidget(left_splitter)
        main_splitter.addWidget(left_panel_container)


        # Painel Direito (Abas de Trabalho)
        self.central_widget = QTabWidget()
        self.central_widget.setTabsClosable(True)
        self.central_widget.tabCloseRequested.connect(self.central_widget.removeTab)
        main_splitter.addWidget(self.central_widget)

        # Define o tamanho inicial dos painéis principais
        main_splitter.setSizes([300, 900]) 

    def _setup_initial_content(self):
        """Popula a árvore de navegação e pode abrir abas iniciais."""
        self._populate_workspace_tree()
        self._populate_file_system_tree()

    def _populate_workspace_tree(self):
        """Popula a seção 'Espaço de Trabalho' da árvore lendo de engenharia.xlsx."""
        workspace_root_item = None
        for i in range(self.tree_widget.topLevelItemCount()):
            item = self.tree_widget.topLevelItem(i)
            if item.text(0) == "Projetos/Espaço de Trabalho":
                workspace_root_item = item
                break
        
        if workspace_root_item:
            while workspace_root_item.childCount() > 0:
                workspace_root_item.removeChild(workspace_root_item.child(0))
        else:
            workspace_root_item = QTreeWidgetItem(self.tree_widget, ["Projetos/Espaço de Trabalho", "Pasta"])
            self.tree_widget.addTopLevelItem(workspace_root_item) 
        
        workspace_root_item.setExpanded(True) 
        
        for item_data in self.workspace_items: 
            new_item = QTreeWidgetItem(workspace_root_item, [item_data["name"], item_data["type"]])
            new_item.setHidden(False) 
            
        self._sort_top_level_items()


    def _populate_file_system_tree(self):
        """Popula as seções 'Arquivos do Usuário' e 'Arquivos do Sistema' da árvore."""
        root_items_to_remove = []
        for i in range(self.tree_widget.topLevelItemCount()):
            item = self.tree_widget.topLevelItem(i)
            if item.text(0) in ["Arquivos do Usuário (user_sheets)", "Arquivos do Sistema (app_sheets)"]:
                root_items_to_remove.append(item)
        
        for item in root_items_to_remove:
            self.tree_widget.takeTopLevelItem(self.tree_widget.indexOfTopLevelItem(item))

        user_files_root = QTreeWidgetItem(self.tree_widget, ["Arquivos do Usuário (user_sheets)", "Pasta"])
        user_files_root.setExpanded(True)
        app_files_root = QTreeWidgetItem(self.tree_widget, ["Arquivos do Sistema (app_sheets)", "Pasta"])
        app_files_root.setExpanded(True)

        self._add_files_to_tree(USER_SHEETS_DIR, user_files_root)
        self._add_files_to_tree(APP_SHEETS_DIR, app_files_root)
        
        self._sort_top_level_items()

    def _sort_top_level_items(self):
        """Garante que os itens de nível superior da árvore estejam em uma ordem consistente."""
        top_level_items = []
        for i in range(self.tree_widget.topLevelItemCount()):
            top_level_items.append(self.tree_widget.topLevelItem(i))
        
        order = {
            "Projetos/Espaço de Trabalho": 0,
            "Arquivos do Usuário (user_sheets)": 1,
            "Arquivos do Sistema (app_sheets)": 2
        }
        
        top_level_items.sort(key=lambda item: order.get(item.text(0), 99))
        
        self.tree_widget.clear() 
        for item in top_level_items:
            self.tree_widget.addTopLevelItem(item)

    def _add_files_to_tree(self, directory, parent_item):
        """Adiciona arquivos .xlsx e subdiretórios de um diretório à árvore."""
        try:
            for filename in os.listdir(directory):
                if filename.startswith('~$'): 
                    continue

                if os.path.basename(directory) == "user_sheets" and filename.lower() == "db.xlsx":
                    continue
                
                file_path = os.path.join(directory, filename)
                if os.path.isdir(file_path):
                    folder_item = QTreeWidgetItem(parent_item, [filename, "Pasta"])
                    folder_item.setExpanded(True)
                    self._add_files_to_tree(file_path, folder_item) 
                elif filename.endswith(".xlsx"): 
                    file_info = QFileInfo(file_path)
                    item = QTreeWidgetItem(parent_item, [file_info.fileName(), "Arquivo Excel"])
                    item.setData(0, Qt.UserRole, file_path) 
        except Exception as e:
            QMessageBox.warning(self, "Erro ao Listar Arquivos", f"Não foi possível listar arquivos em {directory}: {e}")

    def _on_tree_item_double_clicked(self, item, column):
        """Lida com o clique duplo em um item da árvore."""
        file_path = item.data(0, Qt.UserRole) 
        if file_path and os.path.exists(file_path):
            self._open_excel_file_in_viewer(file_path)
        elif not file_path:
            pass
        else:
            QMessageBox.warning(self, "Arquivo Não Encontrado", f"O arquivo '{os.path.basename(file_path)}' não existe ou o caminho está incorreto.")

    def _open_excel_file_in_viewer(self, file_path):
        """Abre um arquivo Excel usando o ExcelViewerTool."""
        tool_name = f"Viewer: {os.path.basename(file_path)}"
        
        for i in range(self.central_widget.count()):
            if self.central_widget.tabText(i) == tool_name:
                self.central_widget.setCurrentIndex(i)
                return

        try:
            excel_viewer_tool = ExcelViewerTool(file_path=file_path)
            self.central_widget.addTab(excel_viewer_tool, tool_name)
            self.central_widget.setCurrentWidget(excel_viewer_tool)
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Abrir Arquivo", f"Não foi possível abrir '{os.path.basename(file_path)}' no visualizador: {e}")
            
    # --- FUNÇÕES DO MINI-CONSOLE ---
    def _handle_console_command(self, command: str):
        """
        Processa comandos recebidos do mini-console.
        Aqui você pode adicionar lógica para interpretar e executar comandos.
        Por enquanto, apenas ecoa a entrada.
        """
        self.mini_console_widget.append_output(f"Comando recebido: '{command}' (Lógica de execução a ser implementada)")
        # Exemplo: Se você quisesse executar código Python diretamente (CUIDADO com segurança!)
        # try:
        #     exec(command, globals(), locals())
        # except Exception as e:
        #     self.mini_console_widget.append_output(f"Erro ao executar: {e}")

    # --- FUNÇÕES PARA EXECUTAR SCRIPTS EXTERNOS (USADAS PELO MENU ADMIN) ---
    def _run_external_python_script(self, script_path, action, *args):
        """
        Executa um script Python externo em um processo separado com uma ação específica.
        Exibe uma caixa de mensagem com o resultado e envia a saída para o mini-console.
        """
        cmd = [sys.executable, script_path, action] + list(args)
        
        self.mini_console_widget.append_output(f"Executando script: {os.path.basename(script_path)} {action}...")
        self.mini_console_widget.append_output("Aguarde a saída...")

        try:
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, bufsize=1)
            
            def read_stream(stream, is_stderr=False):
                for line in stream:
                    output_line = f"ERRO SCRIPT: {line.strip()}" if is_stderr else line.strip()
                    self.mini_console_widget.append_output(output_line)
                    QApplication.processEvents() 

            stdout_thread = threading.Thread(target=read_stream, args=(process.stdout,))
            stderr_thread = threading.Thread(target=read_stream, args=(process.stderr, True))

            stdout_thread.start()
            stderr_thread.start()

            stdout_thread.join() 
            stderr_thread.join()

            process.wait() 

            if process.returncode == 0:
                self.mini_console_widget.append_output(f"\n--- Script '{os.path.basename(script_path)}' Concluído com Sucesso ---")
                QMessageBox.information(self, "Sucesso na Execução do Script", f"Script '{os.path.basename(script_path)}' executado com sucesso.")
                self._refresh_gui_data() 
            else:
                self.mini_console_widget.append_output(f"\n--- Script '{os.path.basename(script_path)}' Concluído com Erros (Código: {process.returncode}) ---")
                self.mini_console_widget.append_output("Por favor, revise a saída do console para detalhes.")
                QMessageBox.critical(self, "Erro na Execução do Script", f"O script '{os.path.basename(script_path)}' retornou um erro. Veja o console para detalhes.")
                
        except FileNotFoundError:
            err_msg = f"ERRO: O executável Python ou o script '{os.path.basename(script_path)}' não foi encontrado. Verifique o PATH ou o caminho do script."
            self.mini_console_widget.append_output(err_msg)
            QMessageBox.critical(self, "Erro de Arquivo", err_msg)
        except subprocess.CalledProcessError as e:
            err_msg = f"ERRO no Processo para '{os.path.basename(script_path)}':\nstdout: {e.stdout}\nstderr: {e.stderr}"
            self.mini_console_widget.append_output(err_msg)
            QMessageBox.critical(self, "Erro no Processo", err_msg)
        except Exception as e:
            err_msg = f"ERRO Inesperado ao tentar executar o script '{os.path.basename(script_path)}': {e}"
            self.mini_console_widget.append_output(err_msg)
            QMessageBox.critical(self, "Erro Inesperado", err_msg)
        finally:
            self.mini_console_widget.append_output("\n--- Execução de script finalizada ---")


    def _run_create_or_update_all_sheets(self):
        """Executa o script para criar/atualizar todas as planilhas definidas no db_db."""
        if not os.path.exists(UPDATE_METADATA_SCRIPT_PATH): 
            QMessageBox.critical(self, "Erro", f"O script de atualização de metadados não foi encontrado em: {UPDATE_METADATA_SCRIPT_PATH}")
            return
        self._run_external_python_script(UPDATE_METADATA_SCRIPT_PATH, "create_or_update_sheets")

    def _run_sync_db_db_schema(self):
        """Executa o script para sincronizar o schema db_db com os cabeçalhos reais dos arquivos."""
        if not os.path.exists(UPDATE_METADATA_SCRIPT_PATH):
            QMessageBox.critical(self, "Erro", f"O script de atualização de metadados não foi encontrado em: {UPDATE_METADATA_SCRIPT_PATH}")
            return
        self._run_external_python_script(UPDATE_METADATA_SCRIPT_PATH, "update_db_schema")

    def _run_validate_db_consistency(self):
        """
        Abre a ferramenta DbHeadersUpdaterTool em uma aba para validar a consistência do DB.
        """
        self._open_tool("MOD000018", refresh_callback=self._refresh_gui_data) 


# === INÍCIO DO APLICATIVO ===
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    login_window = LoginWindow()
    login_window.show()
    
    sys.exit(app.exec_())

