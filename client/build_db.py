import openpyxl
import shutil
import os

def copy_sheet(src_path, sheet_name, target_wb, dest_name):
    src_wb = openpyxl.load_workbook(src_path)
    src_sheet = src_wb.active
    target_sheet = target_wb.create_sheet(dest_name)

    for row in src_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

def build_combined_db():
    target_path = "user_sheets/db.xlsx"
    os.makedirs("user_sheets", exist_ok=True)

    # Remove if exists
    if os.path.exists(target_path):
        os.remove(target_path)

    # Create empty workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add sheets
    copy_sheet("app_sheets/users.xlsx", "users", wb, "users")
    copy_sheet("app_sheets/tools.xlsx", "tools", wb, "tools")
    copy_sheet("app_sheets/access.xlsx", "access", wb, "access")

    wb.save(target_path)
    print(f"✅ db.xlsx generated at {target_path}")

if __name__ == "__main__":
    build_combined_db()
