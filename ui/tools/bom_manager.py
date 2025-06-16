import sys
import os
import openpyxl
import bcrypt
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox
from PyQt5.QtCore import Qt


# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
user_sheets_dir = os.path.join(project_root, 'user_sheets')
# O DB_EXCEL_PATH é necessário para buscar os schemas das tools
DB_EXCEL_PATH = os.path.join(user_sheets_dir, "db.xlsx")

DEFAULT_DATA_EXCEL_FILENAME = "bom_data.xlsx"
DEFAULT_SHEET_NAME = "BOM"

# Mapeamento de colunas de engenharia.xlsx para a visualização BOM.
# Este mapeamento é crucial para a funcionalidade da ferramenta e não é um cabeçalho "desnecessário".
ENGENHARIA_BOM_MAP = {
    "part_number": "ID do Componente",
    "parent_part_number": "ID do BOM",
    "quantidade": "Quantidade",
    "materia_prima": "Tipo (Matéria Prima)" 
}

class BomManagerTool(QWidget):
    """
    GUI para gerenciar Listas de Materiais (BOMs).
    Permite visualizar, adicionar e salvar informações de BOM.
    Os cabeçalhos da tabela são dinamicamente carregados do arquivo Excel,
    priorizando a configuração em db.xlsx ou o arquivo de dados em si.
    Pode mapear dados de engenharia.xlsx para a visualização BOM.
    """
    def __init__(self, file_path=None, sheet_name=None):
        super().__init__()
        if file_path:
            self.file_path = file_path
        else:
            self.file_path = os.path.join(user_sheets_dir, DEFAULT_DATA_EXCEL_FILENAME)
        
        self.sheet_name = sheet_name if sheet_name else DEFAULT_SHEET_NAME

        # Verifica se o arquivo é engenharia.xlsx para ativar o modo de visualização especial
        self.is_engenharia_file = (os.path.basename(self.file_path) == "engenharia.xlsx")

        self.setWindowTitle(f"Gerenciador de BOM: {os.path.basename(self.file_path)}")
        if self.is_engenharia_file:
            self.setWindowTitle(self.windowTitle() + " (Dados de Engenharia - Somente Leitura)")

        self.layout = QVBoxLayout(self)

        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        if self.is_engenharia_file:
            self.file_name_label.setText(self.file_name_label.text() + " (Mapeado para BOM)")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        # Edição permitida por padrão, mas desabilitada para arquivos de engenharia
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        # Configurações de modo somente leitura se for um arquivo de engenharia
        if self.is_engenharia_file:
            self.add_row_btn.setEnabled(False)
            self.save_btn.setEnabled(False)
            self.table.setEditTriggers(QTableWidget.NoEditTriggers)
            QMessageBox.information(self, "Modo Somente Leitura", "Ao visualizar dados de engenharia como BOM, esta ferramenta opera em modo somente leitura. Para editar, use a ferramenta 'Engenharia (Dados)'.")

        self._populate_sheet_selector()

    def _get_default_bom_display_headers(self):
        """
        Tenta carregar os cabeçalhos padrão para a exibição de BOM da planilha 'tool_schemas' em db.xlsx.
        Se não encontrar, retorna um conjunto básico e alerta o usuário.
        """
        try:
            if not os.path.exists(DB_EXCEL_PATH):
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Arquivo de banco de dados '{os.path.basename(DB_EXCEL_PATH)}' não encontrado. "
                                    "Usando cabeçalhos padrão muito básicos para BOM. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["ID do BOM", "ID do Componente", "Quantidade", "Unidade"]

            wb = openpyxl.load_workbook(DB_EXCEL_PATH, read_only=True)
            if "tool_schemas" not in wb.sheetnames:
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Planilha 'tool_schemas' não encontrada em '{os.path.basename(DB_EXCEL_PATH)}'. "
                                    "Usando cabeçalhos padrão muito básicos para BOM. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["ID do BOM", "ID do Componente", "Quantidade", "Unidade"]
            
            sheet = wb["tool_schemas"]
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            # Mapeia cabeçalhos para índices para fácil acesso
            header_map = {h: idx for idx, h in enumerate(headers)}
            
            tool_name_idx = header_map.get("tool_name")
            schema_type_idx = header_map.get("schema_type")
            header_name_idx = header_map.get("header_name")
            order_idx = header_map.get("order")

            if None in [tool_name_idx, schema_type_idx, header_name_idx, order_idx]:
                QMessageBox.warning(self, "Schema Inválido", 
                                    f"Cabeçalhos esperados (tool_name, schema_type, header_name, order) não encontrados na planilha 'tool_schemas'. "
                                    "Usando cabeçalhos padrão muito básicos para BOM.")
                return ["ID do BOM", "ID do Componente", "Quantidade", "Unidade"]

            configured_headers = []
            for row_idx in range(2, sheet.max_row + 1): # Começa da segunda linha para pular cabeçalhos
                row_values = [cell.value for cell in sheet[row_idx]]
                
                # Garante que a linha tem dados suficientes
                if len(row_values) > max(tool_name_idx, schema_type_idx, header_name_idx, order_idx):
                    tool = row_values[tool_name_idx]
                    schema = row_values[schema_type_idx]
                    header = row_values[header_name_idx]
                    order = row_values[order_idx]

                    if tool == "BomManagerTool" and schema == "default_bom_display" and header:
                        configured_headers.append((header, order))
            
            # Ordena os cabeçalhos com base na coluna 'order'
            configured_headers.sort(key=lambda x: x[1] if x[1] is not None else float('inf'))
            return [h[0] for h in configured_headers]

        except Exception as e:
            QMessageBox.critical(self, "Erro de Configuração", 
                                f"Erro ao carregar configurações de cabeçalho de db.xlsx: {e}. "
                                "Usando cabeçalhos padrão muito básicos para BOM.")
            return ["ID do BOM", "ID do Componente", "Quantidade", "Unidade"]


    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        # Obtém os cabeçalhos padrão/configurados para o BOM
        default_bom_headers = self._get_default_bom_display_headers()

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", 
                                f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. "
                                f"Ele será criado com a aba padrão '{DEFAULT_SHEET_NAME}' e os cabeçalhos definidos ao salvar.")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(len(default_bom_headers))
            self.table.setHorizontalHeaderLabels(default_bom_headers)
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", 
                                    f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. "
                                    f"Adicionando a aba padrão '{DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                default_index = self.sheet_selector.findText(self.sheet_name)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names:
                    self.sheet_selector.setCurrentIndex(0)
            
            self._load_data_from_selected_sheet()

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """Carrega dados da planilha Excel atualmente selecionada para o QTableWidget, usando cabeçalhos reais."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                # Se o arquivo não existe, inicializa a tabela com os cabeçalhos padrão do BOM.
                default_bom_headers = self._get_default_bom_display_headers()
                self.table.setRowCount(0)
                self.table.setColumnCount(len(default_bom_headers))
                self.table.setHorizontalHeaderLabels(default_bom_headers)
                return

            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                # Se a planilha não existe, cria-a com cabeçalhos padrão do BOM e salva.
                default_bom_headers = self._get_default_bom_display_headers()
                ws = wb.create_sheet(current_sheet_name)
                ws.append(default_bom_headers)
                wb.save(self.file_path)
                QMessageBox.information(self, "Planilha Criada", 
                                        f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. "
                                        "Uma nova foi criada com cabeçalhos padrão.")
                self._populate_sheet_selector() # Recarrega o seletor de abas para incluir a nova
                return

            sheet = wb[current_sheet_name]

            source_headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            data_to_display = []
            
            if self.is_engenharia_file:
                # Caso seja engenharia.xlsx, aplica o mapeamento e inclui colunas adicionais
                # Os cabeçalhos de exibição preferenciais vêm de db.xlsx (o que costumava ser BOM_HEADERS)
                preferred_display_headers = self._get_default_bom_display_headers()
                
                display_headers = list(preferred_display_headers)
                
                source_header_idx_map = {h: idx for idx, h in enumerate(source_headers)}

                for row_idx in range(2, sheet.max_row + 1):
                    row_values = [cell.value for cell in sheet[row_idx]]
                    # Inicializa a linha mapeada com vazios para o tamanho dos cabeçalhos preferenciais
                    mapped_row = [""] * len(preferred_display_headers) 

                    # Mapeia os dados da engenharia.xlsx para as colunas de exibição preferenciais
                    for eng_header, bom_header in ENGENHARIA_BOM_MAP.items():
                        eng_idx = source_header_idx_map.get(eng_header)
                        if eng_idx is not None and eng_idx < len(row_values):
                            try:
                                bom_idx = preferred_display_headers.index(bom_header)
                                mapped_row[bom_idx] = row_values[eng_idx]
                            except ValueError:
                                # Se o cabeçalho BOM esperado não estiver nos preferred_display_headers, ignora ou adiciona como extra
                                pass

                    # Adiciona quaisquer colunas de engenharia.xlsx que não foram mapeadas explicitamente
                    for h_idx, h_name in enumerate(source_headers):
                        if h_name not in ENGENHARIA_BOM_MAP.keys(): # Verifica se o nome do cabeçalho original não está nas chaves do mapeamento
                            if h_name not in display_headers: # Garante que não adiciona cabeçalhos já presentes (do preferred_display_headers)
                                display_headers.append(h_name)
                                # Adiciona um placeholder vazio para esta nova coluna em todas as linhas existentes
                                for existing_row in data_to_display:
                                    existing_row.append("")
                            
                            # Encontra o índice da coluna recém-adicionada ou existente
                            display_idx = display_headers.index(h_name)
                            # Extende a linha atual se necessário para acomodar a nova coluna
                            while len(mapped_row) <= display_idx:
                                mapped_row.append("")
                            if h_idx < len(row_values):
                                mapped_row[display_idx] = row_values[h_idx]

                    # Garante que a linha tenha o mesmo número de colunas que os display_headers
                    while len(mapped_row) < len(display_headers):
                        mapped_row.append("")

                    data_to_display.append(mapped_row)
                
                # Atualiza display_headers para incluir os cabeçalhos adicionais na ordem correta
                # Remove duplicatas e mantém a ordem original dos preferred_display_headers
                # e adiciona os novos no final.
                final_display_headers = []
                seen_headers = set()
                for h in preferred_display_headers:
                    if h not in seen_headers:
                        final_display_headers.append(h)
                        seen_headers.add(h)
                for h in display_headers: # display_headers agora contém preferred_display_headers + extras
                    if h not in seen_headers:
                        final_display_headers.append(h)
                        seen_headers.add(h)
                display_headers = final_display_headers

            else:
                # Caso seja bom_data.xlsx ou outro arquivo de BOM
                if not source_headers:
                    # Se a planilha está vazia, usa os cabeçalhos padrão do BOM carregados de db.xlsx
                    display_headers = self._get_default_bom_display_headers()
                else:
                    # Caso contrário, usa os cabeçalhos reais da planilha
                    display_headers = source_headers
                
                for row in sheet.iter_rows(min_row=2):
                    row_values = [cell.value for cell in row]
                    # Garante que a linha tenha o mesmo número de colunas que os cabeçalhos
                    while len(row_values) < len(display_headers):
                        row_values.append("")
                    data_to_display.append(row_values)
            
            self.table.setColumnCount(len(display_headers))
            self.table.setHorizontalHeaderLabels(display_headers)
            self.table.setRowCount(len(data_to_display))
            for row_idx, row_data in enumerate(data_to_display):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            # Re-aplica a configuração de somente leitura se for engenharia.xlsx
            if self.is_engenharia_file:
                self.table.setEditTriggers(QTableWidget.NoEditTriggers)
            else:
                self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados do BOM da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(len(self._get_default_bom_display_headers())) 
            self.table.setHorizontalHeaderLabels(self._get_default_bom_display_headers())

    def _save_data(self):
        """Salva dados do QTableWidget de volta para a planilha Excel, mantendo cabeçalhos existentes ou usando padrão."""
        if self.is_engenharia_file: 
            QMessageBox.warning(self, "Ação Não Permitida", "Esta ferramenta está visualizando dados de engenharia em modo somente leitura. Não é possível salvar alterações aqui.")
            return

        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(current_sheet_name)
                # Ao criar uma nova planilha para salvar, usa os cabeçalhos atualmente exibidos na tabela
                headers_to_save = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                if not headers_to_save: # Fallback se a tabela estiver vazia de alguma forma
                    headers_to_save = self._get_default_bom_display_headers()
                ws.append(headers_to_save)
                
                QMessageBox.information(self, "Arquivo e Planilha Criados", f"Nova planilha '{current_sheet_name}' criada em '{os.path.basename(self.file_path)}'.")
            else:
                ws = wb[current_sheet_name]
            
            # Limpa todas as linhas existentes, exceto a primeira (cabeçalhos)
            for row_idx in range(ws.max_row, 1, -1):
                ws.delete_rows(row_idx)

            # Obtém os cabeçalhos atuais da tabela (o que está sendo exibido)
            current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
            existing_sheet_headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []

            # Se os cabeçalhos atuais da tabela forem diferentes dos cabeçalhos da planilha, atualiza a primeira linha.
            if existing_sheet_headers != current_headers:
                if ws.max_row > 0: # Se houver uma linha de cabeçalho existente, a apaga
                    ws.delete_rows(1)
                ws.insert_rows(1) # Insere uma nova primeira linha
                ws.append(current_headers) # Adiciona os novos cabeçalhos
            elif not existing_sheet_headers and current_headers: # Caso a planilha esteja completamente vazia
                ws.append(current_headers)
            
            # Adiciona os dados da tabela
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                ws.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados do BOM: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        if self.is_engenharia_file: 
            QMessageBox.warning(self, "Ação Não Permitida", "Esta ferramenta está visualizando dados de engenharia em modo somente leitura. Não é possível adicionar linhas.")
            return

        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Configura um caminho de teste para db.xlsx para o ambiente de teste da tool
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    db_test_path = os.path.join(user_sheets_dir_test, "db.xlsx")

    # Cria um db.xlsx de teste com a planilha tool_schemas se não existir
    if not os.path.exists(db_test_path):
        db_wb = openpyxl.Workbook()
        db_ws_users = db_wb.active # Renomeia a primeira sheet
        db_ws_users.title = "users"
        db_ws_users.append(["id", "username", "password_hash", "role"])
        db_ws_users.append([1, "admin", bcrypt.hashpw("admin".encode(), bcrypt.gensalt()).decode(), "admin"])
        db_ws_users.append([2, "user", bcrypt.hashpw("user".encode(), bcrypt.gensalt()).decode(), "user"])

        db_ws_access = db_wb.create_sheet("access")
        db_ws_access.append(["role", "allowed_modules"])
        db_ws_access.append(["admin", "all"])
        db_ws_access.append(["user", "mod1,mod3"]) # Exemplo

        db_ws_tools = db_wb.create_sheet("tools")
        db_ws_tools.append(["id", "name", "description", "path"])
        db_ws_tools.append(["mod1", "Gerenciador de BOM", "Gerencia Listas de Materiais", "ui.tools.bom_manager"])
        db_ws_tools.append(["mod4", "Engenharia (Workflow)", "Ferramenta de Workflow", "ui.tools.engenharia_data"])
        
        # Adiciona a planilha tool_schemas
        db_ws_schemas = db_wb.create_sheet("tool_schemas")
        db_ws_schemas.append(["tool_name", "schema_type", "header_name", "order"])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do BOM", 1])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do Componente", 2])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Nome do Componente", 3])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Quantidade", 4])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Unidade", 5])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Ref Designator", 6])


        db_wb.save(db_test_path)
        print(f"Arquivo de teste db.xlsx criado em: {db_test_path}")

    # Cria um arquivo bom_data.xlsx de teste se não existir, sem cabeçalhos inicialmente para testar carregamento dinâmico
    test_file_path = os.path.join(user_sheets_dir_test, DEFAULT_DATA_EXCEL_FILENAME)
    if not os.path.exists(test_file_path):
        wb_bom = openpyxl.Workbook()
        ws_bom = wb_bom.active
        ws_bom.title = DEFAULT_SHEET_NAME
        # Não adiciona headers aqui para forçar a leitura de db.xlsx ou o fallback
        wb_bom.save(test_file_path)
        print(f"Arquivo de teste {DEFAULT_DATA_EXCEL_FILENAME} criado vazio para testar carregamento de headers.")

    # Cria um arquivo engenharia.xlsx de teste se não existir
    engenharia_test_path = os.path.join(user_sheets_dir_test, "engenharia.xlsx")
    if not os.path.exists(engenharia_test_path):
        eng_wb = openpyxl.Workbook()
        eng_ws = eng_wb.active
        eng_ws.title = "Estrutura"
        eng_ws.append(["part_number", "parent_part_number", "quantidade", "materia_prima", "revisao"])
        eng_ws.append(["PROD-001", "", 1, "Não", "A"])
        eng_ws.append(["ASSY-A", "PROD-001", 1, "Não", "B"])
        eng_ws.append(["RAW-MAT-001", "ASSY-A", 10, "Sim", "C"])
        eng_wb.save(engenharia_test_path)
        print(f"Arquivo de teste engenharia.xlsx criado para testar visualização BOM.")


    # Testando a ferramenta com bom_data.xlsx
    window_bom = BomManagerTool(file_path=test_file_path)
    window_bom.show()

    # Testando a ferramenta com engenharia.xlsx
    # window_eng = BomManagerTool(file_path=engenharia_test_path)
    # window_eng.show() # Para testar esta, comente a linha window_bom.show() e descomente esta.

    sys.exit(app.exec_())

