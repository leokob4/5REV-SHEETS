import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox
from PyQt5.QtCore import Qt

# Define o caminho padrão para o arquivo Excel para esta ferramenta.
# Pode ser sobrescrito ao instanciar a ferramenta com um 'file_path'.
DEFAULT_DATA_EXCEL_FILENAME = "output.xlsx"
DEFAULT_SHEET_NAME = "product_data" # Nome da planilha padrão para esta ferramenta

class ProductDataTool(QWidget):
    """
    GUI para gerenciar Dados do Produto.
    Permite visualizar, adicionar e salvar informações do produto em 'output.xlsx'.
    Permite redimensionamento interativo de colunas e linhas.
    """
    def __init__(self, file_path=None):
        super().__init__()
        # Se um file_path for fornecido, use-o; caso contrário, construa o caminho padrão.
        if file_path:
            self.file_path = file_path
        else:
            # Obtém o diretório raiz do projeto para construir o caminho padrão.
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            self.file_path = os.path.join(project_root, 'user_sheets', DEFAULT_DATA_EXCEL_FILENAME)

        self.setWindowTitle(f"Dados do Produto: {os.path.basename(self.file_path)}")
        self.layout = QVBoxLayout(self)

        # Cabeçalho com nome do arquivo e controles
        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch() # Pushes other elements to the right

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        self.table.setAlternatingRowColors(True)
        # Habilitar redimensionamento interativo de colunas e linhas
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        # Botões de controle
        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        self._populate_sheet_selector() # Popula o dropdown e carrega os dados iniciais

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        
        # Garante que o diretório exista antes de tentar listar arquivos
        user_sheets_dir = os.path.dirname(self.file_path)
        os.makedirs(user_sheets_dir, exist_ok=True)

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. Ele será criado com a aba padrão '{DEFAULT_SHEET_NAME}' ao salvar.")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.setHorizontalHeaderLabels(["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]) # Cabeçalhos padrão
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(DEFAULT_SHEET_NAME) # Sempre oferece o padrão
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. Adicionando a aba padrão '{DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                # Define a planilha padrão se ela existir, caso contrário, seleciona a primeira
                default_index = self.sheet_selector.findText(DEFAULT_SHEET_NAME)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names:
                    self.sheet_selector.setCurrentIndex(0) # Seleciona a primeira planilha disponível
                else:
                    self.sheet_selector.setCurrentIndex(0)

            # Aciona manualmente _load_data_from_selected_sheet após popular
            self._load_data_from_selected_sheet()

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME) # Fallback para o nome padrão em caso de erro
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """Carrega dados da planilha Excel atualmente selecionada para o QTableWidget."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                self.table.setRowCount(0)
                self.table.setColumnCount(0)
                self.table.setHorizontalHeaderLabels(["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]) # Cabeçalhos padrão para um arquivo novo
                return

            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                QMessageBox.information(self, "Planilha Não Encontrada", f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. Criando uma nova.")
                ws = wb.create_sheet(current_sheet_name)
                default_headers = ["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]
                ws.append(default_headers)
                wb.save(self.file_path)
                self._populate_sheet_selector() # Atualiza o seletor para incluir a nova planilha
                return # Sai, pois _populate_sheet_selector acionará um novo carregamento

            sheet = wb[current_sheet_name]

            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            data = []
            for row in sheet.iter_rows(min_row=2):
                data.append([cell.value for cell in row])

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive) # Aplicar redimensionamento interativo
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive) # Aplicar redimensionamento interativo
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados do produto da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.setHorizontalHeaderLabels(["Erro", "Erro", "Erro", "Erro", "Erro"])

    def _save_data(self):
        """Salva dados do QTableWidget de volta para a planilha Excel."""
        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
                current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                if not current_headers:
                    current_headers = ["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]
                ws.append(current_headers)
                wb.save(self.file_path)
                QMessageBox.information(self, "Arquivo e Planilha Criados", f"Novo arquivo '{os.path.basename(self.file_path)}' com planilha '{current_sheet_name}' criado.")
                self._populate_sheet_selector() 
                # Continua para salvar os dados na planilha recém-criada
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                    current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                    if not current_headers:
                        current_headers = ["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]
                    ws.append(current_headers)
                    wb.save(self.file_path)
                    QMessageBox.information(self, "Planilha Criada", f"Nova planilha '{current_sheet_name}' criada em '{os.path.basename(self.file_path)}'.")
                    self._populate_sheet_selector()
                    # Continua para salvar os dados na planilha recém-criada

            sheet = wb[current_sheet_name]
            
            # Clear existing data but keep header (row 1)
            for row_idx in range(sheet.max_row, 1, -1):
                sheet.delete_rows(row_idx)

            # Write current headers from table (in case they were changed in GUI, though not expected for now)
            current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
            if not current_headers: # If no headers are set in the table (e.g., table was empty on load)
                current_headers = ["ID", "Nome do Produto", "Código", "Revisão", "Descrição"]
                self.table.setColumnCount(len(current_headers))
                self.table.setHorizontalHeaderLabels(current_headers)
            
            existing_sheet_headers = [cell.value for cell in sheet[1]]
            if existing_sheet_headers != current_headers:
                for col_idx, header_value in enumerate(current_headers):
                    sheet.cell(row=1, column=col_idx + 1, value=header_value)
            
            # Append all rows from the QTableWidget
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                sheet.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados do produto: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    test_file_dir = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(test_file_dir, exist_ok=True)
    test_file_path = os.path.join(test_file_dir, DEFAULT_DATA_EXCEL_FILENAME)
    
    # Criar um workbook vazio se não existir
    if not os.path.exists(test_file_path):
        wb = openpyxl.Workbook()
        wb.save(test_file_path)

    window = ProductDataTool(file_path=test_file_path)
    window.show()
    sys.exit(app.exec_())
