import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox, QFileDialog, QInputDialog
from PyQt5.QtCore import Qt

# Não há headers hardcoded aqui; a ferramenta lê diretamente da primeira linha da planilha.

class SheetEditorWidget(QWidget):
    """
    GUI para editar qualquer arquivo Excel (.xlsx).
    Permite abrir arquivos, navegar entre planilhas, editar dados,
    adicionar/remover linhas e salvar alterações.
    Os cabeçalhos são lidos da primeira linha da planilha.
    """
    def __init__(self):
        super().__init__()
        self.file_path = None # O caminho do arquivo será definido ao carregar
        self.current_workbook = None # Armazenará o objeto openpyxl.Workbook

        self.setWindowTitle("📄 Sheet Editor")
        self.layout = QVBoxLayout(self)

        # Layout superior para o nome do arquivo e seletor de planilha
        file_sheet_layout = QHBoxLayout()
        self.file_name_label = QLabel("<b>Arquivo:</b> Nenhum arquivo carregado")
        file_sheet_layout.addWidget(self.file_name_label)
        file_sheet_layout.addStretch() # Empurra o resto para a direita

        file_sheet_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        file_sheet_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        file_sheet_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(file_sheet_layout)

        # Botão para carregar arquivo
        self.load_btn = QPushButton("📂 Carregar Arquivo Excel")
        self.load_btn.clicked.connect(self._load_excel_file)
        self.layout.addWidget(self.load_btn)

        # Tabela principal para exibir os dados
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed) # Permite edição
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        # Layout para os botões de ação (Salvar, Adicionar Linha, Deletar Linha, Adicionar Planilha)
        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.delete_row_btn = QPushButton("Deletar Linha Selecionada")
        self.delete_row_btn.clicked.connect(self._delete_selected_row)
        self.add_sheet_btn = QPushButton("Adicionar Nova Planilha")
        self.add_sheet_btn.clicked.connect(self._add_new_sheet)
        self.save_btn = QPushButton("💾 Salvar Alterações")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet) # Recarrega a aba, não o arquivo todo

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.delete_row_btn)
        button_layout.addWidget(self.add_sheet_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        # Desabilitar botões inicialmente até que um arquivo seja carregado
        self._set_buttons_enabled(False)


    def _set_buttons_enabled(self, enabled):
        """Controla a habilitação dos botões de edição/salvamento."""
        self.add_row_btn.setEnabled(enabled)
        self.delete_row_btn.setEnabled(enabled)
        self.add_sheet_btn.setEnabled(enabled)
        self.save_btn.setEnabled(enabled)
        self.refresh_btn.setEnabled(enabled)
        self.sheet_selector.setEnabled(enabled)
        self.refresh_sheets_btn.setEnabled(enabled)

    def _load_excel_file(self):
        """Permite ao usuário selecionar e carregar um arquivo Excel."""
        # Caminho inicial para QFileDialog
        initial_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'user_sheets')
        os.makedirs(initial_dir, exist_ok=True) # Garante que a pasta user_sheets existe

        file, _ = QFileDialog.getOpenFileName(self, "Selecionar Arquivo Excel", initial_dir, "Arquivos Excel (*.xlsx)")
        if not file:
            return

        self.file_path = file
        self.file_name_label.setText(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        self._load_workbook() # Carrega o workbook associado ao caminho do arquivo
        self._populate_sheet_selector() # Popula o seletor de planilhas com as sheets do novo arquivo
        self._set_buttons_enabled(True) # Habilita os botões

    def _load_workbook(self):
        """Carrega o workbook do arquivo Excel."""
        try:
            if not os.path.exists(self.file_path):
                QMessageBox.critical(self, "Erro", f"O arquivo '{os.path.basename(self.file_path)}' não foi encontrado.")
                self.current_workbook = None
                return
            self.current_workbook = openpyxl.load_workbook(self.file_path)
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Carregar Arquivo", f"Não foi possível carregar o arquivo '{os.path.basename(self.file_path)}': {e}")
            self.current_workbook = None # Garante que não há workbook carregado em caso de erro

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do workbook atual."""
        self.sheet_selector.clear()
        if self.current_workbook:
            for sheet_name in self.current_workbook.sheetnames:
                self.sheet_selector.addItem(sheet_name)
            
            # Tenta selecionar a primeira sheet por padrão
            if self.current_workbook.sheetnames:
                self.sheet_selector.setCurrentIndex(0)
            
            self._load_data_from_selected_sheet() # Carrega dados da aba selecionada (ou vazia)
        else:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """Carrega dados da planilha Excel atualmente selecionada para o QTableWidget."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.current_workbook:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            sheet = self.current_workbook[current_sheet_name]

            # Carrega cabeçalhos da primeira linha da planilha. Se não houver, assume 0 colunas.
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            data = []
            # Itera a partir da segunda linha para os dados
            for row in sheet.iter_rows(min_row=2): 
                row_values = [cell.value for cell in row]
                # Garante que a linha tenha células suficientes para os cabeçalhos
                while len(row_values) < len(headers):
                    row_values.append("")
                data.append(row_values)

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
            # QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(0) # Limpa a tabela em caso de erro grave

    def _save_data(self):
        """Salva dados do QTableWidget de volta para a planilha Excel, capturando os cabeçalhos da tabela."""
        if not self.current_workbook or not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo Excel está carregado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            # Obtém a planilha atual do workbook
            ws = self.current_workbook[current_sheet_name]
            
            # Limpa todas as linhas existentes na planilha
            for row_idx in range(ws.max_row, 0, -1):
                ws.delete_rows(row_idx)

            # Obtém os cabeçalhos atuais da QTableWidget.
            current_headers = [self.table.horizontalHeaderItem(col).text() 
                               for col in range(self.table.columnCount())]
            
            # Salva os cabeçalhos se existirem (se o usuário digitou ou eles foram carregados)
            if current_headers:
                ws.append(current_headers)
            
            # Percorre o QTableWidget e adiciona as linhas ao Excel
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                ws.append(row_data)

            self.current_workbook.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
            self._populate_sheet_selector() # Recarrega para garantir que o seletor esteja atualizado se uma nova aba foi criada
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        
        # Se a tabela ainda não tem colunas (ex: planilha nova/vazia),
        # esta é a primeira linha, e os valores aqui servirão como cabeçalhos na próxima gravação.
        if self.table.columnCount() == 0 and row_count == 0:
            text, ok = QInputDialog.getText(self, "Definir Cabeçalhos", 
                                            "A planilha está vazia. Insira os nomes das colunas separados por vírgula (ex: ID, Nome, Quantidade):")
            if ok and text:
                headers = [h.strip() for h in text.split(',')]
                self.table.setColumnCount(len(headers))
                self.table.setHorizontalHeaderLabels(headers)
            else:
                QMessageBox.warning(self, "Aviso", "Nenhum cabeçalho fornecido. Nenhuma coluna será adicionada.")
                self.table.removeRow(row_count) # Remove a linha vazia recém-adicionada
                return # Sai da função se o usuário cancelar ou não fornecer cabeçalhos

        # Preenche a nova linha com itens vazios (ou se já houver colunas definidas)
        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

    def _delete_selected_row(self):
        """Deleta a(s) linha(s) selecionada(s) da QTableWidget."""
        selected_rows = sorted(set(index.row() for index in self.table.selectedIndexes()))
        if not selected_rows:
            QMessageBox.warning(self, "Nenhuma Linha Selecionada", "Por favor, selecione uma ou mais linhas para deletar.")
            return

        reply = QMessageBox.question(self, "Confirmar Deleção", 
                                     f"Tem certeza que deseja deletar {len(selected_rows)} linha(s) selecionada(s)?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            for row_idx in reversed(selected_rows): # Deleta em ordem reversa para evitar problemas de índice
                self.table.removeRow(row_idx)
            QMessageBox.information(self, "Linha(s) Deletada(s)", "Linha(s) selecionada(s) deletada(s) com sucesso. Lembre-se de salvar as alterações.")

    def _add_new_sheet(self):
        """Permite ao usuário adicionar uma nova planilha."""
        sheet_name, ok = QInputDialog.getText(self, "Nova Planilha", "Nome da Nova Planilha:")
        if not ok or not sheet_name.strip():
            return

        sheet_name = sheet_name.strip()
        if self.current_workbook and sheet_name in self.current_workbook.sheetnames:
            QMessageBox.warning(self, "Nome Duplicado", f"Uma planilha com o nome '{sheet_name}' já existe neste arquivo.")
            return

        try:
            if not self.current_workbook: # Se nenhum arquivo foi carregado, cria um novo workbook
                self.current_workbook = openpyxl.Workbook()
                if 'Sheet' in self.current_workbook.sheetnames: # Remove a sheet padrão
                    del self.current_workbook['Sheet']
                self.file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'user_sheets', 'new_workbook.xlsx')
                QMessageBox.information(self, "Novo Arquivo Criado", f"Nenhum arquivo estava carregado. Um novo arquivo 'new_workbook.xlsx' foi criado em 'user_sheets'.")
                self._set_buttons_enabled(True)
                self.file_name_label.setText(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")

            self.current_workbook.create_sheet(title=sheet_name)
            self.current_workbook.save(self.file_path)
            QMessageBox.information(self, "Planilha Criada", f"Planilha '{sheet_name}' criada com sucesso.")
            self._populate_sheet_selector() # Recarrega o seletor para incluir a nova planilha
            self.sheet_selector.setCurrentText(sheet_name) # Seleciona a nova planilha
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Criar Planilha", f"Erro ao criar nova planilha: {e}")


# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    test_file_dir = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(test_file_dir, exist_ok=True)
    
    # Crie um arquivo de teste para o editor
    test_editor_file = os.path.join(test_file_dir, "test_editor.xlsx")
    
    if not os.path.exists(test_editor_file):
        wb_test = openpyxl.Workbook()
        ws1 = wb_test.active
        ws1.title = "DadosEditaveis"
        ws1.append(["Col1", "Col2", "Col3"])
        ws1.append(["A1", "B1", "C1"])
        ws1.append(["A2", "B2", "C2"])
        
        ws2 = wb_test.create_sheet("OutraAba")
        ws2.append(["HeaderX", "HeaderY"])
        ws2.append(["X1", "Y1"])

        ws3 = wb_test.create_sheet("AbaVaziaParaCriarHeaders") # Uma planilha sem dados nem cabeçalhos
        
        wb_test.save(test_editor_file)
        print(f"Arquivo de teste '{test_editor_file}' criado para o editor.")
    else:
        print(f"Arquivo de teste '{test_editor_file}' já existe, usando existente para o editor.")


    window = SheetEditorWidget()
    window.show()
    sys.exit(app.exec_())
