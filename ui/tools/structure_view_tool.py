import os
import openpyxl
import sys
import PyQt5
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTreeWidget, QTreeWidgetItem, QLabel, QMessageBox, QHeaderView, QComboBox
from PyQt5.QtCore import Qt

class StructureViewTool(QWidget):
    """
    GUI para visualizar a estrutura hierárquica (e.g., BOM ou estrutura de arquivo)
    de um item selecionado ou de uma planilha específica em um arquivo Excel.
    Os cabeçalhos da árvore são carregados EXCLUSIVAMENTE da primeira linha do arquivo Excel.
    Possui suporte específico para o arquivo 'engenharia.xlsx' e usa 'ParentID'/'ComponentID'
    ou 'parent_part_number'/'part_number' para inferir a hierarquia.
    """
    def __init__(self, file_path=None, sheet_name=None):
        super().__init__()
        # Definindo caminhos de forma dinâmica a partir da localização do script
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
        user_sheets_dir = os.path.join(project_root, 'user_sheets')

        if file_path:
            self.file_path = file_path
        else:
            # Define um caminho padrão se nenhum for fornecido, por exemplo, para um caso de teste padrão
            self.file_path = os.path.join(user_sheets_dir, "engenharia.xlsx") 
            # ou "bom_data.xlsx" dependendo da sua preferência de default

        self.sheet_name = sheet_name # Este será usado para tentar selecionar a sheet inicialmente
        
        self.is_engenharia_file = (os.path.basename(self.file_path).lower() == "engenharia.xlsx")

        self.setWindowTitle(f"Estrutura: {os.path.basename(self.file_path)}")
        if self.is_engenharia_file:
            self.setWindowTitle(self.windowTitle() + " (Engenharia)")

        self.layout = QVBoxLayout(self)

        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_structure_data)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.structure_tree = QTreeWidget()
        self.structure_tree.header().setSectionResizeMode(QHeaderView.Interactive)
        # self.structure_tree.verticalHeader().setSectionResizeMode(QHeaderView.Interactive) # Vertical header not typically interactive
        self.layout.addWidget(self.structure_tree)

        self._populate_sheet_selector() # Initial population and data load

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        
        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. Crie-o para continuar.")
            # Limpa a árvore e retorna se o arquivo não existe
            self.structure_tree.clear()
            self.structure_tree.setHeaderLabels([]) # Limpa cabeçalhos
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'.")
                self.structure_tree.clear()
                self.structure_tree.setHeaderLabels([]) # Limpa cabeçalhos
                return
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                # Prioriza a sheet que foi passada (e.g., "Estrutura" para engenharia.xlsx) ou a primeira disponível
                default_sheet_to_select = self.sheet_name if self.sheet_name else (sheet_names[0] if sheet_names else "")
                default_index = self.sheet_selector.findText(default_sheet_to_select)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names: # Fallback: if specified sheet not found, select the first
                    self.sheet_selector.setCurrentIndex(0)
                
            self._load_structure_data() # Load data for the selected sheet

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.structure_tree.clear()
            self.structure_tree.setHeaderLabels([]) # Limpa cabeçalhos


    def _load_structure_data(self):
        """Carrega e exibe a estrutura hierárquica do arquivo e planilha especificados."""
        self.structure_tree.clear() # Limpa a árvore existente
        self.structure_tree.setHeaderLabels([]) # Limpa os cabeçalhos da árvore

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path or not os.path.exists(self.file_path):
            self.structure_tree.addTopLevelItem(QTreeWidgetItem(["N/A", "Arquivo ou planilha não selecionados/encontrados."]))
            return

        try:
            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                QMessageBox.warning(self, "Planilha Não Encontrada", f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'.")
                self.structure_tree.addTopLevelItem(QTreeWidgetItem(["N/A", "Planilha não encontrada."]))
                return

            sheet = wb[current_sheet_name]
            
            # Carrega cabeçalhos da primeira linha da planilha
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            if not headers:
                QMessageBox.information(self, "Planilha Vazia", 
                                        f"A planilha '{current_sheet_name}' está vazia ou não possui cabeçalhos. Nenhuma estrutura para exibir.")
                self.structure_tree.addTopLevelItem(QTreeWidgetItem(["N/A", "Nenhum dado de estrutura."]))
                return
            
            self.structure_tree.setHeaderLabels(headers)
            
            parent_id_col = -1
            component_id_col = -1
            
            # Tenta encontrar colunas de ID e ParentID com base nos nomes esperados
            # priorizando nomes de engenharia.xlsx, depois nomes genéricos
            try:
                parent_id_col = headers.index("parent_part_number")
            except ValueError:
                if "ParentID" in headers:
                    parent_id_col = headers.index("ParentID")

            try:
                component_id_col = headers.index("part_number")
            except ValueError:
                if "ComponentID" in headers:
                    component_id_col = headers.index("ComponentID")

            if parent_id_col == -1 or component_id_col == -1:
                QMessageBox.critical(self, "Erro de Cabeçalho", 
                                     "Não foi possível identificar as colunas de ID do Componente (ex: 'part_number' ou 'ComponentID') "
                                     "e ID do Pai (ex: 'parent_part_number' ou 'ParentID'). "
                                     "Verifique os cabeçalhos da planilha selecionada.")
                self.structure_tree.addTopLevelItem(QTreeWidgetItem(["Erro", "Cabeçalhos de estrutura não encontrados."]))
                return

            children_map = {}
            all_component_ids = set()
            all_parent_ids_in_data = set()

            for row_idx in range(2, sheet.max_row + 1): # Começa da segunda linha (após os cabeçalhos)
                row_values = [str(cell.value) if cell.value is not None else "" for cell in sheet[row_idx]]
                
                # Garante que a linha tem dados suficientes para as colunas de ID
                if len(row_values) > max(parent_id_col, component_id_col):
                    parent_id = row_values[parent_id_col].strip()
                    component_id = row_values[component_id_col].strip()

                    if component_id: 
                        all_component_ids.add(component_id)
                        if parent_id:
                            all_parent_ids_in_data.add(parent_id)
                            if parent_id not in children_map:
                                children_map[parent_id] = []
                            children_map[parent_id].append({"ID": component_id, "Data": row_values})
            
            # Identifica os itens raiz (aqueles que são componentes mas não são filhos de ninguém listado)
            # ou que explicitamente têm parent_id vazio
            root_items_ids = []
            for component_id in all_component_ids:
                if component_id not in all_parent_ids_in_data:
                    # Verifica também se há explicitamente uma linha onde o parent_id está vazio para este component_id
                    found_empty_parent = False
                    for row_idx in range(2, sheet.max_row + 1):
                        row_values = [str(cell.value) if cell.value is not None else "" for cell in sheet[row_idx]]
                        if len(row_values) > max(parent_id_col, component_id_col) and \
                           row_values[component_id_col].strip() == component_id and \
                           row_values[parent_id_col].strip() == "":
                           found_empty_parent = True
                           break
                    if found_empty_parent:
                        root_items_ids.append(component_id)

            # Fallback para caso não haja raízes claras (e.g., uma lista simples sem hierarquia clara)
            if not root_items_ids and all_component_ids:
                # Se houver apenas um componente, ele é a raiz
                if len(all_component_ids) == 1:
                    root_items_ids = list(all_component_ids)
                else:
                    # Se múltiplos, e não há parent_id vazio, escolha o primeiro como raiz
                    # Isso pode não ser ideal para todas as estruturas, mas fornece um ponto de partida
                    QMessageBox.information(self, "Aviso de Estrutura", 
                                            "Não foi possível identificar raízes claras na estrutura. Exibindo o primeiro componente encontrado como raiz.")
                    root_items_ids = [list(all_component_ids)[0]]

            if not root_items_ids:
                QMessageBox.information(self, "Nenhuma Estrutura Encontrada", 
                                        f"Nenhum dado de estrutura hierárquica válido encontrado na planilha '{current_sheet_name}' do arquivo '{os.path.basename(self.file_path)}'. "
                                        "Verifique se as colunas de ID do Componente e ID do Pai estão presentes e corretas.")
                self.structure_tree.addTopLevelItem(QTreeWidgetItem(["N/A", "Nenhum dado de estrutura."]))
                return

            for root_id in root_items_ids:
                root_data = None
                # Busca a linha completa do item raiz para popular o QTreeWidgetItem
                for row_idx in range(2, sheet.max_row + 1):
                    row_values = [str(cell.value) if cell.value is not None else "" for cell in sheet[row_idx]]
                    if len(row_values) > component_id_col and row_values[component_id_col].strip() == root_id:
                        root_data = row_values
                        break
                
                if root_data:
                    root_q_item = QTreeWidgetItem(root_data)
                    self.structure_tree.addTopLevelItem(root_q_item)
                    self._add_items_to_tree(root_q_item, children_map, root_id, headers, parent_id_col, component_id_col)
                else: 
                    # Cria um item raiz "dummy" se a linha de dados completa para o ID raiz não for encontrada
                    # (isso pode acontecer se um ID for uma raiz mas só aparecer como pai de outros)
                    dummy_root_item = QTreeWidgetItem([root_id] + [""] * (len(headers) - 1)) # Preenche com vazios
                    self.structure_tree.addTopLevelItem(dummy_root_item)
                    self._add_items_to_tree(dummy_root_item, children_map, root_id, headers, parent_id_col, component_id_col)
                    QMessageBox.warning(self, "Aviso de Estrutura", 
                                        f"O item raiz '{root_id}' foi identificado, mas sua linha de dados completa não foi encontrada para exibição. "
                                        "Exibindo apenas a subestrutura (se houver).")

            self.structure_tree.expandAll() # Expande todos os nós por padrão para visualização completa

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Carregar Estrutura", f"Erro ao carregar dados da estrutura de '{os.path.basename(self.file_path)}' ({current_sheet_name}): {e}")
            self.structure_tree.clear()
            self.structure_tree.addTopLevelItem(QTreeWidgetItem(["Erro", "Erro ao carregar dados. Detalhes: " + str(e)]))

    def _add_items_to_tree(self, parent_qtree_item, children_map, current_item_id, headers, parent_id_col, component_id_col):
        """Adiciona recursivamente itens ao QTreeWidget com base nas relações pai-filho."""
        if current_item_id in children_map:
            for child_entry in children_map[current_item_id]:
                child_id = child_entry["ID"]
                child_row_values = child_entry["Data"]
                
                item_values = [""] * len(headers) # Inicializa com strings vazias para o tamanho correto
                for h_idx, header_name in enumerate(headers):
                    if h_idx < len(child_row_values):
                        item_values[h_idx] = child_row_values[h_idx]
                
                # Garante que o item pai não exiba seu próprio ID de componente como o ID do pai de seus filhos
                # A coluna parent_id_col deve ser vazia para os filhos diretos na exibição hierárquica
                # ou exibir o valor real se preferir, mas geralmente a hierarquia já é mostrada pela árvore.
                # Se desejar ocultar o parent_id na exibição dos filhos, descomente a linha abaixo
                # if parent_id_col != -1 and parent_id_col < len(item_values):
                #     item_values[parent_id_col] = "" 

                q_item = QTreeWidgetItem([str(v) if v is not None else "" for v in item_values])
                parent_qtree_item.addChild(q_item)
                self._add_items_to_tree(q_item, children_map, child_id, headers, parent_id_col, component_id_col)

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Define o caminho para a pasta user_sheets
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    
    # --- Exemplo para bom_data.xlsx ---
    # Cria/Atualiza um arquivo bom_data.xlsx de teste com a sheet "structure"
    test_bom_file_path = os.path.join(user_sheets_dir_test, "bom_data.xlsx") 
    if os.path.exists(test_bom_file_path):
        os.remove(test_bom_file_path) # Garante que começamos com um arquivo limpo para o teste

    wb_test_bom = openpyxl.Workbook()
    ws_bom_structure = wb_test_bom.active # Pega a primeira sheet, renomeia
    ws_bom_structure.title = "EstruturaBOM" # Nome da sheet para este exemplo
    
    # Cabeçalhos para bom_data.xlsx
    bom_headers = ["ParentID", "ComponentID", "ComponentName", "Quantity", "Unit", "Type", "Notes"]
    ws_bom_structure.append(bom_headers) 
    
    # Dados de exemplo para bom_data.xlsx
    ws_bom_structure.append(["", "PROD-A", "Produto Principal A", 1, "EA", "Assembly", "Produto final"])
    ws_bom_structure.append(["PROD-A", "SUB-A1", "Sub-montagem A1", 1, "EA", "Assembly", "Primeira sub-montagem"])
    ws_bom_structure.append(["PROD-A", "PART-X", "Parafuso X", 10, "PCS", "Part", "Parafuso de fixação"])
    ws_bom_structure.append(["SUB-A1", "COMP-C", "Componente C", 2, "EA", "Component", "Componente eletrônico"])
    ws_bom_structure.append(["SUB-A1", "PART-Y", "Arruela Y", 2, "PCS", "Part", "Arruela de pressão"])
    ws_bom_structure.append(["COMP-C", "RAW-Z", "Material Bruto Z", 0.5, "KG", "Raw Material", "Plástico granulado"])
    wb_test_bom.save(test_bom_file_path)
    print(f"Arquivo de teste '{test_bom_file_path}' criado com dados de exemplo.")

    # --- Exemplo para engenharia.xlsx ---
    # Cria/Atualiza um arquivo engenharia.xlsx de teste com a sheet "Estrutura"
    test_engenharia_file_path = os.path.join(user_sheets_dir_test, "engenharia.xlsx")
    if os.path.exists(test_engenharia_file_path):
        os.remove(test_engenharia_file_path) # Garante que começamos com um arquivo limpo para o teste

    wb_engenharia = openpyxl.Workbook()
    ws_engenharia_structure = wb_engenharia.active # Pega a primeira sheet, renomeia
    ws_engenharia_structure.title = "Estrutura" # Nome da sheet para este exemplo
    
    # Cabeçalhos para engenharia.xlsx
    engenharia_headers = ["part_number", "parent_part_number", "quantidade", "materia_prima", "descrição_extra"]
    ws_engenharia_structure.append(engenharia_headers)
    
    # Dados de exemplo para engenharia.xlsx
    ws_engenharia_structure.append(["PROD-001", "", 1, "Não", "Produto final principal da Engenharia"])
    ws_engenharia_structure.append(["ASSY-E1", "PROD-001", 1, "Não", "Montagem Eletrônica"])
    ws_engenharia_structure.append(["COMP-EL-001", "ASSY-E1", 3, "Sim", "Chip de controle"])
    ws_engenharia_structure.append(["COMP-MEC-001", "PROD-001", 1, "Não", "Componente Mecânico"])
    ws_engenharia_structure.append(["RAW-M-001", "COMP-MEC-001", 20, "Sim", "Barra de Alumínio"])
    wb_engenharia.save(test_engenharia_file_path)
    print(f"Arquivo de teste '{test_engenharia_file_path}' criado com dados de exemplo.")

    # --- Instanciação e exibição das ferramentas para teste ---
    # Teste com engenharia.xlsx
    print("\nAbrindo StructureViewTool para engenharia.xlsx (sheet 'Estrutura')...")
    window_engenharia = StructureViewTool(file_path=test_engenharia_file_path, sheet_name="Estrutura")
    window_engenharia.show()

    # Teste com bom_data.xlsx
    print("\nAbrindo StructureViewTool para bom_data.xlsx (sheet 'EstruturaBOM')...")
    window_bom = StructureViewTool(file_path=test_bom_file_path, sheet_name="EstruturaBOM")
    window_bom.show()

    sys.exit(app.exec_())
