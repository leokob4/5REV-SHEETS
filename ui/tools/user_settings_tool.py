import os
import openpyxl
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox
from PyQt5.QtCore import Qt

# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
user_sheets_dir = os.path.join(project_root, 'user_sheets')
DB_EXCEL_PATH = os.path.join(user_sheets_dir, "db.xlsx")

class UserSettingsTool(QWidget):
    """
    Ferramenta para gerenciar as configurações do perfil do usuário.
    Permite visualizar e editar informações como nome completo, email, telefone e departamento.
    As alterações são salvas na planilha 'users' dentro de 'db.xlsx'.
    """
    def __init__(self, user_data):
        super().__init__()
        self.user_data = user_data # Dicionário com os dados do usuário logado
        self.setWindowTitle(f"Perfil do Usuário: {self.user_data.get('username', 'N/A')}")
        self._init_ui()
        self._load_user_profile_data()

    def _init_ui(self):
        """Inicializa a interface do usuário da ferramenta de perfil."""
        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignTop) # Alinha o conteúdo ao topo

        # Título da seção
        title_label = QLabel("<h2>Configurações do Perfil</h2>")
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        main_layout.addSpacing(20)

        # Formulário de entrada de dados
        form_layout = QVBoxLayout()
        form_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter) # Centraliza o formulário horizontalmente

        # Campo: Nome de Usuário (somente leitura)
        username_layout = QHBoxLayout()
        username_layout.addWidget(QLabel("<b>Nome de Usuário:</b>"))
        self.username_display = QLabel(self.user_data.get("username", "N/A"))
        username_layout.addWidget(self.username_display)
        username_layout.addStretch()
        form_layout.addLayout(username_layout)

        # Campo: Papel (somente leitura)
        role_layout = QHBoxLayout()
        role_layout.addWidget(QLabel("<b>Papel:</b>"))
        self.role_display = QLabel(self.user_data.get("role", "N/A"))
        role_layout.addWidget(self.role_display)
        role_layout.addStretch()
        form_layout.addLayout(role_layout)
        form_layout.addSpacing(10)

        # Campos editáveis
        self.full_name_input = self._create_input_field(form_layout, "Nome Completo:", "full_name")
        self.email_input = self._create_input_field(form_layout, "Email:", "email")
        self.phone_input = self._create_input_field(form_layout, "Telefone:", "phone")
        self.department_input = self._create_input_field(form_layout, "Departamento:", "department")
        
        main_layout.addLayout(form_layout)
        main_layout.addSpacing(30)

        # Botão Salvar
        save_button = QPushButton("Salvar Alterações")
        save_button.setFixedSize(200, 40) # Tamanho fixo para o botão
        save_button.clicked.connect(self._save_user_profile_data)
        
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(save_button)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)

        main_layout.addStretch() # Empurra o conteúdo para o topo

    def _create_input_field(self, parent_layout, label_text, attribute_name):
        """Cria um QLabel e QLineEdit para um campo de entrada e o adiciona ao layout."""
        h_layout = QHBoxLayout()
        label = QLabel(f"<b>{label_text}</b>")
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(f"Digite seu {label_text.lower().replace(':', '')}")
        line_edit.setClearButtonEnabled(True) # Adiciona um botão para limpar o campo

        h_layout.addWidget(label)
        h_layout.addWidget(line_edit)
        parent_layout.addLayout(h_layout)
        return line_edit

    def _load_user_profile_data(self):
        """Carrega os dados do perfil do usuário da planilha 'users' para os campos da GUI."""
        try:
            if not os.path.exists(DB_EXCEL_PATH):
                QMessageBox.warning(self, "Erro de Carregamento", "O arquivo de banco de dados 'db.xlsx' não foi encontrado.")
                return

            wb = openpyxl.load_workbook(DB_EXCEL_PATH)
            if "users" not in wb.sheetnames:
                QMessageBox.warning(self, "Erro de Carregamento", "A planilha 'users' não foi encontrada em 'db.xlsx'.")
                return

            sheet = wb["users"]
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            header_map = {h: idx for idx, h in enumerate(headers)}

            # Encontra a linha do usuário logado
            user_row_idx = -1
            username_col_idx = header_map.get("username")
            if username_col_idx is None:
                QMessageBox.warning(self, "Erro de Configuração", "A coluna 'username' não foi encontrada na planilha 'users'.")
                return

            for row_idx in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=username_col_idx + 1).value
                if cell_value == self.user_data.get("username"):
                    user_row_idx = row_idx
                    break
            
            if user_row_idx == -1:
                QMessageBox.warning(self, "Usuário Não Encontrado", f"O usuário '{self.user_data.get('username')}' não foi encontrado na planilha 'users'.")
                return

            # Preenche os campos da GUI com os dados do usuário
            row_values = [sheet.cell(row=user_row_idx, column=col_idx + 1).value for col_idx in range(len(headers))]
            
            self.full_name_input.setText(str(row_values[header_map.get("full_name", -1)]) if header_map.get("full_name", -1) != -1 and header_map["full_name"] < len(row_values) else "")
            self.email_input.setText(str(row_values[header_map.get("email", -1)]) if header_map.get("email", -1) != -1 and header_map["email"] < len(row_values) else "")
            self.phone_input.setText(str(row_values[header_map.get("phone", -1)]) if header_map.get("phone", -1) != -1 and header_map["phone"] < len(row_values) else "")
            self.department_input.setText(str(row_values[header_map.get("department", -1)]) if header_map.get("department", -1) != -1 and header_map["department"] < len(row_values) else "")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Ocorreu um erro ao carregar os dados do perfil: {e}")

    def _save_user_profile_data(self):
        """Salva as alterações do perfil do usuário de volta para a planilha 'users'."""
        try:
            if not os.path.exists(DB_EXCEL_PATH):
                QMessageBox.critical(self, "Erro de Salvamento", "O arquivo de banco de dados 'db.xlsx' não foi encontrado.")
                return

            wb = openpyxl.load_workbook(DB_EXCEL_PATH)
            if "users" not in wb.sheetnames:
                QMessageBox.critical(self, "Erro de Salvamento", "A planilha 'users' não foi encontrada em 'db.xlsx'.")
                return

            sheet = wb["users"]
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            header_map = {h: idx for idx, h in enumerate(headers)}

            # Garante que as colunas essenciais existem ou as adiciona se necessário (para nova instalação ou arquivo corrompido)
            required_profile_headers = ["full_name", "email", "phone", "department"]
            missing_headers = [h for h in required_profile_headers if h not in header_map]

            if missing_headers:
                # Adiciona os cabeçalhos que faltam na primeira linha
                for header in missing_headers:
                    headers.append(header)
                sheet.cell(row=1, column=len(headers)).value = missing_headers[-1] # Adiciona o último missing_header
                header_map = {h: idx for idx, h in enumerate(headers)} # Recria o mapa

            # Encontra a linha do usuário logado
            user_row_idx = -1
            username_col_idx = header_map.get("username")
            if username_col_idx is None:
                QMessageBox.critical(self, "Erro de Configuração", "A coluna 'username' é essencial e não foi encontrada na planilha 'users'. Não é possível salvar.")
                return

            for row_idx in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=username_col_idx + 1).value
                if cell_value == self.user_data.get("username"):
                    user_row_idx = row_idx
                    break
            
            if user_row_idx == -1:
                QMessageBox.critical(self, "Erro de Salvamento", f"O usuário '{self.user_data.get('username')}' não foi encontrado para atualização. As alterações não foram salvas.")
                return

            # Atualiza os valores das células
            sheet.cell(row=user_row_idx, column=header_map.get("full_name", len(headers)) + 1).value = self.full_name_input.text()
            sheet.cell(row=user_row_idx, column=header_map.get("email", len(headers)) + 1).value = self.email_input.text()
            sheet.cell(row=user_row_idx, column=header_map.get("phone", len(headers)) + 1).value = self.phone_input.text()
            sheet.cell(row=user_row_idx, column=header_map.get("department", len(headers)) + 1).value = self.department_input.text()
            
            wb.save(DB_EXCEL_PATH)
            
            # Atualiza os dados na memória (self.user_data) para refletir as mudanças
            self.user_data["full_name"] = self.full_name_input.text()
            self.user_data["email"] = self.email_input.text()
            self.user_data["phone"] = self.phone_input.text()
            self.user_data["department"] = self.department_input.text()

            QMessageBox.information(self, "Sucesso", "Dados do perfil atualizados com sucesso!")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Salvamento", f"Ocorreu um erro ao salvar os dados do perfil: {e}")

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    import bcrypt # Importar bcrypt para o bloco de teste

    app = QApplication(sys.argv)

    # Configura um caminho de teste para db.xlsx para o ambiente de teste da tool
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    db_test_path = os.path.join(user_sheets_dir_test, "db.xlsx")

    # Cria um db.xlsx de teste com a planilha users se não existir
    if not os.path.exists(db_test_path):
        db_wb = openpyxl.Workbook()
        db_ws_users = db_wb.active 
        db_ws_users.title = "users"
        # Adiciona cabeçalhos completos para o teste de perfil
        db_ws_users.append(["id", "username", "password_hash", "role", "full_name", "email", "phone", "department"])
        # Use hashes reais para produção
        db_ws_users.append([1, "admin", bcrypt.hashpw("admin_pass".encode(), bcrypt.gensalt()).decode(), "admin", "Admin Teste", "admin@teste.com", "123456789", "TI"]) 
        db_ws_users.append([2, "user", bcrypt.hashpw("user_pass".encode(), bcrypt.gensalt()).decode(), "user", "Usuario Teste", "user@teste.com", "987654321", "Vendas"])
        db_wb.save(db_test_path)
        print(f"Arquivo de teste db.xlsx criado/atualizado em: {db_test_path}")

    # Dados do usuário logado para simular o uso
    # Este dicionário seria passado pelo LoginWindow para o TeamcenterStyleGUI, e então para a ferramenta
    test_user_data = {
        "id": 1,
        "username": "admin",
        "password_hash": bcrypt.hashpw("admin_pass".encode(), bcrypt.gensalt()).decode(),
        "role": "admin",
        "full_name": "Admin Teste",
        "email": "admin@teste.com",
        "phone": "123456789",
        "department": "TI"
    }

    window = UserSettingsTool(test_user_data)
    window.show()
    sys.exit(app.exec_())
