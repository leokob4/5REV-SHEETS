import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox
from PyQt5.QtCore import Qt

class ExcelViewerTool(QWidget):
    """
    A generic tool for viewing and editing any Excel file (.xlsx) with sheet selection.
    """
    def __init__(self, file_path=None):
        super().__init__()
        self.file_path = file_path
        self.setWindowTitle(f"Visualizador Excel: {os.path.basename(file_path) if file_path else 'Novo Arquivo'}")
        self.layout = QVBoxLayout(self)

        # Header with file name and controls
        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(file_path) if file_path else 'Nenhum Arquivo Carregado'}")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch() # Pushes other elements to the right

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        self.table.setAlternatingRowColors(True)
        self.layout.addWidget(self.table)

        # Control buttons
        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        if self.file_path:
            self._populate_sheet_selector() # Populates and loads data from the first sheet
        else:
            QMessageBox.warning(self, "Nenhum Arquivo", "Nenhum arquivo Excel especificado. Crie ou abra um arquivo.")

    def _populate_sheet_selector(self):
        """Populates the QComboBox with sheet names from the Excel file."""
        self.sheet_selector.clear()
        if not self.file_path or not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", f"O arquivo Excel não foi encontrado ou especificado: {self.file_path}. Crie-o ao salvar.")
            # If file doesn't exist, provide a default "Sheet1" for new file creation
            self.sheet_selector.addItem("Sheet1")
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.setHorizontalHeaderLabels(["Coluna 1", "Coluna 2", "Coluna 3"]) # Default headers for new files
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem("Sheet1") # Always offer default
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. Adicionando a aba padrão 'Sheet1'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                # Select the first sheet by default
                self.sheet_selector.setCurrentIndex(0)

            # Manually trigger _load_data_from_selected_sheet after populating
            self._load_data_from_selected_sheet()

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem("Sheet1") # Fallback to default name if error
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """Loads data from the currently selected Excel sheet into the QTableWidget."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                # This should ideally be handled by _populate_sheet_selector or save logic.
                # If we reach here and file doesn't exist, just clear table.
                self.table.setRowCount(0)
                self.table.setColumnCount(0)
                self.table.setHorizontalHeaderLabels(["Coluna 1", "Coluna 2", "Coluna 3"]) # Default headers
                return

            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                # If sheet doesn't exist (e.g., user typed a new name), create it with headers
                QMessageBox.information(self, "Planilha Não Encontrada", f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. Criando uma nova.")
                ws = wb.create_sheet(current_sheet_name)
                default_headers = ["Coluna 1", "Coluna 2", "Coluna 3"]
                ws.append(default_headers)
                wb.save(self.file_path)
                self._populate_sheet_selector() # Refresh selector to include the new sheet
                return # Exit, as _populate_sheet_selector will trigger a new load

            sheet = wb[current_sheet_name]

            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            data = []
            for row in sheet.iter_rows(min_row=2):
                data.append([cell.value for cell in row])

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.setHorizontalHeaderLabels(["Erro", "Erro", "Erro"])

    def _save_data(self):
        """Saves data from the QTableWidget back to the Excel sheet."""
        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
                # If new file, use headers from current table or a default set
                current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                if not current_headers:
                    current_headers = ["Coluna 1", "Coluna 2", "Coluna 3"]
                ws.append(current_headers)
                wb.save(self.file_path)
                QMessageBox.information(self, "Arquivo e Planilha Criados", f"Novo arquivo '{os.path.basename(self.file_path)}' com planilha '{current_sheet_name}' criado.")
                self._populate_sheet_selector() # Refresh selector to include the new sheet
                # Continue to save data to the newly created sheet
                # Fallthrough to the existing sheet saving logic below
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                    # Use current table headers for new sheet
                    current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                    if not current_headers:
                        current_headers = ["Coluna 1", "Coluna 2", "Coluna 3"]
                    ws.append(current_headers)
                    wb.save(self.file_path)
                    QMessageBox.information(self, "Planilha Criada", f"Nova planilha '{current_sheet_name}' criada em '{os.path.basename(self.file_path)}'.")
                    self._populate_sheet_selector() # Refresh selector to include the new sheet
                    # Continue to save data to the newly created sheet
                    # Fallthrough to the existing sheet saving logic below

            sheet = wb[current_sheet_name]
            
            # Clear existing data but keep header (row 1)
            for row_idx in range(sheet.max_row, 1, -1):
                sheet.delete_rows(row_idx)

            # Write current headers from table (in case they were changed in GUI, though not expected for now)
            current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
            if not current_headers: # If no headers are set in the table (e.g., table was empty on load)
                current_headers = ["Coluna 1", "Coluna 2", "Coluna 3"]
                self.table.setColumnCount(len(current_headers))
                self.table.setHorizontalHeaderLabels(current_headers)
            
            existing_sheet_headers = [cell.value for cell in sheet[1]]
            if existing_sheet_headers != current_headers:
                for col_idx, header_value in enumerate(current_headers):
                    sheet.cell(row=1, column=col_idx + 1, value=header_value)
            
            # Append all rows from the QTableWidget
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                sheet.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados: {e}")

    def _add_empty_row(self):
        """Adds an empty row to the QTableWidget for new data entry."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    # Exemplo: Criar um arquivo temporário para teste
    temp_file_path = os.path.join(os.getcwd(), "temp_excel_file.xlsx")
    
    # Criar um workbook vazio se não existir
    if not os.path.exists(temp_file_path):
        wb = openpyxl.Workbook()
        wb.save(temp_file_path)

    window = ExcelViewerTool(file_path=temp_file_path)
    window.show()
    sys.exit(app.exec_())
    # Opcional: Remover o arquivo temporário após o teste
    # if os.path.exists(temp_file_path):
    #     os.remove(temp_file_path)

