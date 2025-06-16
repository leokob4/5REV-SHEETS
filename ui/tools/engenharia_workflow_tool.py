import sys
import os
import openpyxl
import json # Para serializar/desserializar a lista de conexões
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QHBoxLayout, QMessageBox, QGraphicsView, QGraphicsScene, QComboBox, QLabel, QInputDialog
from PyQt5.QtCore import Qt, QPointF
from PyQt5.QtGui import QBrush, QPen, QColor, QFont # Import QFont

# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
user_sheets_dir = os.path.join(project_root, 'user_sheets')
# O DB_EXCEL_PATH é necessário para buscar os schemas das tools
DB_EXCEL_PATH = os.path.join(user_sheets_dir, "db.xlsx")

DEFAULT_DATA_EXCEL_FILENAME = "engenharia.xlsx"
DEFAULT_SHEET_NAME = "Workflows" # Nome da planilha padrão para salvar/carregar workflows

class EngenhariaWorkflowTool(QWidget):
    """
    GUI para criar, visualizar, salvar e carregar diagramas de fluxo de trabalho.
    Permite adicionar nós de tarefa e ligações de dependência.
    Os dados do diagrama (posições, textos, etc.) são salvos/carregados de engenharia.xlsx.
    Os cabeçalhos para a estrutura de salvamento são carregados dinamicamente de db.xlsx.
    """
    def __init__(self, file_path=None, sheet_name=None):
        super().__init__()
        self.setWindowTitle("Engenharia (Fluxo de Trabalho)")
        
        self.file_path = file_path if file_path else os.path.join(
            user_sheets_dir, DEFAULT_DATA_EXCEL_FILENAME
        )
        self.sheet_name = sheet_name if sheet_name else DEFAULT_SHEET_NAME

        self.layout = QVBoxLayout(self)

        # Controles de arquivo e planilha
        file_sheet_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        file_sheet_layout.addWidget(self.file_name_label)
        file_sheet_layout.addStretch()

        file_sheet_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_workflow_from_selected_sheet)
        file_sheet_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        file_sheet_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(file_sheet_layout)

        self.scene = QGraphicsScene()
        self.view = QGraphicsView(self.scene)
        self.layout.addWidget(self.view)

        # Botões de controle de diagrama
        control_layout = QHBoxLayout()
        add_node_btn = QPushButton("Adicionar Nó de Tarefa")
        add_node_btn.clicked.connect(self._add_task_node)
        add_link_btn = QPushButton("Adicionar Ligação de Dependência")
        add_link_btn.clicked.connect(self._add_dependency_link)
        clear_btn = QPushButton("Limpar Diagrama")
        clear_btn.clicked.connect(self._clear_diagram)
        save_btn = QPushButton("Salvar Workflow")
        save_btn.clicked.connect(self._save_workflow_to_excel)
        load_btn = QPushButton("Recarregar Workflow") # Alterado texto do botão
        load_btn.clicked.connect(self._load_workflow_from_selected_sheet)

        control_layout.addWidget(add_node_btn)
        control_layout.addWidget(add_link_btn)
        control_layout.addWidget(clear_btn)
        control_layout.addWidget(save_btn)
        control_layout.addWidget(load_btn)
        self.layout.addLayout(control_layout)

        self.nodes = [] # Para rastrear os nós adicionados (QGraphicsRectItem)
        self.node_properties = {} # Para armazenar propriedades adicionais dos nós (texto, ID, etc.)
        self.links = [] # Para rastrear as ligações (QGraphicsLineItem)
        self.next_node_id = 1 # Contador para IDs de nós

        # Popula o seletor de planilhas e carrega os dados iniciais
        self._populate_sheet_selector()

    def _get_workflow_schema_headers(self):
        """
        Tenta carregar os cabeçalhos para o esquema de salvamento/carregamento do workflow
        da planilha 'tool_schemas' em db.xlsx.
        Se não encontrar, retorna um conjunto básico e alerta o usuário.
        """
        try:
            if not os.path.exists(DB_EXCEL_PATH):
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Arquivo de banco de dados '{os.path.basename(DB_EXCEL_PATH)}' não encontrado. "
                                    "Usando cabeçalhos padrão muito básicos para Workflow. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["Tipo", "ID", "X", "Y", "Largura", "Altura", "Texto", "Cor", "Conexões"]

            wb = openpyxl.load_workbook(DB_EXCEL_PATH, read_only=True)
            if "tool_schemas" not in wb.sheetnames:
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Planilha 'tool_schemas' não encontrada em '{os.path.basename(DB_EXCEL_PATH)}'. "
                                    "Usando cabeçalhos padrão muito básicos para Workflow. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["Tipo", "ID", "X", "Y", "Largura", "Altura", "Texto", "Cor", "Conexões"]
            
            sheet = wb["tool_schemas"]
            headers_row = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            header_map = {h: idx for idx, h in enumerate(headers_row)}
            
            tool_name_idx = header_map.get("tool_name")
            schema_type_idx = header_map.get("schema_type")
            header_name_idx = header_map.get("header_name")
            order_idx = header_map.get("order")

            if None in [tool_name_idx, schema_type_idx, header_name_idx, order_idx]:
                QMessageBox.warning(self, "Schema Inválido", 
                                    f"Cabeçalhos esperados (tool_name, schema_type, header_name, order) não encontrados na planilha 'tool_schemas'. "
                                    "Usando cabeçalhos padrão muito básicos para Workflow.")
                return ["Tipo", "ID", "X", "Y", "Largura", "Altura", "Texto", "Cor", "Conexões"]

            configured_headers = []
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]
                
                if len(row_values) > max(tool_name_idx, schema_type_idx, header_name_idx, order_idx):
                    tool = row_values[tool_name_idx]
                    schema = row_values[schema_type_idx]
                    header = row_values[header_name_idx]
                    order = row_values[order_idx]

                    if tool == "EngenhariaWorkflowTool" and schema == "workflow_diagram_schema" and header:
                        configured_headers.append((header, order))
            
            configured_headers.sort(key=lambda x: x[1] if x[1] is not None else float('inf'))
            return [h[0] for h in configured_headers]

        except Exception as e:
            QMessageBox.critical(self, "Erro de Configuração", 
                                f"Erro ao carregar configurações de cabeçalho de db.xlsx: {e}. "
                                "Usando cabeçalhos padrão muito básicos para Workflow.")
            return ["Tipo", "ID", "X", "Y", "Largura", "Altura", "Texto", "Cor", "Conexões"]

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        # Se o arquivo não existe, adiciona apenas a sheet padrão e configura um estado inicial
        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", 
                                f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. "
                                f"Ele será criado com a aba padrão '{self.DEFAULT_SHEET_NAME}' ao salvar.")
            self.sheet_selector.addItem(self.DEFAULT_SHEET_NAME)
            self.sheet_selector.setCurrentText(self.DEFAULT_SHEET_NAME) # Define o texto atual
            # Como o arquivo não existe, não há dados para carregar, a cena permanecerá vazia ou com elementos de exemplo.
            self._load_workflow_from_selected_sheet() # Chama para inicializar a tabela mesmo sem arquivo
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(self.DEFAULT_SHEET_NAME)
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", 
                                    f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. "
                                    f"Adicionando a aba padrão '{self.DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                # Tenta selecionar a aba que a ferramenta estava aberta
                default_index = self.sheet_selector.findText(self.sheet_name)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names:
                    self.sheet_selector.setCurrentIndex(0) # Seleciona a primeira sheet disponível
                
            self._load_workflow_from_selected_sheet() # Carrega os dados da aba selecionada

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(self.DEFAULT_SHEET_NAME) # Fallback
            self._clear_diagram() # Limpa o diagrama em caso de erro grave


    def _save_workflow_to_excel(self):
        """
        Salva o estado atual do diagrama para a planilha Excel selecionada.
        Cada nó e link é salvo como uma linha.
        """
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Erro", "Selecione uma planilha para salvar.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                else:
                    ws = wb[current_sheet_name]
            
            # Limpa todas as linhas existentes, exceto a primeira (cabeçalhos)
            # ou todas as linhas se não houver cabeçalhos ainda.
            for row_idx in range(ws.max_row, 0, -1): # Começa do fim, apaga tudo
                ws.delete_rows(row_idx)

            workflow_schema_headers = self._get_workflow_schema_headers()
            ws.append(workflow_schema_headers) # Garante que a primeira linha tem os cabeçalhos corretos

            # Salvar Nós
            for node_item in self.nodes:
                node_id = self.node_properties.get(node_item, {}).get("id")
                node_text = self.node_properties.get(node_item, {}).get("text", "")
                
                # Encontrar a posição correta do QGraphicsTextItem associado
                text_item_found = None
                for item in self.scene.items(node_item.boundingRect()):
                    if hasattr(item, 'toPlainText') and item.pos().x() == node_item.rect().x() + 5: # Basic check for text position
                        text_item_found = item
                        break
                if text_item_found:
                    node_text = text_item_found.toPlainText()

                node_x = node_item.rect().x()
                node_y = node_item.rect().y()
                node_width = node_item.rect().width()
                node_height = node_item.rect().height()
                node_color = node_item.brush().color().name() 
                
                # A coluna "Conexões" para nós será uma lista vazia por enquanto ou JSON de propriedades adicionais.
                connections = [] 
                
                row_data = {
                    "Tipo": "Node",
                    "ID": node_id,
                    "X": node_x,
                    "Y": node_y,
                    "Largura": node_width,
                    "Altura": node_height,
                    "Texto": node_text,
                    "Cor": node_color,
                    "Conexões": json.dumps(connections)
                }
                
                # Garante a ordem correta das colunas ao salvar
                ordered_row = [row_data.get(header, "") for header in workflow_schema_headers]
                ws.append(ordered_row)

            # Salvar Links (simplificado: armazena IDs de nós conectados)
            for link_item in self.links:
                # Recupera os IDs dos nós de origem e destino da propriedade dos links, se armazenados
                source_id = self.node_properties.get(link_item.line().p1(), {}).get("id") # Isso está incorreto, link_item.line().p1() não é o nó. Precisa de uma maneira melhor de rastrear isso.
                target_id = self.node_properties.get(link_item.line().p2(), {}).get("id") # Isso está incorreto.
                
                # PARA SIMPLIFICAR AGORA: Vamos supor que você rastreie as conexões de outra forma
                # ou que esta parte será aprimorada quando a lógica de "Add Link" for completa.
                # Por ora, salvamos um placeholder se não houver IDs de nó disponíveis.
                if hasattr(link_item, 'source_node_id') and hasattr(link_item, 'target_node_id'):
                    link_connections = {"source": link_item.source_node_id, "target": link_item.target_node_id}
                else:
                    link_connections = {"source": "N/A", "target": "N/A"} # Placeholder
                
                row_data = {
                    "Tipo": "Link",
                    "ID": "", # Links podem não ter IDs únicos neste esquema simplificado
                    "X": "", "Y": "", "Largura": "", "Altura": "", "Texto": "", "Cor": "", # Campos vazios para links
                    "Conexões": json.dumps(link_connections)
                }
                ordered_row = [row_data.get(header, "") for header in workflow_schema_headers]
                ws.append(ordered_row)


            wb.save(self.file_path)
            QMessageBox.information(self, "Sucesso", f"Workflow salvo em '{current_sheet_name}' em '{os.path.basename(self.file_path)}'.")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Não foi possível salvar o workflow: {e}")

    def _load_workflow_from_selected_sheet(self):
        """
        Carrega um diagrama de fluxo de trabalho da planilha Excel selecionada.
        """
        self._clear_diagram() # Limpa o diagrama antes de carregar
        current_sheet_name = self.sheet_selector.currentText()
        
        if not current_sheet_name or not os.path.exists(self.file_path):
            # Se o arquivo não existe, ou nenhuma planilha selecionada,
            # garante que a cena está limpa e pode adicionar elementos de exemplo se quiser
            self._add_sample_diagram_elements_if_empty()
            return

        try:
            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                QMessageBox.warning(self, "Planilha Não Encontrada", f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'.")
                self._add_sample_diagram_elements_if_empty() # Adiciona exemplos se a sheet não existir
                return

            sheet = wb[current_sheet_name]
            
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            header_map = {h: idx for idx, h in enumerate(headers)}

            loaded_nodes = {} # Mapeia IDs de nós para os objetos QGraphicsRectItem

            for row_idx in range(2, sheet.max_row + 1): # Começa da segunda linha para pular cabeçalhos
                row_values = [cell.value for cell in sheet[row_idx]]
                
                # Garante que a linha tenha os índices mapeados
                get_value = lambda hdr: row_values[header_map.get(hdr)] if hdr in header_map and header_map[hdr] < len(row_values) else None

                row_type = get_value("Tipo")

                if row_type == "Node":
                    node_id = get_value("ID")
                    x = get_value("X") or 0
                    y = get_value("Y") or 0
                    width = get_value("Largura") or 100
                    height = get_value("Altura") or 50
                    text = str(get_value("Texto") or "")
                    color_name = get_value("Cor") or "lightblue"

                    node_rect = self.scene.addRect(x, y, width, height, QPen(Qt.black), QBrush(QColor(color_name)))
                    node_text_item = self.scene.addText(text) 
                    node_text_item.setPos(x + 5, y + 15) # Posição do texto dentro do nó
                    
                    self.nodes.append(node_rect)
                    self.node_properties[node_rect] = {"id": node_id, "text": text}
                    loaded_nodes[node_id] = node_rect
                    if isinstance(node_id, (int, float)):
                        self.next_node_id = max(self.next_node_id, int(node_id) + 1)


                elif row_type == "Link":
                    link_data_str = get_value("Conexões")
                    try:
                        link_data = json.loads(link_data_str)
                        source_id = link_data.get("source")
                        target_id = link_data.get("target")

                        source_node = loaded_nodes.get(source_id)
                        target_node = loaded_nodes.get(target_id)

                        if source_node and target_node:
                            pen = QPen(Qt.darkGray, 2)
                            line = self.scene.addLine(
                                source_node.rect().x() + source_node.rect().width(), source_node.rect().y() + source_node.rect().height() / 2,
                                target_node.rect().x(), target_node.rect().y() + target_node.rect().height() / 2,
                                pen
                            )
                            # Armazena os IDs dos nós conectados para referência futura (se necessário para edição de links)
                            line.source_node_id = source_id
                            line.target_node_id = target_id
                            self.links.append(line)
                    except json.JSONDecodeError:
                        print(f"Aviso: Dados de conexão inválidos para link: {link_data_str}")

            if not self.nodes and not self.links: # Se nada foi carregado, adiciona exemplos
                self._add_sample_diagram_elements_if_empty()
            else:
                QMessageBox.information(self, "Sucesso", f"Workflow carregado de '{current_sheet_name}' em '{os.path.basename(self.file_path)}'.")


        except Exception as e:
            QMessageBox.critical(self, "Erro ao Carregar", f"Não foi possível carregar o workflow da aba '{current_sheet_name}': {e}")
            self._clear_diagram() # Limpa em caso de erro no carregamento
            self._add_sample_diagram_elements_if_empty() # E adiciona exemplos

    def _add_sample_diagram_elements_if_empty(self):
        """Adiciona alguns elementos de exemplo à cena do diagrama SOMENTE se ela estiver vazia."""
        if not self.nodes and not self.links:
            # Nós de tarefa
            node1_rect = self.scene.addRect(50, 50, 100, 50, QPen(Qt.black), QBrush(QColor("lightblue")))
            node1_text = self.scene.addText("Fase de Design")
            node1_text.setPos(55, 65)
            self.nodes.append(node1_rect)
            self.node_properties[node1_rect] = {"id": f"node_{self.next_node_id}", "text": "Fase de Design", "text_item": node1_text}
            node1_id = self.next_node_id
            self.next_node_id += 1

            node2_rect = self.scene.addRect(200, 150, 100, 50, QPen(Qt.black), QBrush(QColor("lightgreen")))
            node2_text = self.scene.addText("Revisão (Aprovado)")
            node2_text.setPos(205, 165)
            self.nodes.append(node2_rect)
            self.node_properties[node2_rect] = {"id": f"node_{self.next_node_id}", "text": "Revisão (Aprovado)", "text_item": node2_text}
            node2_id = self.next_node_id
            self.next_node_id += 1
            
            node3_rect = self.scene.addRect(350, 50, 100, 50, QPen(Qt.black), QBrush(QColor("lightcoral")))
            node3_text = self.scene.addText("Preparação da Produção")
            node3_text.setPos(355, 65)
            self.nodes.append(node3_rect)
            self.node_properties[node3_rect] = {"id": f"node_{self.next_node_id}", "text": "Preparação da Produção", "text_item": node3_text}
            node3_id = self.next_node_id
            self.next_node_id += 1

            # Ligações/Setas
            pen = QPen(Qt.darkGray)
            pen.setWidth(2)
            
            link1 = self.scene.addLine(node1_rect.x() + node1_rect.rect().width(), node1_rect.y() + node1_rect.rect().height() / 2,
                                       node2_rect.x(), node2_rect.y() + node2_rect.rect().height() / 2, pen)
            link1.source_node_id = f"node_{node1_id}" # Adiciona IDs aos links para salvar
            link1.target_node_id = f"node_{node2_id}"
            self.links.append(link1)

            link2 = self.scene.addLine(node2_rect.x() + node2_rect.rect().width(), node2_rect.y() + node2_rect.rect().height() / 2,
                                       node3_rect.x(), node3_rect.y() + node3_rect.rect().height() / 2, pen)
            link2.source_node_id = f"node_{node2_id}"
            link2.target_node_id = f"node_{node3_id}"
            self.links.append(link2)


    def _add_task_node(self):
        """Adiciona um novo nó de tarefa genérico ao diagrama."""
        node_text, ok = QInputDialog.getText(self, "Novo Nó de Tarefa", "Nome da Tarefa:")
        if not ok or not node_text:
            return

        x = 10 + (len(self.nodes) % 5) * 150 # Deslocamento horizontal
        y = 10 + (len(self.nodes) // 5) * 80  # Deslocamento vertical

        node_rect = self.scene.addRect(x, y, 100, 50, QPen(Qt.black), QBrush(QColor("#FFD700"))) # Cor ouro
        
        node_text_item = self.scene.addText(node_text)
        node_text_item.setPos(x + 5, y + 15) # Posição do texto dentro do nó
        
        new_node_id = f"node_{self.next_node_id}"
        self.nodes.append(node_rect)
        self.node_properties[node_rect] = {"id": new_node_id, "text": node_text, "text_item": node_text_item}
        self.next_node_id += 1

        self.view.centerOn(node_rect)
        QMessageBox.information(self, "Nó Adicionado", f"Nó '{node_text}' adicionado com ID: {new_node_id}.")


    def _add_dependency_link(self):
        """
        Lógica para adicionar uma ligação entre dois nós selecionados.
        Este é um exemplo simplificado, a seleção visual exigiria manipulação de eventos do mouse.
        Por ora, usará caixas de diálogo para obter IDs de nó.
        """
        if len(self.nodes) < 2:
            QMessageBox.warning(self, "Adicionar Ligação", "Você precisa de pelo menos dois nós para criar uma ligação.")
            return

        node_ids = [self.node_properties[n].get("id") for n in self.nodes if self.node_properties.get(n) and self.node_properties[n].get("id")]
        if not node_ids:
            QMessageBox.warning(self, "Erro", "Nenhum ID de nó válido encontrado. Adicione nós primeiro.")
            return

        source_id, ok1 = QInputDialog.getItem(self, "Adicionar Ligação", "Selecione o nó de ORIGEM:", node_ids, 0, False)
        if not ok1: return

        target_id, ok2 = QInputDialog.getItem(self, "Adicionar Ligação", "Selecione o nó de DESTINO:", node_ids, 0, False)
        if not ok2: return

        if source_id == target_id:
            QMessageBox.warning(self, "Erro", "Nós de origem e destino não podem ser os mesmos.")
            return

        source_node_obj = next((n for n in self.nodes if self.node_properties.get(n, {}).get("id") == source_id), None)
        target_node_obj = next((n for n in self.nodes if self.node_properties.get(n, {}).get("id") == target_id), None)

        if not source_node_obj or not target_node_obj:
            QMessageBox.critical(self, "Erro", "Um ou ambos os nós selecionados não foram encontrados.")
            return
        
        # Desenha a linha
        pen = QPen(Qt.darkGray, 2)
        line = self.scene.addLine(
            source_node_obj.rect().x() + source_node_obj.rect().width(), source_node_obj.rect().y() + source_node_obj.rect().height() / 2,
            target_node_obj.x(), target_node_obj.y() + target_node_obj.rect().height() / 2,
            pen
        )
        # Armazena os IDs dos nós conectados no próprio objeto de linha para salvar/carregar
        line.source_node_id = source_id
        line.target_node_id = target_id
        self.links.append(line)
        QMessageBox.information(self, "Ligação Adicionada", f"Ligação criada de '{source_id}' para '{target_id}'.")


    def _clear_diagram(self):
        """Limpa todos os elementos do diagrama e reinicia o contador de IDs."""
        self.scene.clear()
        self.nodes = [] 
        self.node_properties = {}
        self.links = []
        self.next_node_id = 1
        QMessageBox.information(self, "Diagrama Limpo", "O diagrama foi limpo.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Configura um caminho de teste para db.xlsx para o ambiente de teste da tool
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    db_test_path = os.path.join(user_sheets_dir_test, "db.xlsx")

    # Cria/Atualiza um db.xlsx de teste com a planilha tool_schemas se não existir
    import bcrypt # Importar bcrypt para o bloco de teste
    if not os.path.exists(db_test_path):
        db_wb = openpyxl.Workbook()
        db_ws_users = db_wb.active 
        db_ws_users.title = "users"
        db_ws_users.append(["id", "username", "password_hash", "role"])
        db_ws_users.append([1, "admin", bcrypt.hashpw("admin_pass".encode(), bcrypt.gensalt()).decode(), "admin"]) 
        db_ws_users.append([2, "user", bcrypt.hashpw("user_pass".encode(), bcrypt.gensalt()).decode(), "user"])

        db_ws_access = db_wb.create_sheet("access")
        db_ws_access.append(["role", "allowed_modules"])
        db_ws_access.append(["admin", "all"])
        db_ws_access.append(["user", "mod1,mod3,modX,mod4,mod_workflow"]) # Exemplo de módulos permitidos (adicionei mod_workflow)

        db_ws_tools = db_wb.create_sheet("tools")
        db_ws_tools.append(["id", "name", "description", "path"])
        db_ws_tools.append(["mod1", "Gerenciador de BOM", "Gerencia Listas de Materiais", "ui.tools.bom_manager"])
        db_ws_tools.append(["mod3", "Colaboradores", "Gerencia dados de colaboradores", "ui.tools.colaboradores"])
        db_ws_tools.append(["modX", "Configurador", "Gerencia configurações do produto", "ui.tools.configurador"])
        db_ws_tools.append(["mod4", "Engenharia (Dados)", "Gerencia dados de estrutura de engenharia", "ui.tools.engenharia_data"])
        db_ws_tools.append(["mod_workflow", "Engenharia (Fluxo de Trabalho)", "Ferramenta de criação e gerenciamento de diagramas de fluxo de trabalho de engenharia", "ui.tools.engenharia_workflow_tool"]) # Adicionado

        # Adiciona a planilha tool_schemas
        db_ws_schemas = db_wb.create_sheet("tool_schemas")
        db_ws_schemas.append(["tool_name", "schema_type", "header_name", "order"])
        
        # Schemas para BomManagerTool
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do BOM", 1])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do Componente", 2])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Nome do Componente", 3])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Quantidade", 4])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Unidade", 5])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Ref Designator", 6])

        # Schemas para ColaboradoresTool
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "id_colab", 1])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "matricula_colab", 2])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "nome_colab", 3])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_nasc", 4])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_contrat", 5])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_disp", 6])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "setor_colab", 7])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "recurso_colab", 8])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "enabled_colab", 9])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "cpf", 10])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_nascimento", 11])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "endereco", 12])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "telefone", 13])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "email", 14])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "cargo", 15])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "departamento", 16])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_contratacao", 17])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "status_contrato", 18])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "salario_base", 19])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "horas_trabalho_semanais", 20])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "habilidades_principais", 21])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_ultima_avaliacao", 22])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "supervisor", 23])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "turno_trabalho", 24])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "custo_hora_colaborador", 25])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "motivo_saida", 26])

        # Schemas para ConfiguradorTool
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "ID da Configuração", 1])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Nome da Configuração", 2])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Versão", 3])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Descrição", 4])

        # Schemas para EngenhariaDataTool
        db_ws_schemas.append(["EngenhariaDataTool", "default_engenharia_display", "part_number", 1])
        db_ws_schemas.append(["EngenhariaDataTool", "default_engenharia_display", "parent_part_number", 2])
        db_ws_schemas.append(["EngenhariaDataTool", "default_engenharia_display", "quantidade", 3])
        db_ws_schemas.append(["EngenhariaDataTool", "default_engenharia_display", "materia_prima", 4])

        # Schemas para EngenhariaWorkflowTool
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Tipo", 1])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "ID", 2])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "X", 3])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Y", 4])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Largura", 5])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Altura", 6])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Texto", 7])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Cor", 8])
        db_ws_schemas.append(["EngenhariaWorkflowTool", "workflow_diagram_schema", "Conexões", 9])

        db_wb.save(db_test_path)
        print(f"Arquivo de teste db.xlsx criado/atualizado em: {db_test_path}")

    # Cria/Atualiza um arquivo engenharia.xlsx de teste
    # Apenas para garantir que o arquivo existe para que a ferramenta possa tentar carregá-lo
    test_file_path = os.path.join(user_sheets_dir_test, DEFAULT_DATA_EXCEL_FILENAME)
    if not os.path.exists(test_file_path):
        wb_eng = openpyxl.Workbook()
        # Adiciona a aba 'Estrutura' para dados de engenharia
        ws_estrutura = wb_eng.active
        ws_estrutura.title = "Estrutura"
        ws_estrutura.append(["part_number", "parent_part_number", "quantidade", "materia_prima"])
        ws_estrutura.append(["PROD-001", "", 1, "Não"])
        ws_estrutura.append(["COMP-001", "PROD-001", 2, "Não"])
        
        # Adiciona a aba 'Workflows' para dados de workflow de diagrama
        ws_workflow = wb_eng.create_sheet(DEFAULT_SHEET_NAME)
        # Não adiciona headers aqui para forçar a leitura de db.xlsx ou o fallback
        # O _load_workflow_from_selected_sheet cuidará disso
        wb_eng.save(test_file_path)
        print(f"Arquivo de teste {DEFAULT_DATA_EXCEL_FILENAME} criado/atualizado com abas de exemplo.")

    # Testando a ferramenta
    window = EngenhariaWorkflowTool(file_path=test_file_path, sheet_name=DEFAULT_SHEET_NAME)
    window.show()
    sys.exit(app.exec_())
