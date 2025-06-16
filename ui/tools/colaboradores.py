import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox, QInputDialog
from PyQt5.QtCore import Qt

# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
user_sheets_dir = os.path.join(project_root, 'user_sheets')

DEFAULT_DATA_EXCEL_FILENAME = "colaboradores.xlsx"
DEFAULT_SHEET_NAME = "Colaboradores" # Nome da planilha padrão para dados de colaboradores

# COLABORADORES_HEADERS foi removido daqui e será carregado dinamicamente da planilha

class ColaboradoresTool(QWidget):
    """
    GUI para gerenciar dados de colaboradores.
    Permite visualizar, adicionar, editar e salvar informações de colaboradores em planilhas Excel.
    Os cabeçalhos da tabela são carregados EXCLUSIVAMENTE da primeira linha da planilha Excel.
    Se a planilha estiver vazia, os cabeçalhos serão definidos ao adicionar a primeira linha de dados.
    """
    def __init__(self, file_path=None):
        super().__init__()
        if file_path:
            self.file_path = file_path
        else:
            self.file_path = os.path.join(user_sheets_dir, DEFAULT_DATA_EXCEL_FILENAME)

        self.setWindowTitle(f"Colaboradores: {os.path.basename(self.file_path)}")
        self.layout = QVBoxLayout(self)

        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        self._populate_sheet_selector()

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        
        user_sheets_dir = os.path.dirname(self.file_path)
        os.makedirs(user_sheets_dir, exist_ok=True)

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", 
                                f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. "
                                f"Ele será criado com a aba padrão '{DEFAULT_SHEET_NAME}' ao salvar. "
                                "Os cabeçalhos serão definidos ao adicionar e salvar a primeira linha de dados.")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(0) # Inicia com 0 colunas, aguardando cabeçalhos do arquivo ou do usuário
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", 
                                    f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. "
                                    f"Adicionando a aba padrão '{DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                default_index = self.sheet_selector.findText(DEFAULT_SHEET_NAME)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names:
                    self.sheet_selector.setCurrentIndex(0)
                
            self._load_data_from_selected_sheet()

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME) # Fallback para o nome da sheet se der erro
            self.table.setRowCount(0)
            self.table.setColumnCount(0) # Limpa a tabela em caso de erro

    def _load_data_from_selected_sheet(self):
        """Carrega dados da planilha Excel atualmente selecionada para o QTableWidget."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path or not os.path.exists(self.file_path):
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            wb = openpyxl.load_workbook(self.file_path)
            
            if current_sheet_name not in wb.sheetnames:
                QMessageBox.information(self, "Planilha Não Encontrada", 
                                        f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. "
                                        "Criando uma nova. Os cabeçalhos serão definidos ao adicionar e salvar a primeira linha de dados.")
                ws = wb.create_sheet(current_sheet_name)
                wb.save(self.file_path) 
                self._populate_sheet_selector() 
                return 

            sheet = wb[current_sheet_name]

            # Carrega cabeçalhos da primeira linha da planilha
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            if not headers: 
                self.table.setColumnCount(0)
                self.table.setRowCount(0)
                QMessageBox.information(self, "Planilha Vazia", f"A planilha '{current_sheet_name}' está vazia ou não possui cabeçalhos. Adicione uma linha para definir os cabeçalhos.")
                return
            
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            data = []
            for row in sheet.iter_rows(min_row=2):
                row_values = [cell.value for cell in row]
                while len(row_values) < len(headers):
                    row_values.append("")
                data.append(row_values)

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados de colaboradores da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(0) 

    def _save_data(self):
        """Salva dados do QTableWidget de volta para a planilha Excel, capturando os cabeçalhos da tabela."""
        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                else:
                    ws = wb[current_sheet_name]
            
            for row_idx in range(ws.max_row, 0, -1):
                ws.delete_rows(row_idx)

            current_headers = [self.table.horizontalHeaderItem(col).text() 
                               for col in range(self.table.columnCount())]
            
            if current_headers:
                ws.append(current_headers)
            
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                ws.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
            self._populate_sheet_selector() 
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados de colaboradores: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        
        if self.table.columnCount() == 0 and row_count == 0:
            text, ok = QInputDialog.getText(self, "Definir Cabeçalhos", 
                                            "A planilha está vazia. Insira os nomes das colunas separados por vírgula (ex: Matrícula, Nome, Setor):")
            if ok and text:
                headers = [h.strip() for h in text.split(',')]
                self.table.setColumnCount(len(headers))
                self.table.setHorizontalHeaderLabels(headers)
            else:
                QMessageBox.warning(self, "Aviso", "Nenhum cabeçalho fornecido. Nenhuma coluna será adicionada.")
                self.table.removeRow(row_count) 
                return 

        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    
    # Cria/Atualiza um arquivo colaboradores.xlsx de teste com a sheet padrão 'Colaboradores'
    test_file_path = os.path.join(user_sheets_dir_test, DEFAULT_DATA_EXCEL_FILENAME)
    if os.path.exists(test_file_path):
        os.remove(test_file_path) # Garante que começamos com um arquivo limpo para o teste

    # Cria um novo workbook e salva-o com a sheet 'Colaboradores'
    wb_colab = openpyxl.Workbook()
    ws_colab = wb_colab.active # Ativa a primeira sheet
    ws_colab.title = DEFAULT_SHEET_NAME # Define o título como "Colaboradores"
    # Adiciona cabeçalhos e alguns dados de exemplo (ou deixar vazio para testar a criação de headers)
    ws_colab.append(["Matrícula", "Nome", "Setor", "Cargo", "Email"])
    ws_colab.append(["001", "João Silva", "Engenharia", "Engenheiro de Projeto", "joao.silva@empresa.com"])
    ws_colab.append(["002", "Maria Oliveira", "Produção", "Líder de Produção", "maria.oliveira@empresa.com"])
    
    wb_colab.save(test_file_path)
    print(f"Arquivo de teste '{DEFAULT_DATA_EXCEL_FILENAME}' criado/atualizado com abas de exemplo.")

    window = ColaboradoresTool(file_path=test_file_path)
    window.show()
    sys.exit(app.exec_())
