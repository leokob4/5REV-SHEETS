import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox
from PyQt5.QtCore import Qt

# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
user_sheets_dir = os.path.join(project_root, 'user_sheets')
# O DB_EXCEL_PATH é necessário para buscar os schemas das tools
DB_EXCEL_PATH = os.path.join(user_sheets_dir, "db.xlsx")

DEFAULT_DATA_EXCEL_FILENAME = "configurador.xlsx"
DEFAULT_SHEET_NAME = "Configurations"

# CONFIGURATOR_HEADERS foi removido daqui e será carregado dinamicamente
# ou terá um fallback muito básico se o db.xlsx não estiver configurado.

class ConfiguradorTool(QWidget):
    """
    GUI para gerenciar configurações de produto.
    Permite visualizar, adicionar e salvar configurações.
    Os cabeçalhos da tabela são dinamicamente carregados da primeira linha do arquivo Excel
    ou de uma configuração em db.xlsx se a planilha estiver vazia/nova.
    """
    def __init__(self, file_path=None):
        super().__init__()
        if file_path:
            self.file_path = file_path
        else:
            self.file_path = os.path.join(user_sheets_dir, DEFAULT_DATA_EXCEL_FILENAME)

        self.setWindowTitle(f"Configurador: {os.path.basename(self.file_path)}")
        self.layout = QVBoxLayout(self)

        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        self._populate_sheet_selector()

    def _get_default_configurador_headers(self):
        """
        Tenta carregar os cabeçalhos padrão para a exibição de Configurações
        da planilha 'tool_schemas' em db.xlsx.
        Se não encontrar, retorna um conjunto básico e alerta o usuário.
        """
        try:
            if not os.path.exists(DB_EXCEL_PATH):
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Arquivo de banco de dados '{os.path.basename(DB_EXCEL_PATH)}' não encontrado. "
                                    "Usando cabeçalhos padrão muito básicos para Configurador. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["ID da Configuração", "Nome da Configuração", "Versão", "Descrição"]

            wb = openpyxl.load_workbook(DB_EXCEL_PATH, read_only=True)
            if "tool_schemas" not in wb.sheetnames:
                QMessageBox.warning(self, "Configuração Ausente", 
                                    f"Planilha 'tool_schemas' não encontrada em '{os.path.basename(DB_EXCEL_PATH)}'. "
                                    "Usando cabeçalhos padrão muito básicos para Configurador. "
                                    "Por favor, configure 'db.xlsx' e sua planilha 'tool_schemas'.")
                return ["ID da Configuração", "Nome da Configuração", "Versão", "Descrição"]
            
            sheet = wb["tool_schemas"]
            headers_row = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            # Mapeia cabeçalhos para índices para fácil acesso
            header_map = {h: idx for idx, h in enumerate(headers_row)}
            
            tool_name_idx = header_map.get("tool_name")
            schema_type_idx = header_map.get("schema_type")
            header_name_idx = header_map.get("header_name")
            order_idx = header_map.get("order")

            if None in [tool_name_idx, schema_type_idx, header_name_idx, order_idx]:
                QMessageBox.warning(self, "Schema Inválido", 
                                    f"Cabeçalhos esperados (tool_name, schema_type, header_name, order) não encontrados na planilha 'tool_schemas'. "
                                    "Usando cabeçalhos padrão muito básicos para Configurador.")
                return ["ID da Configuração", "Nome da Configuração", "Versão", "Descrição"]

            configured_headers = []
            for row_idx in range(2, sheet.max_row + 1): # Começa da segunda linha para pular cabeçalhos
                row_values = [cell.value for cell in sheet[row_idx]]
                
                # Garante que a linha tem dados suficientes
                if len(row_values) > max(tool_name_idx, schema_type_idx, header_name_idx, order_idx):
                    tool = row_values[tool_name_idx]
                    schema = row_values[schema_type_idx]
                    header = row_values[header_name_idx]
                    order = row_values[order_idx]

                    if tool == "ConfiguradorTool" and schema == "default_configurador_display" and header:
                        configured_headers.append((header, order))
            
            # Ordena os cabeçalhos com base na coluna 'order'
            configured_headers.sort(key=lambda x: x[1] if x[1] is not None else float('inf'))
            return [h[0] for h in configured_headers]

        except Exception as e:
            QMessageBox.critical(self, "Erro de Configuração", 
                                f"Erro ao carregar configurações de cabeçalho de db.xlsx: {e}. "
                                "Usando cabeçalhos padrão muito básicos para Configurador.")
            return ["ID da Configuração", "Nome da Configuração", "Versão", "Descrição"]


    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        
        user_sheets_dir = os.path.dirname(self.file_path)
        os.makedirs(user_sheets_dir, exist_ok=True)

        default_configurador_headers = self._get_default_configurador_headers()

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", 
                                f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. "
                                f"Ele será criado com a aba padrão '{DEFAULT_SHEET_NAME}' e os cabeçalhos definidos ao salvar.")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(len(default_configurador_headers))
            self.table.setHorizontalHeaderLabels(default_configurador_headers)
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", 
                                    f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. "
                                    f"Adicionando a aba padrão '{DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                default_index = self.sheet_selector.findText(DEFAULT_SHEET_NAME)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names:
                    self.sheet_selector.setCurrentIndex(0)
                # else: caso não haja sheets e a padrão não for encontrada, o current index será -1,
                # tratado por _load_data_from_selected_sheet

            self._load_data_from_selected_sheet()

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """Carrega dados da planilha Excel atualmente selecionada para o QTableWidget, usando cabeçalhos reais."""
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            # Tenta carregar os cabeçalhos padrão do DB antes de qualquer outra coisa
            default_configurador_headers = self._get_default_configurador_headers()

            wb = None
            if not os.path.exists(self.file_path):
                # Se o arquivo não existe, inicializa a tabela com os cabeçalhos padrão.
                self.table.setRowCount(0)
                self.table.setColumnCount(len(default_configurador_headers))
                self.table.setHorizontalHeaderLabels(default_configurador_headers)
                return

            wb = openpyxl.load_workbook(self.file_path)
            if current_sheet_name not in wb.sheetnames:
                # Se a planilha não existe, cria-a com cabeçalhos padrão e salva.
                QMessageBox.information(self, "Planilha Não Encontrada", 
                                        f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. "
                                        "Criando uma nova com cabeçalhos padrão.")
                ws = wb.create_sheet(current_sheet_name)
                ws.append(default_configurador_headers)
                wb.save(self.file_path)
                self._populate_sheet_selector() # Recarrega o seletor de abas para incluir a nova
                return

            sheet = wb[current_sheet_name]

            # Carrega os cabeçalhos da primeira linha da planilha existente
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            # Fallback se a planilha estiver completamente vazia (sem cabeçalhos na primeira linha)
            if not headers: 
                headers = default_configurador_headers
            
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            # Carrega os dados a partir da segunda linha
            data = []
            for row in sheet.iter_rows(min_row=2):
                row_values = [cell.value for cell in row]
                # Se a linha for mais curta que os cabeçalhos, preenche com vazios
                while len(row_values) < len(headers):
                    row_values.append("")
                data.append(row_values)

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados de configuração da aba '{current_sheet_name}': {e}")
            # Em caso de erro, define uma tabela vazia com os cabeçalhos padrão como fallback final
            default_configurador_headers = self._get_default_configurador_headers()
            self.table.setRowCount(0)
            self.table.setColumnCount(len(default_configurador_headers)) 
            self.table.setHorizontalHeaderLabels(default_configurador_headers)

    def _save_data(self):
        """Salva dados do QTableWidget de volta para a planilha Excel, mantendo cabeçalhos existentes ou usando padrão."""
        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
                
                # Usa os cabeçalhos da tabela atual ou os padrões se a tabela estiver vazia
                headers_to_save = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                if not headers_to_save:
                    headers_to_save = self._get_default_configurador_headers()
                ws.append(headers_to_save)
                
                wb.save(self.file_path)
                QMessageBox.information(self, "Arquivo e Planilha Criados", f"Novo arquivo '{os.path.basename(self.file_path)}' com planilha '{current_sheet_name}' criado.")
                self._populate_sheet_selector() # Recarregar para mostrar a nova aba
                return
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                    headers_to_save = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
                    if not headers_to_save:
                        headers_to_save = self._get_default_configurador_headers()
                    ws.append(headers_to_save)
                    wb.save(self.file_path)
                    QMessageBox.information(self, "Planilha Criada", f"Nova planilha '{current_sheet_name}' criada em '{os.path.basename(self.file_path)}'.")
                    self._populate_sheet_selector() # Recarregar para mostrar a nova aba

            sheet = wb[current_sheet_name]
            
            # Limpa os dados existentes na planilha, mantendo a primeira linha (cabeçalho)
            for row_idx in range(sheet.max_row, 1, -1):
                sheet.delete_rows(row_idx)

            # Obtém os cabeçalhos da tabela atual para salvar
            current_headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
            
            # Garante que a primeira linha (cabeçalho) da planilha corresponda aos cabeçalhos da tabela
            # Pega os cabeçalhos atuais da planilha (pode ser None se a planilha foi recém-criada ou limpa)
            existing_sheet_headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []

            if existing_sheet_headers != current_headers:
                # Se houver uma linha de cabeçalho existente e ela for diferente, a apaga
                if sheet.max_row > 0 and existing_sheet_headers: 
                    sheet.delete_rows(1)
                # Insere os novos cabeçalhos na primeira linha
                sheet.insert_rows(1)
                # Anexa os cabeçalhos (eles irão para a primeira linha que agora está vazia)
                sheet.append(current_headers) 
            elif not existing_sheet_headers and current_headers: # Caso a planilha estivesse completamente vazia e a tabela tem headers
                sheet.append(current_headers)

            # Percorre o QTableWidget e adiciona as linhas ao Excel
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    row_data.append(item.text() if item is not None else "")
                sheet.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados de configuração: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        for col_idx in range(self.table.columnCount()):
            self.table.setItem(row_count, col_idx, QTableWidgetItem(""))

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Configura um caminho de teste para db.xlsx para o ambiente de teste da tool
    # Ajuste o caminho 'project_root_test' para apontar para a raiz do seu projeto 5REV-SHEETS
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    db_test_path = os.path.join(user_sheets_dir_test, "db.xlsx")

    # Cria/Atualiza um db.xlsx de teste com a planilha tool_schemas se não existir
    # Este bloco é crucial para o teste isolado da ferramenta
    if not os.path.exists(db_test_path):
        db_wb = openpyxl.Workbook()
        db_ws_users = db_wb.active 
        db_ws_users.title = "users"
        db_ws_users.append(["id", "username", "password_hash", "role"])
        # Use hashes reais para produção
        db_ws_users.append([1, "admin", "dummy_admin_hash", "admin"]) 
        db_ws_users.append([2, "user", "dummy_user_hash", "user"])

        db_ws_access = db_wb.create_sheet("access")
        db_ws_access.append(["role", "allowed_modules"])
        db_ws_access.append(["admin", "all"])
        db_ws_access.append(["user", "mod1,mod3,modX"]) # Exemplo de módulos permitidos

        db_ws_tools = db_wb.create_sheet("tools")
        db_ws_tools.append(["id", "name", "description", "path"])
        db_ws_tools.append(["mod1", "Gerenciador de BOM", "Gerencia Listas de Materiais", "ui.tools.bom_manager"])
        db_ws_tools.append(["mod3", "Colaboradores", "Gerencia dados de colaboradores", "ui.tools.colaboradores"])
        db_ws_tools.append(["modX", "Configurador", "Gerencia configurações do produto", "ui.tools.configurador"]) # Exemplo: modX para ConfiguradorTool
        
        # Adiciona a planilha tool_schemas
        db_ws_schemas = db_wb.create_sheet("tool_schemas")
        db_ws_schemas.append(["tool_name", "schema_type", "header_name", "order"])
        
        # Schemas para BomManagerTool
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do BOM", 1])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "ID do Componente", 2])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Nome do Componente", 3])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Quantidade", 4])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Unidade", 5])
        db_ws_schemas.append(["BomManagerTool", "default_bom_display", "Ref Designator", 6])

        # Schemas para ColaboradoresTool
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "id_colab", 1])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "matricula_colab", 2])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "nome_colab", 3])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_nasc", 4])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_contrat", 5])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_disp", 6])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "setor_colab", 7])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "recurso_colab", 8])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "enabled_colab", 9])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "cpf", 10])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_nascimento", 11])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "endereco", 12])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "telefone", 13])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "email", 14])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "cargo", 15])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "departamento", 16])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_contratacao", 17])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "status_contrato", 18])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "salario_base", 19])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "horas_trabalho_semanais", 20])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "habilidades_principais", 21])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "data_ultima_avaliacao", 22])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "supervisor", 23])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "turno_trabalho", 24])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "custo_hora_colaborador", 25])
        db_ws_schemas.append(["ColaboradoresTool", "default_colaboradores_display", "motivo_saida", 26])

        # Schemas para ConfiguradorTool
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "ID da Configuração", 1])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Nome da Configuração", 2])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Versão", 3])
        db_ws_schemas.append(["ConfiguradorTool", "default_configurador_display", "Descrição", 4])


        db_wb.save(db_test_path)
        print(f"Arquivo de teste db.xlsx criado/atualizado em: {db_test_path}")

    # Criar um arquivo configurador.xlsx de teste se não existir, sem cabeçalhos inicialmente para testar carregamento dinâmico
    test_file_path = os.path.join(user_sheets_dir_test, DEFAULT_DATA_EXCEL_FILENAME)
    if not os.path.exists(test_file_path):
        wb_config = openpyxl.Workbook()
        ws_config = wb_config.active
        ws_config.title = DEFAULT_SHEET_NAME
        # Não adiciona headers aqui para forçar a leitura de db.xlsx ou o fallback
        wb_config.save(test_file_path)
        print(f"Arquivo de teste {DEFAULT_DATA_EXCEL_FILENAME} criado vazio para testar carregamento de headers.")

    # Testando a ferramenta
    window = ConfiguradorTool(file_path=test_file_path)
    window.show()
    sys.exit(app.exec_())

