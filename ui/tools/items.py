import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QHeaderView, QLabel, QComboBox, QInputDialog
from PyQt5.QtCore import Qt, QVariant # Importar QVariant para tipos de dados (embora pouco usado diretamente agora)
import datetime # Para validação de data

# Define o nome do arquivo Excel padrão para esta ferramenta
DEFAULT_DATA_EXCEL_FILENAME = "estoque.xlsx" # O nome do arquivo parece ser "estoque.xlsx" para itens/movimentações
DEFAULT_SHEET_NAME = "Movimentacoes" # Nome da planilha padrão alterado para "Movimentacoes" para clareza

# ITEMS_HEADERS FOI REMOVIDO. Os cabeçalhos serão lidos da primeira linha da planilha.

# Mapeamento para tipos de dados esperados para validação (cabeçalho da coluna: tipo)
# Este mapeamento define as REGRAS de validação para colunas COM ESTES NOMES.
# Se um cabeçalho de coluna não estiver nesta lista, o tipo padrão será str.
ITEM_COLUMN_TYPES = {
    "quantidade_movimentada": float, 
    "custo_unitario_movimentacao": float, 
    "data_movimentacao": datetime.date, 
    "validade_lote": datetime.date, 
    "id_movimentacao": int,
    "id_item": int,
    "responsavel_movimentacao": int,
    "saldo_final_deposito": float,
    "estoque_em_transito": float,
    "estoque_disponivel_para_venda": float
}

class ValidatingTableWidgetItem(QTableWidgetItem):
    """
    QTableWidgetItem personalizado que realiza validação básica de tipo
    quando os dados são definidos, com base no nome da coluna e tipo esperado.
    """
    def __init__(self, text="", col_name="", col_type=str):
        super().__init__(text)
        self.col_name = col_name
        self.col_type = col_type
        # Armazena o valor bruto para evitar problemas de formatação na exibição vs. valor real
        self._raw_value = text 

    def data(self, role: int):
        if role == Qt.EditRole:
            # Retorna o valor bruto quando o editor solicita os dados
            return self._raw_value
        return super().data(role)

    def setData(self, role, value):
        if role == Qt.EditRole:
            try:
                # Tenta converter o valor de entrada para o tipo esperado
                converted_value = None
                if self.col_type == int:
                    converted_value = int(value)
                elif self.col_type == float:
                    # Substitui vírgula por ponto para conversão de float
                    converted_value = float(str(value).replace(',', '.'))
                elif self.col_type == datetime.date:
                    if isinstance(value, datetime.date): # Se já for um objeto de data
                        converted_value = value
                    else:
                        date_str = str(value).strip()
                        if not date_str: # Permite datas vazias
                            converted_value = None # Salva como None ou string vazia no Excel
                        else:
                            # Tenta múltiplos formatos de data
                            formats = ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]
                            parsed = False
                            for fmt in formats:
                                try:
                                    converted_value = datetime.datetime.strptime(date_str, fmt).date()
                                    parsed = True
                                    break
                                except ValueError:
                                    continue
                            if not parsed:
                                raise ValueError(f"Formato de data inválido para '{self.col_name}': '{value}'. Esperado um formato como YYYY-MM-DD ou DD/MM/YYYY.")
                else: # Padrão para string para outros tipos
                    converted_value = str(value)
                
                # Armazena o valor bruto convertido
                self._raw_value = converted_value

                # Define o texto exibido na célula
                value_to_display = ""
                if converted_value is not None:
                    if self.col_type == float:
                        # Exibe float com vírgula como separador decimal
                        value_to_display = str(converted_value).replace('.', ',')
                    elif self.col_type == datetime.date:
                        # Exibe data formatada como DD/MM/AAAA
                        value_to_display = converted_value.strftime("%d/%m/%Y")
                    else:
                        value_to_display = str(converted_value)
                
                super().setData(role, value_to_display)
                self.setToolTip("") # Limpa qualquer dica de ferramenta de erro anterior
            except ValueError as e:
                # Se a validação falhar, exibe uma mensagem e reverte o valor na célula
                QMessageBox.warning(self.tableWidget().parentWidget(), "Erro de Validação", 
                                    f"Valor inválido para a coluna '{self.col_name}': '{value}'. "
                                    f"Esperado tipo '{self.col_type.__name__}'. Detalhes: {e}")
                self.setToolTip(f"Erro: Esperado {self.col_type.__name__}. Inserido: '{value}'")
                # Não chama super().setData(role, self.text()) para não sobrescrever o valor antigo
                # e manter o valor inválido visualmente para o usuário se preferir
                # ou pode reverter para o valor anterior (self._raw_value)
                super().setData(role, str(self._raw_value) if self._raw_value is not None else "") # Reverte para o último valor válido
            except Exception as e:
                QMessageBox.critical(self.tableWidget().parentWidget(), "Erro Inesperado", f"Um erro inesperado ocorreu na validação: {e}")
                super().setData(role, str(self._raw_value) if self._raw_value is not None else "")
        else:
            super().setData(role, value)

class ItemsTool(QWidget):
    """
    GUI para gerenciar movimentações de estoque.
    Permite visualizar, adicionar, editar e salvar informações de estoque em 'estoque.xlsx'.
    Os cabeçalhos da tabela são carregados EXCLUSIVAMENTE da primeira linha do arquivo Excel.
    Se a planilha estiver vazia, os cabeçalhos serão definidos pelo usuário ao adicionar a primeira linha.
    Pode operar em modo somente leitura se o arquivo for 'engenharia.xlsx' ou explicitamente definido.
    Inclui validação de input para tipos de dados usando ValidatingTableWidgetItem.
    """
    def __init__(self, file_path=None, read_only=False): 
        super().__init__()
        # Definindo caminhos de forma dinâmica
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(current_dir)) # Navega de ui/tools para a raiz do projeto
        user_sheets_dir = os.path.join(project_root, 'user_sheets')

        if file_path:
            self.file_path = file_path
        else:
            self.file_path = os.path.join(user_sheets_dir, DEFAULT_DATA_EXCEL_FILENAME)

        # Força somente leitura se o arquivo passado for especificamente 'engenharia.xlsx'
        # ou se 'read_only' for True.
        self.is_read_only = read_only or (os.path.basename(self.file_path).lower() == "engenharia.xlsx") 

        self.setWindowTitle(f"Movimentações de Estoque: {os.path.basename(self.file_path)}")
        if self.is_read_only:
            self.setWindowTitle(self.windowTitle() + " (Somente Leitura)")

        self.layout = QVBoxLayout(self)

        header_layout = QHBoxLayout()
        self.file_name_label = QLabel(f"<b>Arquivo:</b> {os.path.basename(self.file_path)}")
        if self.is_read_only:
            self.file_name_label.setText(self.file_name_label.text() + " (Somente Leitura)")
        header_layout.addWidget(self.file_name_label)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Planilha:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setMinimumWidth(150)
        self.sheet_selector.currentIndexChanged.connect(self._load_data_from_selected_sheet)
        header_layout.addWidget(self.sheet_selector)

        self.refresh_sheets_btn = QPushButton("Atualizar Abas")
        self.refresh_sheets_btn.clicked.connect(self._populate_sheet_selector)
        header_layout.addWidget(self.refresh_sheets_btn)
        self.layout.addLayout(header_layout)

        self.table = QTableWidget()
        # Define o trigger de edição com base no modo somente leitura
        if self.is_read_only:
            self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        else:
            self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.layout.addWidget(self.table)

        button_layout = QHBoxLayout()
        self.add_row_btn = QPushButton("Adicionar Linha")
        self.add_row_btn.clicked.connect(self._add_empty_row)
        self.save_btn = QPushButton("Salvar Dados")
        self.save_btn.clicked.connect(self._save_data)
        self.refresh_btn = QPushButton("Recarregar Dados da Aba Atual")
        self.refresh_btn.clicked.connect(self._load_data_from_selected_sheet)

        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(button_layout)

        # Desabilita botões de edição/salvamento se estiver em modo somente leitura
        if self.is_read_only:
            self.add_row_btn.setEnabled(False)
            self.save_btn.setEnabled(False)
            QMessageBox.information(self, "Modo Somente Leitura", f"A ferramenta está operando em modo somente leitura para {os.path.basename(self.file_path)}. Edições não são permitidas.")

        self._populate_sheet_selector()

    def _populate_sheet_selector(self):
        """Popula o QComboBox com os nomes das planilhas do arquivo Excel."""
        self.sheet_selector.clear()
        
        user_sheets_dir = os.path.dirname(self.file_path)
        os.makedirs(user_sheets_dir, exist_ok=True)

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Arquivo Não Encontrado", 
                                f"O arquivo de dados não foi encontrado: {os.path.basename(self.file_path)}. "
                                f"Ele será criado com a aba padrão '{DEFAULT_SHEET_NAME}' ao salvar. "
                                "Os cabeçalhos serão definidos ao adicionar e salvar a primeira linha de dados.")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
            self.table.setRowCount(0)
            self.table.setColumnCount(0) # Inicia com 0 colunas, aguardando cabeçalhos do arquivo ou do usuário
            return

        try:
            # Abre em modo somente leitura para listar planilhas para evitar travar o arquivo
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                self.sheet_selector.addItem(DEFAULT_SHEET_NAME)
                QMessageBox.warning(self, "Nenhuma Planilha Encontrada", 
                                    f"Nenhuma planilha encontrada em '{os.path.basename(self.file_path)}'. "
                                    f"Adicionando a aba padrão '{DEFAULT_SHEET_NAME}'.")
            else:
                for sheet_name in sheet_names:
                    self.sheet_selector.addItem(sheet_name)
                
                # Prioriza a sheet padrão se existir, caso contrário, seleciona a primeira
                default_index = self.sheet_selector.findText(DEFAULT_SHEET_NAME)
                if default_index != -1:
                    self.sheet_selector.setCurrentIndex(default_index)
                elif sheet_names: # Se a padrão não for encontrada, mas outras sheets existirem, seleciona a primeira
                    self.sheet_selector.setCurrentIndex(0)
                
            self._load_data_from_selected_sheet() # Aciona o carregamento de dados da planilha selecionada (ou a padrão/primeira)

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Listar Planilhas", f"Erro ao listar planilhas em '{os.path.basename(self.file_path)}': {e}")
            self.sheet_selector.addItem(DEFAULT_SHEET_NAME) # Fallback
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def _load_data_from_selected_sheet(self):
        """
        Carrega dados da planilha Excel atualmente selecionada para o QTableWidget.
        Os cabeçalhos são lidos da primeira linha da planilha.
        """
        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name or not self.file_path or not os.path.exists(self.file_path):
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        try:
            # Sempre carregamos o workbook em modo de escrita/leitura para poder editar e salvar.
            # O controle de somente leitura é feito a nível de GUI (desabilitando botões e triggers de edição).
            wb = openpyxl.load_workbook(self.file_path)
            
            if current_sheet_name not in wb.sheetnames:
                QMessageBox.information(self, "Planilha Não Encontrada", 
                                        f"A planilha '{current_sheet_name}' não foi encontrada em '{os.path.basename(self.file_path)}'. "
                                        "Criando uma nova. Os cabeçalhos serão definidos ao adicionar e salvar a primeira linha de dados.")
                ws = wb.create_sheet(current_sheet_name)
                wb.save(self.file_path) # Salva a nova sheet vazia para que apareça no seletor
                self._populate_sheet_selector() # Recarrega o seletor para incluir a nova planilha
                return # Retorna para carregar a sheet recém-criada, que estará vazia

            sheet = wb[current_sheet_name]

            # Carrega cabeçalhos da primeira linha da planilha
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            
            if not headers: 
                self.table.setColumnCount(0)
                self.table.setRowCount(0)
                QMessageBox.information(self, "Planilha Vazia", 
                                        f"A planilha '{current_sheet_name}' está vazia ou não possui cabeçalhos. "
                                        "Adicione uma linha para definir os cabeçalhos.")
                return
            
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)

            data = []
            for row in sheet.iter_rows(min_row=2): # Começa da segunda linha para os dados
                row_values = [cell.value for cell in row]
                # Preenche com vazios se a linha for mais curta que o número de cabeçalhos
                while len(row_values) < len(headers):
                    row_values.append("")
                data.append(row_values)

            self.table.setRowCount(len(data))
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    header_name = headers[col_idx]
                    # Tenta obter o tipo da coluna, padrão para string se não estiver no mapeamento
                    col_type = ITEM_COLUMN_TYPES.get(header_name, str) 
                    
                    # Cria o ValidatingTableWidgetItem para cada célula
                    item = ValidatingTableWidgetItem(str(cell_value) if cell_value is not None else "", header_name, col_type)
                    
                    # Para tipos como datetime.date, certifique-se de que o valor bruto seja o objeto de data
                    if col_type == datetime.date and isinstance(cell_value, datetime.datetime):
                        item._raw_value = cell_value.date()
                        item.setText(cell_value.strftime("%d/%m/%Y")) # Formato de exibição
                    elif col_type == float and isinstance(cell_value, (int, float)):
                        item._raw_value = float(cell_value)
                        item.setText(str(float(cell_value)).replace('.', ',')) # Formato de exibição
                    else:
                        item._raw_value = cell_value # Armazena o valor bruto
                        item.setText(str(cell_value) if cell_value is not None else "") # Exibe como string
                        

                    self.table.setItem(row_idx, col_idx, item)

            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
            QMessageBox.information(self, "Dados Carregados", f"Dados de '{current_sheet_name}' carregados com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Carregamento", f"Erro ao carregar dados de itens da aba '{current_sheet_name}': {e}")
            self.table.setRowCount(0)
            self.table.setColumnCount(0) 

    def _save_data(self):
        """
        Salva dados do QTableWidget de volta para a planilha Excel, capturando os cabeçalhos da tabela.
        Respeita o modo somente leitura.
        """
        if self.is_read_only: 
            QMessageBox.warning(self, "Ação Não Permitida", "Esta ferramenta está em modo somente leitura. Não é possível salvar alterações.")
            return
            
        if not self.file_path:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo especificado para salvar.")
            return

        current_sheet_name = self.sheet_selector.currentText()
        if not current_sheet_name:
            QMessageBox.warning(self, "Nome da Planilha Inválido", "O nome da planilha não pode estar vazio. Por favor, selecione ou adicione uma aba.")
            return

        try:
            wb = None
            if not os.path.exists(self.file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = current_sheet_name
            else:
                wb = openpyxl.load_workbook(self.file_path)
                if current_sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(current_sheet_name)
                else:
                    ws = wb[current_sheet_name]
            
            # Limpa todas as linhas existentes na planilha
            for row_idx in range(ws.max_row, 0, -1):
                ws.delete_rows(row_idx)

            # Obtém os cabeçalhos atuais da QTableWidget.
            current_headers = [self.table.horizontalHeaderItem(col).text() 
                               for col in range(self.table.columnCount())]
            
            # Salva os cabeçalhos se existirem na tabela
            if current_headers:
                ws.append(current_headers)
            
            # Percorre o QTableWidget e adiciona as linhas ao Excel
            for row_idx in range(self.table.rowCount()):
                row_data = []
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    
                    # Obtém o valor bruto armazenado no ValidatingTableWidgetItem
                    cell_value_to_save = item._raw_value if item else ""
                    
                    # Se o valor é um objeto datetime.date, salva-o como tal para openpyxl
                    if isinstance(cell_value_to_save, datetime.date):
                        row_data.append(cell_value_to_save)
                    else:
                        row_data.append(str(cell_value_to_save) if cell_value_to_save is not None else "")
                ws.append(row_data)

            wb.save(self.file_path)
            QMessageBox.information(self, "Dados Salvos", f"Dados de '{current_sheet_name}' salvos com sucesso em '{os.path.basename(self.file_path)}'.")
            self._populate_sheet_selector() # Recarrega para garantir que o seletor esteja atualizado
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Salvar", f"Erro ao salvar dados de itens: {e}")

    def _add_empty_row(self):
        """Adiciona uma linha vazia ao QTableWidget para nova entrada de dados."""
        if self.is_read_only: 
            QMessageBox.warning(self, "Ação Não Permitida", "Esta ferramenta está em modo somente leitura. Não é possível adicionar linhas.")
            return

        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        
        # Se a tabela ainda não tem colunas (ex: planilha nova/vazia),
        # esta é a primeira linha, e o usuário precisa definir os cabeçalhos.
        if self.table.columnCount() == 0 and row_count == 0:
            text, ok = QInputDialog.getText(self, "Definir Cabeçalhos", 
                                            "A planilha está vazia. Insira os nomes das colunas separados por vírgula (ex: part_number, quantidade_movimentada, data_movimentacao):")
            if ok and text:
                headers = [h.strip() for h in text.split(',')]
                self.table.setColumnCount(len(headers))
                self.table.setHorizontalHeaderLabels(headers)
            else:
                QMessageBox.warning(self, "Aviso", "Nenhum cabeçalho fornecido. Nenhuma coluna será adicionada.")
                self.table.removeRow(row_count) # Remove a linha vazia recém-adicionada
                return # Sai da função se o usuário cancelar ou não fornecer cabeçalhos

        # Preenche a nova linha com instâncias ValidatingTableWidgetItem
        for col_idx in range(self.table.columnCount()):
            header_name = self.table.horizontalHeaderItem(col_idx).text()
            # Obtém o tipo da coluna, padrão para string se não estiver no mapeamento
            col_type = ITEM_COLUMN_TYPES.get(header_name, str) 
            self.table.setItem(row_count, col_idx, ValidatingTableWidgetItem("", header_name, col_type))

# Exemplo de uso (para testar este módulo individualmente)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    project_root_test = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    user_sheets_dir_test = os.path.join(project_root_test, 'user_sheets')
    os.makedirs(user_sheets_dir_test, exist_ok=True)
    
    # Cria/Atualiza um arquivo estoque.xlsx de teste com a sheet padrão 'Movimentacoes'
    test_file_path = os.path.join(user_sheets_dir_test, DEFAULT_DATA_EXCEL_FILENAME)
    if os.path.exists(test_file_path):
        os.remove(test_file_path) # Garante que começamos com um arquivo limpo para o teste

    # Cria um novo workbook e salva-o com a sheet 'Movimentacoes'
    wb_items = openpyxl.Workbook()
    ws_items = wb_items.active # Ativa a primeira sheet
    ws_items.title = DEFAULT_SHEET_NAME # Define o título como "Movimentacoes"
    
    # Define os cabeçalhos para o arquivo de teste (eles serão lidos na primeira linha)
    test_headers = [
        "part_number", "id_movimentacao", "data_movimentacao", "id_item",
        "tipo_movimentacao", "quantidade_movimentada", "deposito_origem",
        "deposito_destino", "lote_item", "validade_lote",
        "custo_unitario_movimentacao", "referencia_documento",
        "responsavel_movimentacao", "saldo_final_deposito", "motivo_ajuste",
        "status_inspecao_recebimento", "posicao_estoque_fisica",
        "reserva_para_ordem_producao", "reserva_para_pedido_venda",
        "estoque_em_transito", "estoque_disponivel_para_venda"
    ]
    ws_items.append(test_headers)
    
    # Adiciona alguns dados de exemplo (com tipos corretos)
    ws_items.append(["PART-001", 101, datetime.date(2024, 6, 15), 1, "Entrada", 50.5, "DP1", "DP1", "LOTE-XYZ", datetime.date(2025, 12, 31), 10.25, "NF-123", 123, 150.0, "", "Aprovado", "", "", 0.0, 150.0])
    ws_items.append(["PART-002", 102, datetime.date(2024, 6, 16), 2, "Saída", 25.0, "DP1", "DP2", "LOTE-ABC", None, 5.50, "OP-456", 456, 125.0, "Venda", "Aprovado", "OP-1", "", 0.0, 125.0])
    
    wb_items.save(test_file_path)
    print(f"Arquivo de teste '{DEFAULT_DATA_EXCEL_FILENAME}' criado/atualizado com abas e dados de exemplo.")

    window = ItemsTool(file_path=test_file_path)
    window.show()
    sys.exit(app.exec_())
