import os
import openpyxl

# Define os caminhos dos diretórios relativos à raiz do projeto
# O script assume que será executado da raiz do projeto 5REV-SHEETS
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USER_SHEETS_DIR = os.path.join(PROJECT_ROOT, "user_sheets")
APP_SHEETS_DIR = os.path.join(PROJECT_ROOT, "app_sheets")
DB_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "db.xlsx")

def get_db_db_schema():
    """
    Carrega o schema esperado da planilha 'db_db' em db.xlsx.
    Retorna um dicionário aninhado:
    {
        'caminho/do/arquivo.xlsx': {
            'NomeDaPlanilha': {'Header1', 'Header2', ...} # Set de nomes de cabeçalho
        }
    }
    """
    expected_schema = {}
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            print(f"Erro: O arquivo db.xlsx não foi encontrado em: {DB_EXCEL_PATH}")
            return expected_schema

        wb = openpyxl.load_workbook(DB_EXCEL_PATH)
        if "db_db" not in wb.sheetnames:
            print(f"Erro: A planilha 'db_db' não foi encontrada em {DB_EXCEL_PATH}")
            return expected_schema

        sheet = wb["db_db"]
        headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
        
        # Mapeia os índices das colunas relevantes
        header_map = {
            "Arquivo (Caminho)": -1,
            "Nome da Coluna (Cabeçalho)": -1,
            "pagina_arquivo": -1 # Usaremos este para o nome da planilha
        }
        for idx, h in enumerate(headers):
            if h in header_map:
                header_map[h] = idx

        if any(idx == -1 for key, idx in header_map.items() if key != "descr_variavel"): # 'descr_variavel' é opcional para esta função
            print(f"Aviso: A planilha 'db_db' em {DB_EXCEL_PATH} não possui os cabeçalhos essenciais: {list(header_map.keys())}. "
                  "O carregamento pode estar incompleto.")
            
        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            file_path_raw = row_values[header_map["Arquivo (Caminho)"]] if header_map["Arquivo (Caminho)"] != -1 and header_map["Arquivo (Caminho)"] < len(row_values) else None
            column_name = row_values[header_map["Nome da Coluna (Cabeçalho)"]] if header_map["Nome da Coluna (Cabeçalho)"] != -1 and header_map["Nome da Coluna (Cabeçalho)"] < len(row_values) else None
            sheet_name_from_db = row_values[header_map["pagina_arquivo"]] if header_map["pagina_arquivo"] != -1 and header_map["pagina_arquivo"] < len(row_values) else None

            if file_path_raw and column_name and sheet_name_from_db:
                normalized_file_path = file_path_raw.replace('\\', '/') # Normaliza para consistência
                
                if normalized_file_path not in expected_schema:
                    expected_schema[normalized_file_path] = {}
                
                if sheet_name_from_db not in expected_schema[normalized_file_path]:
                     expected_schema[normalized_file_path][sheet_name_from_db] = set()

                expected_schema[normalized_file_path][sheet_name_from_db].add(str(column_name))
    except Exception as e:
        print(f"Erro ao carregar schema de db_db: {e}")
    return expected_schema

def get_actual_sheet_headers(directory_path, base_path_for_rel):
    """
    Coleta os cabeçalhos reais de todas as planilhas Excel em um determinado diretório.
    Retorna um dicionário aninhado:
    {
        'caminho/do/arquivo.xlsx': {
            'NomeDaPlanilha': {'Header1', 'Header2', ...} # Set de nomes de cabeçalho
        }
    }
    """
    actual_headers = {}
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_full_path = os.path.join(directory_path, filename)
            file_rel_path = os.path.relpath(file_full_path, base_path_for_rel).replace('\\', '/')
            try:
                wb = openpyxl.load_workbook(file_full_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    if sheet.max_row > 0:
                        headers = [cell.value for cell in sheet[1] if cell.value is not None]
                        if file_rel_path not in actual_headers:
                            actual_headers[file_rel_path] = {}
                        actual_headers[file_rel_path][sheet_name] = {str(h) for h in headers} # Usa um set para facilitar comparação
            except Exception as e:
                print(f"Aviso: Não foi possível ler o arquivo {filename} ou uma de suas planilhas: {e}")
    return actual_headers

def validate_db_db_consistency():
    """
    Verifica a consistência dos dados de db_db com os cabeçalhos reais das outras planilhas.
    Retorna um dicionário de diferenças.
    """
    print("\n--- Iniciando validação de consistência de db_db ---")
    
    expected_schema = get_db_db_schema()
    if not expected_schema:
        print("Nenhum schema de db_db carregado. Validação não pode prosseguir.")
        return {"error": "No db_db schema loaded."}

    all_actual_headers = {}
    all_actual_headers.update(get_actual_sheet_headers(USER_SHEETS_DIR, PROJECT_ROOT))
    all_actual_headers.update(get_actual_sheet_headers(APP_SHEETS_DIR, PROJECT_ROOT))

    differences = {
        "missing_files_in_actual": [], # Arquivos/planilhas em db_db mas não no sistema de arquivos
        "extra_files_in_actual": [],   # Arquivos/planilhas no sistema de arquivos mas não em db_db
        "header_mismatches": []        # Discrepâncias de cabeçalho dentro das planilhas mapeadas
    }

    # 1. Verificar arquivos e planilhas que estão em db_db mas não existem no sistema de arquivos ou não contêm a sheet mapeada
    for db_file_rel_path, db_sheets_data in expected_schema.items():
        for db_sheet_name, db_expected_headers_set in db_sheets_data.items():
            if db_file_rel_path not in all_actual_headers or db_sheet_name not in all_actual_headers[db_file_rel_path]:
                differences["missing_files_in_actual"].append(
                    f"'{db_file_rel_path}' -> Planilha: '{db_sheet_name}' (Mapeado em db_db, mas não encontrado no sistema de arquivos/planilha)."
                )

    # 2. Verificar arquivos e planilhas que estão no sistema de arquivos mas não em db_db
    for actual_file_rel_path, actual_sheets_data in all_actual_headers.items():
        for actual_sheet_name, actual_headers_set in actual_sheets_data.items():
            is_mapped_in_db_db = False
            if actual_file_rel_path in expected_schema and actual_sheet_name in expected_schema[actual_file_rel_path]:
                is_mapped_in_db_db = True
            
            if not is_mapped_in_db_db:
                differences["extra_files_in_actual"].append(
                    f"'{actual_file_rel_path}' -> Planilha: '{actual_sheet_name}' (Encontrado no sistema de arquivos, mas não mapeado em db_db)."
                )

    # 3. Verificar discrepâncias de cabeçalho para planilhas mapeadas
    for db_file_rel_path, db_sheets_data in expected_schema.items():
        for db_sheet_name, db_expected_headers_set in db_sheets_data.items():
            if db_file_rel_path in all_actual_headers and db_sheet_name in all_actual_headers[db_file_rel_path]:
                actual_headers_set = all_actual_headers[db_file_rel_path][db_sheet_name]

                missing_in_actual = db_expected_headers_set - actual_headers_set
                extra_in_actual = actual_headers_set - db_expected_headers_set

                if missing_in_actual or extra_in_actual:
                    diff_detail = {
                        "file": db_file_rel_path,
                        "sheet": db_sheet_name,
                        "missing_headers": sorted(list(missing_in_actual)),
                        "extra_headers": sorted(list(extra_in_actual))
                    }
                    differences["header_mismatches"].append(diff_detail)
    
    # Imprimir o relatório
    if not any(differences.values()):
        print("\n✅ Todos os arquivos e cabeçalhos estão consistentes com 'db_db'.")
    else:
        print("\n--- Relatório de Diferenças de Consistência ---")
        if differences["missing_files_in_actual"]:
            print("\n❌ Arquivos/planilhas mapeados em 'db_db' mas ausentes no sistema:")
            for msg in differences["missing_files_in_actual"]:
                print(f"  - {msg}")
        
        if differences["extra_files_in_actual"]:
            print("\n❓ Arquivos/planilhas presentes no sistema mas não mapeados em 'db_db':")
            for msg in differences["extra_files_in_actual"]:
                print(f"  - {msg}")

        if differences["header_mismatches"]:
            print("\n⚠️ Discrepâncias de cabeçalho (db_db vs. Planilha Real):")
            for diff in differences["header_mismatches"]:
                print(f"  Arquivo: '{diff['file']}' -> Planilha: '{diff['sheet']}'")
                if diff['missing_headers']:
                    print(f"    - Ausentes na planilha: {', '.join(diff['missing_headers'])}")
                if diff['extra_headers']:
                    print(f"    - Extras na planilha: {', '.join(diff['extra_headers'])}")

    print("\n--- Validação concluída ---")
    return differences

if __name__ == "__main__":
    validate_db_db_consistency()
