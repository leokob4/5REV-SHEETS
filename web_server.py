import os
from flask import Flask, send_from_directory, request, jsonify, redirect, url_for
import bcrypt
import openpyxl
from functools import wraps

# --- Flask App Setup ---
app = Flask(__name__, static_folder='js') # Serve static files from 'js' directory

# --- File Paths Configuration ---
USER_SHEETS_DIR = "user_sheets"
APP_SHEETS_DIR = "app_sheets"
DB_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "db.xlsx")
TOOLS_EXCEL_PATH = os.path.join(APP_SHEETS_DIR, "tools.xlsx")

# Ensure directories exist
os.makedirs(USER_SHEETS_DIR, exist_ok=True)
os.makedirs(APP_SHEETS_DIR, exist_ok=True)

# --- Basic Authentication for Web (Placeholder) ---
# In a real application, this token would be generated securely (e.g., JWT)
# and validated with more sophistication (e.g., signature check, expiration).
# For this example, we'll just check for its presence.
VALID_AUTH_TOKEN = "fake-jwt-token-123"

def login_required(f):
    """
    Decorator to ensure a route requires authentication.
    Checks for a simple 'Authorization' header with a valid token.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header or auth_header.split(" ")[1] != VALID_AUTH_TOKEN:
            return jsonify({"message": "Authentication required."}), 401
        return f(*args, **kwargs)
    return decorated_function

# --- Sheet Helpers ---
def load_users_from_excel_backend():
    """Loads user data from the database Excel file for backend use."""
    users = {}
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            return {}

        wb = openpyxl.load_workbook(DB_EXCEL_PATH)
        if 'users' not in wb.sheetnames:
            return {}

        users_sheet = wb["users"]
        # Iterate from the second row to skip headers
        for row in users_sheet.iter_rows(min_row=2):
            if len(row) >= 4 and all(cell.value is not None for cell in row[:4]):
                users[row[1].value] = { # username as key
                    "id": row[0].value,
                    "username": row[1].value,
                    "password_hash": row[2].value,
                    "role": row[3].value
                }
    except Exception as e:
        print(f"Error loading users in backend: {e}")
    return users

def register_user_backend(username, password, role="user"):
    """Registers a new user into the database Excel file for backend use."""
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            wb = openpyxl.Workbook()
            ws_users = wb.active
            ws_users.title = "users"
            ws_users.append(["id", "username", "password_hash", "role"])
            ws_access = wb.create_sheet("access")
            ws_access.append(["role", "allowed_tools"])
            ws_access.append(["user", "mod1,mod2,mes_pcp"]) # Default user permissions for 'user'
            ws_access.append(["admin", "all"]) # Default user permissions for 'admin'
        else:
            wb = openpyxl.load_workbook(DB_EXCEL_PATH)

        if 'users' not in wb.sheetnames:
            ws_users = wb.create_sheet("users")
            ws_users.append(["id", "username", "password_hash", "role"])
        else:
            ws_users = wb["users"]

        # Check for existing username
        for row in ws_users.iter_rows(min_row=2):
            if row[1].value == username:
                raise ValueError("Username already exists.")

        next_id = ws_users.max_row # Get the next available row number for ID
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        ws_users.append([next_id, username, password_hash, role])
        wb.save(DB_EXCEL_PATH)
        return True
    except Exception as e:
        print(f"Error registering user in backend: {e}")
        return False

def load_tools_from_excel_backend():
    """Loads tool data from the dedicated tools Excel file for backend use."""
    tools = {}
    try:
        if not os.path.exists(TOOLS_EXCEL_PATH):
            return {}

        wb = openpyxl.load_workbook(TOOLS_EXCEL_PATH)
        if 'tools' not in wb.sheetnames:
            return {}

        sheet = wb["tools"]
        for row in sheet.iter_rows(min_row=2): # Skip header row
            if len(row) >= 4 and all(cell.value is not None for cell in row[:4]):
                tools[row[0].value] = { # id as key
                    "id": row[0].value,
                    "name": row[1].value,
                    "description": row[2].value,
                    "path": row[3].value
                }
    except Exception as e:
        print(f"Error loading tools in backend: {e}")
    return tools

def load_role_permissions_backend():
    """Loads role permissions from the database Excel file for backend use."""
    perms = {}
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            return {}

        wb = openpyxl.load_workbook(DB_EXCEL_PATH)
        if 'access' not in wb.sheetnames:
            return {}

        sheet = wb["access"]
        for row in sheet.iter_rows(min_row=2): # Skip header row
            if len(row) >= 2 and row[1].value is not None:
                perms[row[0].value] = row[1].value.split(",") if row[1].value.lower() != "all" else "all"
            else:
                print(f"Warning: Skipping malformed row in 'access' sheet: {', '.join(str(c.value) for c in row)}")
    except Exception as e:
        print(f"Error loading permissions in backend: {e}")
    return perms

# --- Routes for serving HTML files ---
@app.route('/')
@app.route('/login')
def serve_login():
    """Serves the login.html file."""
    return send_from_directory(app.static_folder, 'login.html')

@app.route('/dashboard') # Changed from /dashboard.html for cleaner URLs
def serve_dashboard():
    """Serves the dashboard.html file."""
    return send_from_directory(app.static_folder, 'dashboard.html')

# --- API Endpoints ---
@app.route('/api/login', methods=['POST'])
def api_login():
    """Handles user login requests from the web frontend."""
    data = request.json
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"message": "Username and password are required."}), 400

    users = load_users_from_excel_backend()
    user = users.get(username)

    if user and bcrypt.checkpw(password.encode('utf-8'), user["password_hash"].encode('utf-8')):
        # In a real app, generate and return a JWT or session token here
        # Return user details for frontend to store (username, role)
        return jsonify({"message": "Login successful!", "token": VALID_AUTH_TOKEN, "user": {"username": user["username"], "role": user["role"]}}), 200
    else:
        return jsonify({"message": "Invalid username or password."}), 401

@app.route('/api/register', methods=['POST'])
def api_register():
    """Handles user registration requests from the web frontend."""
    data = request.json
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"message": "Username and password are required."}), 400

    users = load_users_from_excel_backend()
    if username in users:
        return jsonify({"message": "Username already exists."}), 409 # Conflict

    if register_user_backend(username, password):
        return jsonify({"message": f"User '{username}' registered successfully."}), 201 # Created
    else:
        return jsonify({"message": "An error occurred during registration."}), 500

@app.route('/api/tools', methods=['GET'])
@login_required # Protect this endpoint
def get_tools():
    """Returns tool data for the dashboard."""
    tools = load_tools_from_excel_backend()
    return jsonify(tools), 200

@app.route('/api/permissions', methods=['GET'])
@login_required # Protect this endpoint
def get_permissions():
    """Returns role permissions for the dashboard."""
    permissions = load_role_permissions_backend()
    return jsonify(permissions), 200

# --- Main entry point for running the Flask app ---
if __name__ == '__main__':
    print(f"Flask app is serving static files from: {app.static_folder}")

    # Initialize the users and tools databases if they don't exist for the first run
    if not os.path.exists(DB_EXCEL_PATH):
        print(f"Creating initial database at {DB_EXCEL_PATH}...")
        try:
            wb = openpyxl.Workbook()
            ws_users = wb.active
            ws_users.title = "users"
            ws_users.append(["id", "username", "password_hash", "role"])
            ws_users.append([1, "admin", bcrypt.hashpw("adminpass".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'), "admin"])
            ws_users.append([2, "user", bcrypt.hashpw("userpass".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'), "user"])

            ws_access = wb.create_sheet("access")
            ws_access.append(["role", "allowed_tools"])
            ws_access.append(["user", "mod1,mod2,mes_pcp"])
            ws_access.append(["admin", "all"])

            wb.save(DB_EXCEL_PATH)
            print("User database created successfully with default users and access rules.")
        except Exception as e:
            print(f"Failed to create initial user database: {e}")

    if not os.path.exists(TOOLS_EXCEL_PATH):
        print(f"Creating initial tools database at {TOOLS_EXCEL_PATH}...")
        try:
            wb = openpyxl.Workbook()
            ws_tools = wb.active
            ws_tools.title = "tools"
            ws_tools.append(["id", "name", "description", "path"])
            ws_tools.append(["mod1", "Tool One", "Description for Tool One", "path/to/tool1"])
            ws_tools.append(["mod2", "Tool Two", "Description for Tool Two", "path/to/tool2"])
            ws_tools.append(["mod3", "Tool Three", "Description for Tool Three", "path/to/tool3"])
            ws_tools.append(["mod4", "Engenharia (Workflow)", "Diagram interface for task workflows and revisions", "path/to/engenharia_tool"])
            ws_tools.append(["mes_pcp", "MES (Apontamento Fábrica)", "Manufacturing Execution System for shop floor data entry (PCP Module)", "path/to/mes_tool"])

            wb.save(TOOLS_EXCEL_PATH)
            print("Tools database created successfully with default tools.")
        except Exception as e:
            print(f"Failed to create initial tools database: {e}")


    app.run(host='0.0.0.0', port=8000, debug=True)
