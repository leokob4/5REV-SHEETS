import os
import openpyxl
import sys

# Define o caminho do diretório user_sheets
# O script assume que estará em '5REV-SHEETS/app_sheets/tools/create_engenharia_xlsx.py'
# project_root deve apontar para '5REV-SHEETS'
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
user_sheets_dir = os.path.join(project_root, 'user_sheets')
os.makedirs(user_sheets_dir, exist_ok=True) # Garante que o diretório user_sheets exista

file_path = os.path.join(user_sheets_dir, "engenharia.xlsx")
sheet_name_estrutura = "Estrutura" # Nome da planilha para a estrutura de engenharia
sheet_name_workflows = "Workflows" # Nome da planilha para os workflows de engenharia

# CORRIGIDO: Define os NOVOS cabeçalhos para a planilha 'Estrutura' em engenharia.xlsx
# Estes cabeçalhos serão a base para a criação inicial da planilha.
ENGENHARIA_STRUCTURE_HEADERS = [
    "part_number", "child_part_number", "parent_part_number", 
    "quantidade_parent_part", "quantidade_child_part", 
    "materia_prima_unidade", "materia_prima_quantidade"
]

# CORRIGIDO: Dados de exemplo para a estrutura de engenharia
# Cada linha representa um item e sua relação pai/filho e atributos.
# A interpretação aqui é que 'part_number' é o item descrito,
# 'parent_part_number' é seu pai, e 'child_part_number' é um exemplo de um de seus filhos.
sample_data_estrutura = [
    # Produto Final (PROD-001) - É um item, mas no topo da hierarquia (sem parent_part_number)
    # Lista um de seus filhos (ASSY-A) para demonstrar a coluna child_part_number
    ["PROD-001", "ASSY-A", "", 1, 1, "", ""], 
    
    # Submontagem A (ASSY-A) - É filho de PROD-001
    ["ASSY-A", "COMP-001", "PROD-001", 1, 1, "", ""], 
    
    # Componente 001 (COMP-001) - É filho de ASSY-A
    # Também tem um filho (RAW-MAT-001) e consome matéria-prima
    ["COMP-001", "RAW-MAT-001", "ASSY-A", 1, 5, "", ""],
    
    # Matéria-prima 001 (RAW-MAT-001) - É filho de COMP-001
    # Note que a "quantidade_parent_part" refere-se à quantidade do COMP-001
    ["RAW-MAT-001", "", "COMP-001", 1, 0.5, "KG", 0.5], 
    
    # Componente 002 (COMP-002) - Outro filho de ASSY-A
    ["COMP-002", "", "ASSY-A", 1, 2, "", ""],
    
    # Componente 003 (COMP-003) - Filho direto de PROD-001
    ["COMP-003", "", "PROD-001", 1, 1, "", ""],
    
    # Matéria-prima 002 (RAW-MAT-002) - Filho direto de COMP-003
    ["RAW-MAT-002", "", "COMP-003", 1, 2.5, "MTR", 2.5],
]

# Cabeçalhos fixos para a planilha de Workflows
# Estes são os cabeçalhos que a EngenhariaWorkflowTool espera e salvará
ENGENHARIA_WORKFLOW_HEADERS = ["Tipo", "ID", "X", "Y", "Largura", "Altura", "Texto", "Cor", "Conexões"]

def create_engenharia_xlsx():
    """Cria ou atualiza o arquivo engenharia.xlsx com os cabeçalhos e dados de exemplo."""
    try:
        wb = None
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            # Se for um novo workbook, remove a planilha padrão 'Sheet'
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        # --- Criação/Atualização da planilha 'Estrutura' ---
        if sheet_name_estrutura in wb.sheetnames:
            ws_estrutura = wb[sheet_name_estrutura]
            # Limpa todas as linhas para recriar com os novos cabeçalhos e dados
            ws_estrutura.delete_rows(1, ws_estrutura.max_row)
        else:
            ws_estrutura = wb.create_sheet(sheet_name_estrutura)
            print(f"Criada nova planilha '{sheet_name_estrutura}' em '{os.path.basename(file_path)}'.")

        # Adiciona os cabeçalhos para a Estrutura
        ws_estrutura.append(ENGENHARIA_STRUCTURE_HEADERS)
        # Adiciona os dados de exemplo para a Estrutura
        for row_data in sample_data_estrutura:
            ws_estrutura.append(row_data)

        # --- Criação/Atualização da planilha 'Workflows' ---
        if sheet_name_workflows in wb.sheetnames:
            ws_workflows = wb[sheet_name_workflows]
            # Limpa apenas os dados, não os cabeçalhos se já existirem no formato esperado
            # Se já houver dados, o EngenhariaWorkflowTool espera um formato específico
            # A forma mais segura é sempre garantir os cabeçalhos na primeira linha.
            ws_workflows.delete_rows(1, ws_workflows.max_row) 
        else:
            ws_workflows = wb.create_sheet(sheet_name_workflows)
            print(f"Criada nova planilha '{sheet_name_workflows}' em '{os.path.basename(file_path)}'.")
        
        # Garante que os cabeçalhos de workflow existam
        ws_workflows.append(ENGENHARIA_WORKFLOW_HEADERS)
        # A ferramenta de workflow adicionará os dados de diagrama ao salvar
        # ou carregará exemplos se a planilha estiver vazia.

        wb.save(file_path)
        print(f"Arquivo '{os.path.basename(file_path)}' criado/atualizado com as planilhas '{sheet_name_estrutura}' e '{sheet_name_workflows}'.")
    except Exception as e:
        print(f"Erro ao criar/atualizar {os.path.basename(file_path)}: {e}")

if __name__ == "__main__":
    create_engenharia_xlsx()

