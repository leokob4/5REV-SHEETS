import os
import openpyxl
import sys
from openpyxl.utils import get_column_letter

# Define os caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
# Navega de app_sheets/tools para a raiz do projeto
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))

USER_SHEETS_DIR = os.path.join(project_root, "user_sheets")
APP_SHEETS_DIR = os.path.join(project_root, "app_sheets")
DB_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "db.xlsx")

# Garante que os diretórios existam
os.makedirs(USER_SHEETS_DIR, exist_ok=True)
os.makedirs(APP_SHEETS_DIR, exist_ok=True)
# Garante que a pasta 'tools' dentro de 'app_sheets' exista, onde este script reside
os.makedirs(os.path.dirname(os.path.abspath(__file__)), exist_ok=True)


def _load_db_db_schema():
    """
    Carrega o esquema de todas as planilhas a partir da planilha 'db_db' em db.xlsx.
    Retorna: um dicionário aninhado {file_path_relative: {sheet_name: [list_of_headers]}}
    """
    schema = {}
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            print(f"Erro: db.xlsx não encontrado em {DB_EXCEL_PATH}. Não é possível carregar o esquema.")
            return schema

        wb = openpyxl.load_workbook(DB_EXCEL_PATH, data_only=True) # data_only para ler valores, não fórmulas
        if "db_db" not in wb.sheetnames:
            print("Erro: Planilha 'db_db' não encontrada em db.xlsx. Por favor, crie-a.")
            return schema

        sheet = wb["db_db"]
        # Garante que a planilha db_db tenha pelo menos uma linha (cabeçalhos)
        if sheet.max_row < 1:
            print("Erro: Planilha 'db_db' está vazia. Precisa de cabeçalhos.")
            return schema

        # Mapeia os cabeçalhos da db_db
        db_db_headers = [cell.value for cell in sheet[1]]
        db_db_header_map = {h: idx for idx, h in enumerate(db_db_headers)}

        required_db_db_headers = ["Arquivo (Caminho)", "Nome da Coluna (Cabeçalho)", "pagina_arquivo"]
        if not all(h in db_db_header_map for h in required_db_db_headers):
            print(f"Erro: Planilha 'db_db' não possui todos os cabeçalhos esperados: {', '.join(required_db_db_headers)}")
            return schema

        for row_idx in range(2, sheet.max_row + 1): # Ignora a linha de cabeçalho da db_db
            row_values = [cell.value for cell in sheet[row_idx]]
            
            # Ignora linhas completamente vazias na db_db
            if all(v is None for v in row_values):
                continue

            file_path_raw = row_values[db_db_header_map["Arquivo (Caminho)"]] if "Arquivo (Caminho)" in db_db_header_map and db_db_header_map["Arquivo (Caminho)"] < len(row_values) else None
            column_name = row_values[db_db_header_map["Nome da Coluna (Cabeçalho)"]] if "Nome da Coluna (Cabeçalho)" in db_db_header_map and db_db_header_map["Nome da Coluna (Cabeçalho)"] < len(row_values) else None
            sheet_name = row_values[db_db_header_map["pagina_arquivo"]] if "pagina_arquivo" in db_db_header_map and db_db_header_map["pagina_arquivo"] < len(row_values) else None

            if file_path_raw and column_name and sheet_name:
                normalized_file_path = os.path.normpath(file_path_raw).replace('\\', '/')
                if normalized_file_path not in schema:
                    schema[normalized_file_path] = {}
                if sheet_name not in schema[normalized_file_path]:
                    schema[normalized_file_path][sheet_name] = []
                schema[normalized_file_path][sheet_name].append(str(column_name)) # Garante que o nome da coluna é string
            else:
                print(f"Aviso: Ignorando linha incompleta em 'db_db' (linha {row_idx}): {row_values}")
    except Exception as e:
        print(f"Erro ao carregar o esquema de db_db: {e}")
    return schema

def _update_sheet_headers(workbook, sheet_name, expected_headers):
    """
    Garante que uma planilha tenha os cabeçalhos esperados, adicionando-os se ausentes.
    Preserva dados existentes.
    """
    if sheet_name not in workbook.sheetnames:
        ws = workbook.create_sheet(sheet_name)
        print(f"  Criando planilha '{sheet_name}'.")
        ws.append(expected_headers)
    else:
        ws = workbook[sheet_name]
        
        # Remove linhas completamente vazias do final da planilha para evitar problemas de max_row
        for r_idx in range(ws.max_row, 0, -1):
            if all(cell.value is None for cell in ws[r_idx]):
                ws.delete_rows(r_idx)
            else:
                break

        # Se a planilha ainda estiver vazia após a limpeza, apenas adicione os cabeçalhos
        if ws.max_row == 0:
            print(f"  Planilha '{sheet_name}' está vazia, adicionando cabeçalhos.")
            ws.append(expected_headers)
            # Ajusta a largura das colunas para os novos cabeçalhos
            for col_idx, header in enumerate(expected_headers):
                try:
                    max_length = len(str(header))
                    ws.column_dimensions[get_column_letter(col_idx + 1)].width = max_length + 2
                except Exception as e:
                    print(f"Aviso: Não foi possível ajustar a largura da coluna para '{header}' em '{sheet_name}': {e}")
            return 

        current_headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]] # Ler a primeira linha
        
        # Encontrar novos cabeçalhos para adicionar (aqueles no esquema que não estão na planilha)
        new_headers_to_add = [h for h in expected_headers if h not in current_headers]
        
        if new_headers_to_add:
            print(f"  Adicionando novos cabeçalhos à planilha '{sheet_name}': {new_headers_to_add}")
            # Encontra a próxima coluna disponível na linha 1
            next_col_idx = len(current_headers) + 1
            for new_h in new_headers_to_add:
                ws.cell(row=1, column=next_col_idx, value=new_h)
                next_col_idx += 1
        else:
            print(f"  Cabeçalhos para '{sheet_name}' estão atualizados.")

    # Reajusta a largura das colunas para todos os cabeçalhos esperados
    for header in expected_headers:
        try:
            # Encontra o índice da coluna do cabeçalho atualizado
            col_idx_in_sheet = -1
            for c_idx, cell in enumerate(ws[1]):
                if (cell.value is not None and str(cell.value) == header) or \
                   (cell.value is None and header == ""): # Para caso de cabeçalho esperado ser vazio e célula ser None
                    col_idx_in_sheet = c_idx + 1
                    break
            
            if col_idx_in_sheet != -1:
                max_length = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col_idx_in_sheet)])
                ws.column_dimensions[get_column_letter(col_idx_in_sheet)].width = max_length + 2

        except Exception as e:
            print(f"Aviso: Não foi possível ajustar a largura da coluna para '{header}' em '{sheet_name}': {e}")

def _create_or_update_all_sheets_from_schema():
    """
    Varre o esquema de db_db e cria/atualiza todas as planilhas definidas,
    garantindo que os cabeçalhos estejam corretos e preservando dados.
    """
    print("\nIniciando criação/atualização de planilhas com base no esquema db_db...")
    schema = _load_db_db_schema()

    if not schema:
        print("Erro: Esquema db_db não carregado. Nenhuma planilha será criada/atualizada.")
        return

    # Mapeia caminhos relativos do esquema para caminhos absolutos completos
    all_target_files_abs = {
        os.path.join(project_root, rel_path) for rel_path in schema.keys()
    }
    
    # Adiciona os diretórios user_sheets e app_sheets para garantir que arquivos ausentes sejam criados
    # ou que arquivos com o path correto mas não listados individualmente sejam tratados.
    # No entanto, a lógica primária é baseada no SCHEMA.
    # Para garantir que todos os arquivos *esperados pelo schema* sejam criados/atualizados,
    # itera sobre as chaves do SCHEMA.
    
    for file_rel_path_from_schema, file_schema_data in schema.items():
        file_full_path = os.path.join(project_root, file_rel_path_from_schema)
        
        # Ignora o próprio db.xlsx ao processar
        if os.path.normpath(file_full_path).replace('\\', '/') == os.path.normpath(DB_EXCEL_PATH).replace('\\', '/'):
            continue

        print(f"\nProcessando arquivo: {os.path.basename(file_full_path)}")
        
        wb = None
        if os.path.exists(file_full_path):
            try:
                wb = openpyxl.load_workbook(file_full_path)
            except Exception as e:
                print(f"Erro ao carregar '{os.path.basename(file_full_path)}': {e}. Tentando criar um novo.")
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"]) # Remove default sheet
        else:
            # Garante que o diretório para o novo arquivo exista
            os.makedirs(os.path.dirname(file_full_path), exist_ok=True)
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"]) # Remove default sheet
            print(f"  Criando novo arquivo: {os.path.basename(file_full_path)}")

        # Itera sobre as planilhas esperadas para este arquivo conforme o esquema
        for sheet_name, expected_headers in file_schema_data.items():
            _update_sheet_headers(wb, sheet_name, expected_headers)
        
        # Remove planilhas que existem no arquivo, mas não estão no esquema
        sheets_to_remove = [s for s in wb.sheetnames if s not in file_schema_data]
        for sheet_to_remove in sheets_to_remove:
            print(f"  Removendo planilha '{sheet_to_remove}' de '{os.path.basename(file_full_path)}' (não está no esquema db_db).")
            wb.remove(wb[sheet_to_remove])

        # Se o workbook ficou sem sheets, adiciona uma padrão para não dar erro ao salvar
        if not wb.sheetnames:
            wb.create_sheet("Default") 
            print(f"  Adicionada planilha 'Default' ao arquivo '{os.path.basename(file_full_path)}' pois ficou vazio após processamento.")

        try:
            wb.save(file_full_path)
            print(f"  Arquivo '{os.path.basename(file_full_path)}' salvo com sucesso.")
        except Exception as e:
            print(f"Erro ao salvar '{os.path.basename(file_full_path)}': {e}")
            
    print("\nCriação/atualização de planilhas concluída.")
    sys.exit(0) # Saída de sucesso


def _sync_db_db_with_actual_files():
    """
    Coleta os cabeçalhos de todas as planilhas .xlsx nas pastas user_sheets e app_sheets
    (exceto o próprio db.xlsx) e os registra na planilha 'db_db' em 'db.xlsx'.
    Esta ação reconstrói o dicionário de dados do sistema, que é a base para a validação de consistência.
    Preserva descrições existentes em 'descr_variavel'.
    """
    print("\nIniciando sincronização da planilha 'db_db' com os arquivos reais...")
    collected_headers_data = []
    
    # Carrega dados existentes da db_db para preservar descrições
    existing_db_db_data = {}
    try:
        if os.path.exists(DB_EXCEL_PATH):
            wb_temp = openpyxl.load_workbook(DB_EXCEL_PATH, data_only=True)
            if "db_db" in wb_temp.sheetnames:
                sheet_temp = wb_temp["db_db"]
                temp_headers = [cell.value for cell in sheet_temp[1]] if sheet_temp.max_row >= 1 else []
                
                # Check for required headers in db_db itself
                if all(h in temp_headers for h in ["Arquivo (Caminho)", "Nome da Coluna (Cabeçalho)", "pagina_arquivo", "descr_variavel"]):
                    temp_header_map = {h: idx for idx, h in enumerate(temp_headers)}
                    for row_idx in range(2, sheet_temp.max_row + 1):
                        row_values = [cell.value for cell in sheet_temp[row_idx]]
                        # Skip completely empty rows
                        if all(v is None for v in row_values):
                            continue
                        
                        file_path_raw = row_values[temp_header_map["Arquivo (Caminho)"]]
                        column_name = row_values[temp_header_map["Nome da Coluna (Cabeçalho)"]]
                        descr_variavel = row_values[temp_header_map["descr_variavel"]]
                        if file_path_raw and column_name:
                            normalized_path = os.path.normpath(file_path_raw).replace('\\', '/')
                            existing_db_db_data[(normalized_path, str(column_name))] = descr_variavel if descr_variavel is not None else ""
    except Exception as e:
        print(f"Aviso: Erro ao carregar dados existentes de db_db para preservação: {e}. As descrições podem ser perdidas.")
    
    directories_to_scan = [USER_SHEETS_DIR, APP_SHEETS_DIR]

    for base_dir in directories_to_scan:
        for root, _, files in os.walk(base_dir):
            for file_name in files:
                # Ignora arquivos temporários e o próprio db.xlsx
                if file_name.endswith(".xlsx") and not file_name.startswith('~$') and os.path.basename(file_name).lower() != "db.xlsx":
                    file_full_path = os.path.join(root, file_name)
                    
                    try:
                        wb = openpyxl.load_workbook(file_full_path, read_only=True, data_only=True)
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            
                            # Ignora planilhas vazias ou sem cabeçalhos (primeira linha totalmente vazia)
                            if sheet.max_row < 1 or sheet.max_column < 1:
                                continue 
                            
                            headers = []
                            # Itarate over the first row to get headers, handling None values
                            first_row_values = [cell.value for cell in sheet[1]]
                            if all(v is None for v in first_row_values): # If first row is completely empty, skip
                                continue

                            for cell_value in first_row_values:
                                headers.append(str(cell_value) if cell_value is not None else "") 

                            # If all collected headers are empty strings after conversion
                            if all(h == "" for h in headers): 
                                continue 

                            relative_file_path = os.path.relpath(file_full_path, project_root).replace('\\', '/')

                            for header_name in headers:
                                lookup_key = (relative_file_path, str(header_name))
                                # Pega a descrição existente ou vazia, para preservar
                                descr_variavel = existing_db_db_data.get(lookup_key, "") 
                                
                                collected_headers_data.append([
                                    relative_file_path,
                                    str(header_name),
                                    sheet_name,
                                    descr_variavel
                                ])
                        print(f"Coletado cabeçalhos de: {os.path.basename(file_full_path)}")
                    except Exception as e:
                        print(f"Erro ao processar arquivo {file_name}: {e}")

    try:
        # Abre o db.xlsx. Se não existir, cria um novo.
        if os.path.exists(DB_EXCEL_PATH):
            wb_db = openpyxl.load_workbook(DB_EXCEL_PATH)
            if "db_db" in wb_db.sheetnames:
                del wb_db["db_db"] # Remove a planilha existente para recriar
            else:
                # Se "db_db" não existe, mas "Sheet" existe e é a única, remove-a
                if "Sheet" in wb_db.sheetnames and len(wb_db.sheetnames) == 1:
                    del wb_db["Sheet"]
        else:
            wb_db = openpyxl.Workbook()
            if "Sheet" in wb_db.sheetnames: wb_db.remove(wb_db["Sheet"]) # Remove a sheet padrão se for a única
            
        ws_db_db = wb_db.create_sheet("db_db")
        db_db_headers = ["Arquivo (Caminho)", "Nome da Coluna (Cabeçalho)", "pagina_arquivo", "descr_variavel"]
        ws_db_db.append(db_db_headers)
        
        for row_data in collected_headers_data:
            ws_db_db.append(row_data)
        
        wb_db.save(DB_EXCEL_PATH)
        print(f"\nSincronização da planilha 'db_db' em '{DB_EXCEL_PATH}' concluída com sucesso.")
        sys.exit(0) # Saída de sucesso
    except Exception as e:
        print(f"Erro ao salvar metadados em db.xlsx: {e}")
        sys.exit(1) # Saída de erro

def _validate_db_consistency():
    """
    Compara a estrutura real dos cabeçalhos das planilhas do projeto
    com o esquema registrado na planilha 'db_db' em 'db.xlsx',
    identificando inconsistências e erros.
    """
    print("\nIniciando validação de consistência do banco de dados...")
    schema = _load_db_db_schema()

    if not schema:
        print("Erro: Esquema db_db não carregado. Não é possível validar a consistência.")
        sys.exit(1)

    inconsistencies_found = False
    directories_to_scan = [USER_SHEETS_DIR, APP_SHEETS_DIR]

    for base_dir in directories_to_scan:
        for root, _, files in os.walk(base_dir):
            for file_name in files:
                if file_name.endswith(".xlsx") and not file_name.startswith('~$') and os.path.basename(file_name).lower() != "db.xlsx":
                    file_full_path = os.path.join(root, file_name)
                    relative_file_path = os.path.relpath(file_full_path, project_root).replace('\\', '/')

                    expected_sheets_for_file = schema.get(relative_file_path, {}) # Use relative_file_path here
                    
                    if not os.path.exists(file_full_path):
                        if expected_sheets_for_file: # Se o esquema espera o arquivo, mas ele não existe
                            print(f"INCONSISTÊNCIA: Arquivo '{os.path.basename(file_full_path)}' esperado pelo esquema, mas não encontrado.")
                            inconsistencies_found = True
                        continue # Pula para o próximo arquivo se não existe ou não é esperado

                    try:
                        wb = openpyxl.load_workbook(file_full_path, read_only=True, data_only=True)
                        actual_sheets_in_file = set(wb.sheetnames)

                        # Verifica planilhas ausentes (no arquivo, mas esperadas no esquema)
                        for expected_sheet_name in expected_sheets_for_file.keys():
                            if expected_sheet_name not in actual_sheets_in_file:
                                print(f"INCONSISTÊNCIA: Planilha '{expected_sheet_name}' esperada em '{os.path.basename(file_full_path)}' (pelo esquema), mas não encontrada.")
                                inconsistencies_found = True
                        
                        # Verifica planilhas inesperadas (no arquivo, mas não no esquema)
                        for actual_sheet_name in actual_sheets_in_file:
                            if actual_sheet_name not in expected_sheets_for_file and actual_sheet_name != "Default": # Ignore "Default" sheet if it's auto-created by openpyxl
                                print(f"AVISO/INCONSISTÊNCIA: Planilha '{actual_sheet_name}' encontrada em '{os.path.basename(file_full_path)}', mas não definida no esquema db_db.")
                                # Não é necessariamente um erro fatal, mas um aviso importante.
                                # Se isso for um erro, setar inconsistencies_found = True

                        # Verifica cabeçalhos em cada planilha
                        for sheet_name, expected_headers in expected_sheets_for_file.items():
                            if sheet_name in actual_sheets_in_file:
                                sheet = wb[sheet_name]
                                # Considera planilhas vazias como inconsistentes se houver cabeçalhos esperados
                                if sheet.max_row < 1 or all(cell.value is None for cell in sheet[1]): 
                                    if expected_headers: # Se há cabeçalhos esperados
                                        print(f"INCONSISTÊNCIA: Planilha '{sheet_name}' em '{os.path.basename(file_full_path)}' está vazia ou sem cabeçalhos válidos, mas espera cabeçalhos.")
                                        inconsistencies_found = True
                                    continue

                                actual_headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
                                
                                # Verifica cabeçalhos ausentes na planilha real
                                for eh in expected_headers:
                                    if eh not in actual_headers:
                                        print(f"INCONSISTÊNCIA: Cabeçalho '{eh}' esperado na planilha '{sheet_name}' de '{os.path.basename(file_full_path)}' (pelo esquema), mas não encontrado.")
                                        inconsistencies_found = True
                                
                                # Verifica cabeçalhos inesperados na planilha real
                                for ah in actual_headers:
                                    if ah not in expected_headers and ah != "": # Ignora células vazias
                                        print(f"AVISO/INCONSISTÊNCIA: Cabeçalho '{ah}' encontrado na planilha '{sheet_name}' de '{os.path.basename(file_full_path)}', mas não definido no esquema db_db.")
                                        # Dependendo da sua regra, isso pode ser uma inconsistência real.
                                        # Se for um erro, setar inconsistencies_found = True

                    except Exception as e:
                        print(f"Erro ao validar arquivo {file_name}: {e}")
                        inconsistencies_found = True

    if inconsistencies_found:
        print("\nValidação de consistência concluída. Foram encontradas INCONSISTÊNCIAS.")
        sys.exit(1)
    else:
        print("\nValidação de consistência concluída. Nenhuma inconsistência encontrada.")
        sys.exit(0)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        action = sys.argv[1]
        if action == "update_db_schema":
            _sync_db_db_with_actual_files()
        elif action == "create_or_update_sheets":
            _create_or_update_all_sheets_from_schema()
        elif action == "validate":
            _validate_db_consistency()
        else:
            print(f"Ação desconhecida: {action}")
            sys.exit(1)
    else:
        print("Uso: python update_user_sheets_metadata.py <ação>")
        print("Ações disponíveis: update_db_schema, create_or_update_sheets, validate")
        sys.exit(1)

