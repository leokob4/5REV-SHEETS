import os
import sys
import openpyxl
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QMessageBox, QComboBox, QPlainTextEdit
)
from PyQt5.QtCore import Qt

# Define o caminho para a raiz do projeto de forma robusta
# Este script está em app_sheets/tools/, então '..' leva a app_sheets, e '..' novamente leva ao project_root
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))

TOOLS_EXCEL_PATH = os.path.join(project_root, "app_sheets", "tools.xlsx")

class ToolsLineGeneratorTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Linha para tools.xlsx")
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()

        # Campo para o nome da ferramenta (mod_name)
        self.mod_name_input = QLineEdit()
        self.mod_name_input.setPlaceholderText("Nome da Ferramenta (mod_name)")
        
        # Campo para a descrição da ferramenta (mod_description)
        self.mod_description_input = QLineEdit()
        self.mod_description_input.setPlaceholderText("Descrição da Ferramenta (mod_description)")
        
        # Seletor de arquivo para o caminho do módulo Python (module_path)
        file_path_layout = QHBoxLayout()
        self.module_path_input = QLineEdit()
        self.module_path_input.setPlaceholderText("Caminho do Módulo Python (ex: ui/tools/my_tool.py)")
        self.browse_file_btn = QPushButton("Procurar...")
        self.browse_file_btn.clicked.connect(self._browse_python_file)
        file_path_layout.addWidget(self.module_path_input)
        file_path_layout.addWidget(self.browse_file_btn)

        # Campo para o nome da planilha de trabalho (MOD_WORK_TABLE)
        self.mod_work_table_input = QLineEdit()
        self.mod_work_table_input.setPlaceholderText("Nome do Arquivo Excel de Trabalho (MOD_WORK_TABLE, ex: meu_data.xlsx)")
        
        # Seletor de arquivo para o caminho da planilha de trabalho (MOD_WORK_TABLE_PATH)
        work_table_path_layout = QHBoxLayout()
        self.mod_work_table_path_input = QLineEdit()
        self.mod_work_table_path_input.setPlaceholderText("Caminho da Planilha de Trabalho (MOD_WORK_TABLE_PATH, ex: /user_sheets/meu_data.xlsx)")
        self.browse_work_table_btn = QPushButton("Procurar Planilha...")
        self.browse_work_table_btn.clicked.connect(self._browse_excel_file)
        work_table_path_layout.addWidget(self.mod_work_table_path_input)
        work_table_path_layout.addWidget(self.browse_work_table_btn)

        # Campos de texto simples para comentários (QPlainTextEdit para múltiplas linhas)
        self.mod_comment_old_input = QPlainTextEdit()
        self.mod_comment_old_input.setPlaceholderText("Comentários Antigos (mod_comment_old) - Opcional")
        self.mod_comment_old_input.setFixedHeight(60) # Altura fixa para visualização

        self.mod_comment_new_input = QPlainTextEdit()
        self.mod_comment_new_input.setPlaceholderText("Novos Comentários (mod_comment_new) - Opcional")
        self.mod_comment_new_input.setFixedHeight(60) # Altura fixa para visualização

        # Botão para adicionar a ferramenta
        add_button = QPushButton("Adicionar Nova Ferramenta")
        add_button.clicked.connect(self._add_tool_to_excel)

        # Adicionar widgets ao layout principal
        layout.addWidget(QLabel("<h2>Adicionar Nova Ferramenta ao tools.xlsx</h2>"))
        layout.addWidget(QLabel("Nome da Ferramenta:"))
        layout.addWidget(self.mod_name_input)
        layout.addWidget(QLabel("Descrição da Ferramenta:"))
        layout.addWidget(self.mod_description_input)
        layout.addWidget(QLabel("Caminho do Módulo Python (.py):"))
        layout.addLayout(file_path_layout)
        layout.addWidget(QLabel("Nome do Arquivo Excel de Trabalho (opcional):"))
        layout.addWidget(self.mod_work_table_input)
        layout.addWidget(QLabel("Caminho Relativo do Arquivo Excel de Trabalho (opcional, ex: /user_sheets/):"))
        layout.addLayout(work_table_path_layout)
        layout.addWidget(QLabel("Comentários Antigos:"))
        layout.addWidget(self.mod_comment_old_input)
        layout.addWidget(QLabel("Novos Comentários:"))
        layout.addWidget(self.mod_comment_new_input)
        layout.addWidget(add_button)

        layout.addStretch() # Empurra tudo para o topo da janela

        self.setLayout(layout)

    def _browse_python_file(self):
        """Abre um diálogo para selecionar um arquivo Python e preenche o campo de input."""
        file_dialog = QFileDialog(self)
        file_path, _ = file_dialog.getOpenFileName(self, "Selecionar Arquivo Python", project_root, "Python Files (*.py)")
        if file_path:
            # Converte o caminho absoluto para um caminho relativo ao project_root
            relative_path = os.path.relpath(file_path, project_root).replace('\\', '/')
            self.module_path_input.setText(relative_path)

    def _browse_excel_file(self):
        """Abre um diálogo para selecionar um arquivo Excel e preenche os campos de caminho/nome de trabalho."""
        file_dialog = QFileDialog(self)
        # Começa a procurar na raiz do projeto ou em user_sheets se existir
        start_dir = os.path.join(project_root, "user_sheets")
        if not os.path.exists(start_dir):
            start_dir = project_root

        file_path, _ = file_dialog.getOpenFileName(self, "Selecionar Arquivo Excel", start_dir, "Excel Files (*.xlsx)")
        if file_path:
            # Converte o caminho absoluto para um caminho relativo ao project_root
            relative_path = os.path.relpath(file_path, project_root).replace('\\', '/')
            self.mod_work_table_path_input.setText(relative_path)
            self.mod_work_table_input.setText(os.path.basename(file_path)) # Preenche o nome do arquivo também

    def _get_next_mod_id(self, worksheet):
        """
        Gera o próximo ID sequencial (MOD0000XX) baseado nos IDs existentes na planilha tools.
        """
        max_id_num = 0
        mod_id_col_idx = -1
        
        # Encontra o índice da coluna 'mod_id'
        if worksheet.max_row >= 1:
            headers = [str(cell.value) if cell.value is not None else "" for cell in worksheet[1]]
            try:
                mod_id_col_idx = headers.index("mod_id")
            except ValueError:
                QMessageBox.critical(self, "Erro de Planilha", "Coluna 'mod_id' não encontrada em tools.xlsx. Não é possível gerar novo ID.")
                return None

        # Itera sobre os IDs existentes para encontrar o maior número
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=mod_id_col_idx + 1).value
            if cell_value and isinstance(cell_value, str) and cell_value.startswith("MOD"):
                try:
                    num = int(cell_value[3:]) # Extrai o número após "MOD"
                    if num > max_id_num:
                        max_id_num = num
                except ValueError:
                    continue # Ignora IDs malformados que não são números
        
        return f"MOD{max_id_num + 1:06d}" # Formata como MOD000001, MOD000002 etc.

    def _add_tool_to_excel(self):
        """
        Adiciona uma nova linha com os dados da ferramenta à planilha tools.xlsx.
        """
        mod_name = self.mod_name_input.text().strip()
        mod_description = self.mod_description_input.text().strip()
        module_path = self.module_path_input.text().strip()
        mod_work_table = self.mod_work_table_input.text().strip()
        mod_work_table_path = self.mod_work_table_path_input.text().strip()
        mod_comment_old = self.mod_comment_old_input.toPlainText().strip()
        mod_comment_new = self.mod_comment_new_input.toPlainText().strip()

        if not mod_name or not module_path:
            QMessageBox.warning(self, "Campos Obrigatórios", "Nome da Ferramenta e Caminho do Módulo Python são obrigatórios.")
            return
        
        if not os.path.exists(TOOLS_EXCEL_PATH):
            QMessageBox.critical(self, "Erro", f"Arquivo tools.xlsx não encontrado em: {TOOLS_EXCEL_PATH}. Por favor, verifique o caminho e crie o arquivo se necessário.")
            return

        try:
            wb = openpyxl.load_workbook(TOOLS_EXCEL_PATH)
            
            # Se a planilha 'tools' não existir, cria
            if "tools" not in wb.sheetnames:
                sheet = wb.create_sheet("tools")
                # Remove a planilha padrão se ela for a única e não "tools"
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and wb["Sheet"] == sheet:
                    wb.remove(wb["Sheet"])
            else:
                sheet = wb["tools"]

            # Obter cabeçalhos para garantir a ordem correta das colunas
            headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]] if sheet.max_row >= 1 else []
            header_map = {h: idx for idx, h in enumerate(headers)}

            # Garante que todos os cabeçalhos necessários para a nova entrada existam na planilha
            required_excel_headers = [
                "mod_id", "mod_name", "mod_description", "module_path",
                "MOD_WORK_TABLE", "MOD_WORK_TABLE_PATH", "mod_comment_old", "mod_comment_new"
            ]
            
            # Se a planilha estiver vazia ou com cabeçalhos incompletos, avisa o usuário.
            # O ideal é que o script de sincronização de schema já tenha preparado isso.
            if not headers or not all(h in header_map for h in required_excel_headers):
                QMessageBox.warning(self, "Aviso de Estrutura da Planilha", 
                                    "Os cabeçalhos da planilha 'tools.xlsx' estão incompletos ou a planilha está vazia. "
                                    "Por favor, execute 'Sincronizar pagina db_db com planilhas das pastas' e "
                                    "'Criar/Reinicializar/atualizar planilhas' no menu 'Ferramentas Admin' "
                                    "para corrigir a estrutura da planilha tools.xlsx antes de adicionar novas ferramentas.")
                return 

            next_mod_id = self._get_next_mod_id(sheet)
            if next_mod_id is None:
                return # Mensagem de erro já mostrada por _get_next_mod_id

            # Prepara os dados da nova linha de acordo com a ordem dos cabeçalhos da planilha
            new_row_data = [""] * len(headers)
            new_row_data[header_map["mod_id"]] = next_mod_id
            new_row_data[header_map["mod_name"]] = mod_name
            new_row_data[header_map["mod_description"]] = mod_description
            new_row_data[header_map["module_path"]] = module_path
            
            # Preenche MOD_WORK_TABLE e MOD_WORK_TABLE_PATH apenas se os cabeçalhos existirem
            if "MOD_WORK_TABLE" in header_map:
                new_row_data[header_map["MOD_WORK_TABLE"]] = mod_work_table
            if "MOD_WORK_TABLE_PATH" in header_map:
                new_row_data[header_map["MOD_WORK_TABLE_PATH"]] = mod_work_table_path
            if "mod_comment_old" in header_map:
                new_row_data[header_map["mod_comment_old"]] = mod_comment_old
            if "mod_comment_new" in header_map:
                new_row_data[header_map["mod_comment_new"]] = mod_comment_new

            sheet.append(new_row_data)
            wb.save(TOOLS_EXCEL_PATH)
            
            QMessageBox.information(self, "Sucesso", 
                                    f"Ferramenta '{mod_name}' (ID: {next_mod_id}) adicionada com sucesso a tools.xlsx.\n\n"
                                    "Lembre-se de rodar 'Sincronizar pagina db_db com planilhas das pastas' no menu Admin "
                                    "para atualizar o banco de dados principal do sistema (db.xlsx) com o novo schema.")
            
            # Limpa os campos após o sucesso
            self.mod_name_input.clear()
            self.mod_description_input.clear()
            self.module_path_input.clear()
            self.mod_work_table_input.clear()
            self.mod_work_table_path_input.clear()
            self.mod_comment_old_input.clear()
            self.mod_comment_new_input.clear()

        except FileNotFoundError:
            QMessageBox.critical(self, "Erro de Arquivo", f"Arquivo tools.xlsx não encontrado em: {TOOLS_EXCEL_PATH}. Por favor, verifique o caminho.")
        except KeyError as ke:
            QMessageBox.critical(self, "Erro de Planilha", f"Um cabeçalho esperado não foi encontrado na planilha 'tools': {ke}. Por favor, verifique a estrutura da planilha.")
        except Exception as e:
            QMessageBox.critical(self, "Erro Inesperado", f"Ocorreu um erro ao adicionar a ferramenta: {e}")

