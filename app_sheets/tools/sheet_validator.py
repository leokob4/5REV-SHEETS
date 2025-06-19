import os
import openpyxl
import sys
from PyQt5.QtWidgets import QApplication, QMessageBox # Importa QApplication e QMessageBox

# Definindo caminhos de forma dinâmica a partir da localização do script
current_dir = os.path.dirname(os.path.abspath(__file__))
# Navega de 'sheet_validator' para a raiz do projeto
project_root = os.path.dirname(current_dir)

USER_SHEETS_DIR = os.path.join(project_root, "user_sheets")
APP_SHEETS_DIR = os.path.join(project_root, "app_sheets")
# O db.xlsx está em user_sheets
DB_EXCEL_PATH = os.path.join(USER_SHEETS_DIR, "db.xlsx")

# Garante que os diretórios existam
os.makedirs(USER_SHEETS_DIR, exist_ok=True)
os.makedirs(APP_SHEETS_DIR, exist_ok=True)
# Garante que o diretório 'sheet_validator' (onde este script reside) exista
os.makedirs(os.path.dirname(os.path.abspath(__file__)), exist_ok=True)


def _load_db_db_schema():
    """
    Carrega o esquema de cabeçalhos da planilha 'db_db' em db.xlsx.
    A chave do esquema é (caminho_completo_arquivo_normalizado, nome_da_planilha).
    O nome_da_planilha é lido da coluna 'pagina_arquivo' de db_db.
    Retorna: um dicionário onde a chave é (caminho_completo_arquivo_normalizado, nome_da_planilha)
             e o valor é uma lista de cabeçalhos esperados para aquela planilha.
    """
    schema = {}
    try:
        if not os.path.exists(DB_EXCEL_PATH):
            print(f"Erro: Arquivo db.xlsx não encontrado em {DB_EXCEL_PATH}. Não é possível carregar o esquema de validação.")
            return schema

        wb = openpyxl.load_workbook(DB_EXCEL_PATH)
        if "db_db" not in wb.sheetnames:
            print("Erro: Planilha 'db_db' não encontrada em db.xlsx. Não é possível carregar o esquema de validação.")
            return schema

        sheet = wb["db_db"]
        headers_row = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
        header_map = {h: idx for idx, h in enumerate(headers_row)}

        required_headers_for_schema = ["Arquivo (Caminho)", "Nome da Coluna (Cabeçalho)", "pagina_arquivo"]
        if not all(h in header_map for h in required_headers_for_schema):
            print(f"Erro: A planilha 'db_db' não possui todos os cabeçalhos obrigatórios para o esquema de validação: {', '.join(required_headers_for_schema)}")
            return schema

        for row_idx in range(2, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            file_path_raw = row_values[header_map["Arquivo (Caminho)"]] if "Arquivo (Caminho)" in header_map and header_map["Arquivo (Caminho)"] < len(row_values) else None
            column_name = row_values[header_map["Nome da Coluna (Cabeçalho)"]] if "Nome da Coluna (Cabeçalho)" in header_map and header_map["Nome da Coluna (Cabeçalho)"] < len(row_values) else None
            
            # 'sheet_name_in_db' é o valor da coluna 'pagina_arquivo' da db_db, que representa o nome da aba/planilha.
            sheet_name_in_db = row_values[header_map["pagina_arquivo"]] if "pagina_arquivo" in header_map and header_map["pagina_arquivo"] < len(row_values) else None

            if file_path_raw and column_name and sheet_name_in_db:
                # Normaliza o caminho do arquivo completo (resolvendo o relativo para absoluto e depois normalizando barras)
                full_absolute_path = os.path.normpath(os.path.join(project_root, file_path_raw))
                key = (full_absolute_path.replace('\\', '/'), str(sheet_name_in_db)) # Usa barras normais para a chave
                
                if key not in schema:
                    schema[key] = []
                schema[key].append(str(column_name))

    except Exception as e:
        print(f"Erro ao carregar o esquema de validação de db.xlsx: {e}")
    return schema


def _validate_all_sheets():
    """
    Valida a consistência dos cabeçalhos das planilhas nas pastas user_sheets e app_sheets
    com base no esquema carregado de db.xlsx.
    """
    print("\nIniciando validação de consistência das planilhas...")
    schema = _load_db_db_schema() # Carrega o esquema
    
    if not schema:
        # A _load_db_db_schema já imprime o erro, mas podemos adicionar um QMessageBox aqui.
        QMessageBox.warning(None, "Validação Cancelada", "Não foi possível carregar o esquema de validação de db.xlsx. Verifique se 'db.xlsx' e a planilha 'db_db' estão corretos.")
        return

    validation_results = []
    
    directories_to_scan = [USER_SHEETS_DIR, APP_SHEETS_DIR]

    for base_dir in directories_to_scan:
        for root, _, files in os.walk(base_dir):
            for file_name in files:
                # Processa apenas arquivos .xlsx que não são temporários e não é o próprio db.xlsx
                if file_name.endswith(".xlsx") and not file_name.startswith('~$') and file_name.lower() != "db.xlsx":
                    file_full_path = os.path.join(root, file_name)
                    
                    try:
                        wb = openpyxl.load_workbook(file_full_path, read_only=True)
                        actual_sheet_names = wb.sheetnames

                        # Normaliza o caminho do arquivo real para comparação com o schema
                        normalized_actual_file_path = os.path.normpath(file_full_path).replace('\\', '/')

                        # Validação de sheets ausentes no arquivo real que deveriam estar no schema
                        for (schema_file_path, schema_sheet_name), expected_headers_list in schema.items():
                            if normalized_actual_file_path == schema_file_path: # Se é o mesmo arquivo
                                if schema_sheet_name not in actual_sheet_names:
                                    validation_results.append(f"AVISO: Planilha '{schema_sheet_name}' esperada em '{os.path.basename(file_full_path)}' (mapeado em db_db) está ausente no arquivo real.")

                        for sheet_name in actual_sheet_names:
                            sheet = wb[sheet_name]
                            actual_headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
                            actual_headers_set = {str(h) for h in actual_headers if h is not None} # Converte para set para comparação fácil

                            # Tenta encontrar os cabeçalhos esperados para esta combinação de arquivo/planilha no schema
                            expected_headers_list = schema.get((normalized_actual_file_path, sheet_name), [])
                            expected_headers_set = {str(h) for h in expected_headers_list if h is not None}

                            if not expected_headers_set: # Se a planilha não está registrada no schema ou o schema está vazio
                                if actual_headers_set: # Se a planilha real tem cabeçalhos mas não está no schema
                                    validation_results.append(f"INFO: Planilha '{sheet_name}' em '{os.path.basename(file_full_path)}' tem cabeçalhos (mas não está registrada no esquema db_db). Cabeçalhos encontrados: {', '.join(sorted(list(actual_headers_set)))}")
                                continue # Não há esquema para validar, então pula para a próxima sheet

                            # Compara os cabeçalhos
                            missing_headers = expected_headers_set - actual_headers_set
                            extra_headers = actual_headers_set - expected_headers_set

                            if missing_headers:
                                validation_results.append(f"ERRO: Na planilha '{sheet_name}' de '{os.path.basename(file_full_path)}', faltam os cabeçalhos: {', '.join(sorted(list(missing_headers)))}")
                            if extra_headers:
                                validation_results.append(f"AVISO: Na planilha '{sheet_name}' de '{os.path.basename(file_full_path)}', existem cabeçalhos extras: {', '.join(sorted(list(extra_headers)))}")
                            
                            # Validação da ordem dos cabeçalhos (opcional, pode ser muito rigoroso para alguns casos)
                            # if list(actual_headers_set) != sorted(list(expected_headers_headers)): # Esta é uma verificação simplificada
                            #     validation_results.append(f"AVISO: A ordem dos cabeçalhos na planilha '{sheet_name}' de '{os.path.basename(file_full_path)}' difere da ordem no esquema db_db.")

                    except Exception as e:
                        validation_results.append(f"ERRO: Não foi possível processar '{os.path.basename(file_full_path)}' (Planilha: {sheet_name if 'sheet_name' in locals() else 'N/A'}): {e}")
            
    if not validation_results:
        final_message = "✅ Validação concluída: Nenhuma inconsistência encontrada. Todas as planilhas estão consistentes com o esquema db_db."
        QMessageBox.information(None, "Validação Concluída", final_message)
        print(f"\n{final_message}")
    else:
        final_message = "Validação concluída com avisos/erros:\n" + "\n".join(validation_results)
        QMessageBox.warning(None, "Validação Concluída com Problemas", final_message)
        print(f"\n{final_message}")


if __name__ == "__main__":
    # É crucial criar uma instância QApplication antes de usar qualquer widget PyQt
    app = QApplication(sys.argv) 
    
    if len(sys.argv) > 1:
        action = sys.argv[1]
        if action == "validate":
            _validate_all_sheets()
        else:
            print(f"Ação desconhecida: {action}")
    else:
        print("Uso: python sheet_validator.py <ação>")
        print("Ações disponíveis: validate")
    
    # app.exec_() # Não é necessário chamar exec_() se o único propósito é exibir QMessageBox e sair
